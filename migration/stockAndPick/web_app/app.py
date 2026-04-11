#!/usr/bin/env python3
"""
Dockerized Flask web application for Stock and Pick PCB inventory management.
All database connections use container networking.
"""

import os
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, g, make_response
from expiration_manager import ExpirationManager, ExpirationStatus
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect, CSRFError
from wtforms import StringField, IntegerField, SelectField, SubmitField, HiddenField
from wtforms.validators import DataRequired, InputRequired, NumberRange, Length, ValidationError, Optional
import psycopg2
from psycopg2 import pool
from psycopg2.extras import RealDictCursor
import re
from functools import wraps, lru_cache
import hashlib
import secrets
import bcrypt
from flask_caching import Cache
from flask_compress import Compress
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import time

# Cache buster - changes on every app restart so browsers get fresh assets
APP_START_TIME = str(int(time.time()))
from io import BytesIO

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
# Use environment variable for secret key, fallback to a consistent key
# IMPORTANT: In production, always set SECRET_KEY environment variable
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'kosh-pcb-inventory-secret-key-2025-production-v1')

# Session Configuration
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=14)  # 14 hour session timeout
app.config['SESSION_COOKIE_SECURE'] = bool(os.environ.get('VERCEL')) or os.environ.get('SESSION_COOKIE_SECURE', 'false').lower() == 'true'
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Prevent JavaScript access to session cookie
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF protection
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max upload, matches nginx

# Enable Flask-Caching for performance
app.config['CACHE_TYPE'] = 'simple'  # In-memory cache
app.config['CACHE_DEFAULT_TIMEOUT'] = 300  # 5 minutes default
cache = Cache(app)

# Enable gzip compression for all responses
app.config['COMPRESS_MIMETYPES'] = [
    'text/html', 'text/css', 'text/xml', 'application/json',
    'application/javascript', 'text/javascript'
]
app.config['COMPRESS_LEVEL'] = 6  # Balance between compression and speed
app.config['COMPRESS_MIN_SIZE'] = 500  # Only compress responses > 500 bytes
compress = Compress(app)

# Column definitions for shortage/job report exports
SHORTAGE_EXPORT_COLUMNS = [
    {'key': 'line_no',      'label': 'LINE #',       'width': 8,  'default': False},
    {'key': 'aci_pn',       'label': 'ACI PN',       'width': 15, 'default': True},
    {'key': 'pcn',          'label': 'PCN',           'width': 12, 'default': True},
    {'key': 'mpn',          'label': 'MPN',           'width': 25, 'default': True},
    {'key': 'manufacturer', 'label': 'MANUFACTURER',  'width': 20, 'default': False},
    {'key': 'description',  'label': 'DESCRIPTION',   'width': 30, 'default': False},
    {'key': 'qty',          'label': 'QTY',           'width': 8,  'default': True},
    {'key': 'order_qty',    'label': 'ORDER QTY',     'width': 12, 'default': True},
    {'key': 'req',          'label': 'REQ',           'width': 8,  'default': True},
    {'key': 'item',         'label': 'ITEM',          'width': 15, 'default': True},
    {'key': 'qty_on_hand',  'label': 'ON HAND QTY',   'width': 14, 'default': True},
    {'key': 'location',     'label': 'LOCATION',      'width': 15, 'default': True},
    {'key': 'unit_cost',    'label': 'UNIT COST',     'width': 12, 'default': False},
    {'key': 'line_cost',    'label': 'LINE COST',     'width': 12, 'default': False},
]

def get_export_cell_value(item, column_key, order_qty=None):
    """Extract cell value for a report column from a DB row."""
    mapping = {
        'line_no':      lambda i: i.get('line_no', ''),
        'aci_pn':       lambda i: i.get('aci_pn', ''),
        'pcn':          lambda i: i.get('pcn') or '',
        'mpn':          lambda i: i.get('mpn', ''),
        'manufacturer': lambda i: i.get('manufacturer') or '',
        'description':  lambda i: i.get('description') or '',
        'qty':          lambda i: i.get('qty', 0),
        'order_qty':    lambda i: i.get('order_qty') or order_qty or '',
        'req':          lambda i: i.get('req') or (int(i.get('qty', 0) or 0) * (order_qty or 1)),
        'item':         lambda i: i.get('item') or i.get('aci_pn') or '',
        'qty_on_hand':  lambda i: i.get('qty_on_hand') if i.get('qty_on_hand') is not None else i.get('on_hand', 0),
        'location':     lambda i: i.get('location') or '',
        'unit_cost':    lambda i: i.get('unit_cost') or '',
        'line_cost':    lambda i: i.get('line_cost') or '',
    }
    extractor = mapping.get(column_key, lambda i: '')
    return extractor(item)

# CSRF Configuration
# Enable CSRF protection with proper configuration
app.config['WTF_CSRF_ENABLED'] = True
app.config['WTF_CSRF_TIME_LIMIT'] = None  # No time limit on CSRF tokens
app.config['WTF_CSRF_SSL_STRICT'] = False  # Allow non-HTTPS in development
app.config['WTF_CSRF_CHECK_DEFAULT'] = True
app.config['WTF_CSRF_HEADERS'] = ['X-CSRFToken', 'X-CSRF-Token']  # Accept CSRF token from headers
csrf = CSRFProtect(app)

# Enable rate limiting (protects against brute force attacks)
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per hour", "50 per minute"],
    storage_uri="memory://",
    strategy="fixed-window"
)

# Input validation functions
def validate_job_number(job: str) -> bool:
    """Validate job number format."""
    if not job or len(job) > 50:
        return False
    # Allow alphanumeric characters, dashes, underscores
    return re.match(r'^[a-zA-Z0-9_-]+$', job) is not None

def validate_pcb_type(pcb_type: str) -> bool:
    """Validate PCB type against allowed values."""
    allowed_types = ['Bare', 'Partial', 'Completed', 'Ready to Ship']
    return pcb_type in allowed_types

def validate_quantity(quantity: Any) -> tuple[bool, int]:
    """Validate quantity is a positive integer."""
    try:
        qty = int(quantity)
        return (1 <= qty <= 10000, qty)
    except (ValueError, TypeError):
        return (False, 0)

def validate_location(location: str) -> bool:
    """Validate location format - must be exactly 7 digits or a standard text location."""
    if not location:
        return False
    location = location.strip()
    standard_locations = ['Receiving Area', 'Rec Area', 'Count Area', 'Stock Room', 'MFG Floor']
    if location.lower() in [loc.lower() for loc in standard_locations]:
        return True
    # Must be exactly 7 digits
    return bool(re.match(r'^\d{7}$', location))

def validate_api_request(required_fields: list):
    """Decorator to validate API request data."""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            try:
                data = request.get_json()
                if not data:
                    return jsonify({'success': False, 'error': 'No JSON data provided'}), 400
                
                # Check required fields
                for field in required_fields:
                    if field not in data:
                        return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400
                
                # Validate specific fields
                if 'job' in data and not validate_job_number(data['job']):
                    return jsonify({'success': False, 'error': 'Invalid job number format'}), 400
                
                if 'pcb_type' in data and not validate_pcb_type(data['pcb_type']):
                    return jsonify({'success': False, 'error': 'Invalid PCB type'}), 400
                
                if 'quantity' in data:
                    is_valid, qty = validate_quantity(data['quantity'])
                    if not is_valid:
                        return jsonify({'success': False, 'error': 'Invalid quantity (must be 1-10000)'}), 400
                    data['quantity'] = qty
                
                if 'location' in data and not validate_location(data['location']):
                    return jsonify({'success': False, 'error': 'Invalid location format'}), 400
                
                return f(*args, **kwargs)
            except Exception as e:
                logger.error(f"API validation error: {e}")
                return jsonify({'success': False, 'error': 'Request validation failed'}), 400
        return decorated_function
    return decorator

# Secure error handling
def get_safe_error_message(error: Exception, operation: str = "operation") -> str:
    """Return a safe error message that doesn't expose sensitive information."""
    # Log the full error for debugging
    logger.error(f"Error in {operation}: {str(error)}", exc_info=True)
    
    # Return generic error messages to users
    if isinstance(error, psycopg2.Error):
        return f"Database {operation} failed. Please try again."
    elif isinstance(error, ValueError):
        return f"Invalid data provided for {operation}."
    elif isinstance(error, KeyError):
        return f"Missing required information for {operation}."
    else:
        return f"An error occurred during {operation}. Please try again."

# CORS: Allow Vercel frontend and tunnel URLs to call local backend
@app.after_request
def add_cors_headers(response):
    """Add CORS headers for Vercel frontend accessing local backend via tunnel."""
    origin = request.headers.get('Origin', '')
    allowed_origins = [
        'https://aci-kosh.vercel.app',
        'https://aci-forge.vercel.app',
        'https://aci-nexus.vercel.app',
    ]
    # Also allow any trycloudflare.com origin (tunnel URLs)
    if origin in allowed_origins or origin.endswith('.trycloudflare.com'):
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, X-CSRFToken, X-CSRF-Token'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    if request.method == 'OPTIONS':
        response.status_code = 200
    return response

# Security headers and performance optimization
@app.after_request
def add_security_headers(response):
    """Add comprehensive security headers and caching to all responses."""
    # Content Security Policy
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' cdn.jsdelivr.net cdnjs.cloudflare.com 'unsafe-inline'; "
        "style-src 'self' cdn.jsdelivr.net fonts.googleapis.com 'unsafe-inline'; "
        "font-src 'self' cdn.jsdelivr.net fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self' cdn.jsdelivr.net *.trycloudflare.com *.americancircuits.net"
    )
    # HTTP Strict Transport Security (force HTTPS in production)
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    # Prevent clickjacking - allow same origin for print preview iframes
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    # Prevent MIME type sniffing
    response.headers['X-Content-Type-Options'] = 'nosniff'
    # XSS Protection
    response.headers['X-XSS-Protection'] = '1; mode=block'
    # Referrer Policy
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    # Feature Policy
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'

    # Cache static assets (CSS, JS, fonts, images) for performance
    # but disable caching for HTML pages to ensure fresh content
    if response.content_type and any(
        t in response.content_type for t in ['text/css', 'javascript', 'font', 'image/', 'woff', 'svg']
    ):
        response.headers['Cache-Control'] = 'public, max-age=3600'
    elif 'no-store' not in response.headers.get('Cache-Control', ''):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'

    return response

@app.context_processor
def inject_current_time():
    return {
        'current_time': datetime.now().strftime('%B %d, %Y %I:%M %p'),
        'current_year': datetime.now().year,
        'current_user': g.get('current_user', {}),
        'user_can_see_itar': g.get('user_can_see_itar', False),
        'is_admin': is_admin_user(),
        'can_manage': can_manage_parts(),
        'cache_version': APP_START_TIME
    }

@app.template_filter('moment_fromnow')
def moment_fromnow_filter(dt):
    """Calculate time ago from a datetime object"""
    if not dt:
        return "Unknown"

    # Handle string timestamps from database
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
        except (ValueError, TypeError) as e:
            logger.warning(f"Failed to parse timestamp: {dt}, error: {e}")
            return "Unknown"

    now = datetime.now()
    if dt.tzinfo is not None:
        # Convert to naive datetime for comparison
        dt = dt.replace(tzinfo=None)

    diff = now - dt
    seconds = diff.total_seconds()

    if seconds < 60:
        return "just now"
    elif seconds < 3600:
        minutes = int(seconds / 60)
        return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
    elif seconds < 86400:
        hours = int(seconds / 3600)
        return f"{hours} hour{'s' if hours != 1 else ''} ago"
    else:
        days = int(seconds / 86400)
        return f"{days} day{'s' if days != 1 else ''} ago"

@app.template_filter('expiration_status')
def expiration_status_filter(item):
    """Calculate expiration status for an inventory item"""
    dc = item.get('dc')
    pcb_type = item.get('pcb_type', 'Bare')
    msd = item.get('msd')
    return expiration_manager.calculate_expiration_status(dc, pcb_type, msd)

@app.template_filter('expiration_badge_class')
def expiration_badge_class_filter(status_text):
    """Get Bootstrap badge class for expiration status"""
    try:
        status = ExpirationStatus(status_text)
        return expiration_manager.get_expiration_badge_class(status)
    except ValueError:
        return 'bg-secondary'

@app.template_filter('expiration_icon')
def expiration_icon_filter(status_text):
    """Get Bootstrap icon for expiration status"""
    try:
        status = ExpirationStatus(status_text)
        return expiration_manager.get_expiration_icon(status)
    except ValueError:
        return 'bi-question-circle'

@app.template_filter('expiration_display')
def expiration_display_filter(expiration_info):
    """Format expiration information for display"""
    return expiration_manager.format_expiration_display(expiration_info)

@app.template_filter('format_number')
def format_number_filter(value):
    """Format number with thousands separator"""
    try:
        return f"{int(value):,}"
    except (ValueError, TypeError):
        return value

# Database configuration from environment variables (local PostgreSQL)
DB_CONFIG = {
    'host': os.getenv('POSTGRES_HOST', 'aci-database'),
    'port': int(os.getenv('POSTGRES_PORT', 5432)),
    'database': os.getenv('POSTGRES_DB', 'pcb_inventory'),
    'user': os.getenv('POSTGRES_USER', 'stockpick_user'),
    'password': os.getenv('POSTGRES_PASSWORD', 'stockpick_pass')
}
logger.info("Using local database")

# PCB Types and Locations (matching the original application)
PCB_TYPES = [
    ('Bare', 'Bare PCB'),
    ('Partial', 'Partial Assembly'),
    ('Completed', 'Completed Assembly'),
    ('Ready to Ship', 'Ready to Ship')
]

# ITAR Classifications
ITAR_CLASSIFICATIONS = [
    ('NONE', 'Non-ITAR (Public)'),
    ('EAR99', 'Export Administration Regulations'),
    ('SENSITIVE', 'Company Sensitive'),
    ('ITAR', 'ITAR Controlled')
]

# User Roles
USER_ROLES = [
    ('Super User', 'Super User'),
    ('User', 'User'),
    ('Manager', 'Manager'),
    ('Operator', 'Operator'),
    ('ITAR', 'ITAR')
]

# Users authorized to access admin pages (user management, notifications)
# in addition to users with the 'Admin' role
ADMIN_AUTHORIZED_USERS = {'kanav', 'preet'}

def is_admin_user():
    """Check if current session user has admin-level access (by role or authorized list)."""
    if session.get('role') == 'Admin':
        return True
    return session.get('username', '').lower() in ADMIN_AUTHORIZED_USERS

# Users allowed to access ACI Numbers and Locations (in addition to admins)
MANAGE_AUTHORIZED_USERS = {'parts@americancircuits.com'}

def can_manage_parts():
    """Check if user can access ACI Numbers and Locations (admins + Theresa)."""
    if is_admin_user():
        return True
    return session.get('username', '').lower() in MANAGE_AUTHORIZED_USERS

LOCATION_RANGES = [
    ('1000-1999', '1000-1999'),
    ('2000-2999', '2000-2999'),
    ('3000-3999', '3000-3999'),
    ('4000-4999', '4000-4999'),
    ('5000-5999', '5000-5999'),
    ('6000-6999', '6000-6999'),
    ('7000-7999', '7000-7999'),
    ('8000-8999', '8000-8999'),  # Default in original app
    ('9000-9999', '9000-9999'),
    ('10000-10999', '10000-10999')
]

def validate_pcb_type_field(form, field):
    """Custom validator for PCB type field."""
    allowed_types = ['Bare', 'Partial', 'Completed', 'Ready to Ship']
    if field.data not in allowed_types:
        raise ValidationError(f'Component type must be one of: {", ".join(allowed_types)}')

STANDARD_LOCATIONS = ['Receiving Area', 'Rec Area', 'Count Area', 'Stock Room', 'MFG Floor']

def validate_location_field(form, field):
    """Custom validator: location must be exactly 7 digits or a standard text location."""
    if not field.data or not field.data.strip():
        return  # Let Optional/DataRequired handle empty
    location = field.data.strip()
    if location.lower() in [loc.lower() for loc in STANDARD_LOCATIONS]:
        return
    if not re.match(r'^\d{7}$', location):
        raise ValidationError('Location must be exactly 7 digits (e.g. 1101101) or a standard location.')

class StockForm(FlaskForm):
    """Form for stocking electronic parts."""
    pcn_number = StringField('PCN Number', validators=[Length(max=10)])
    job = StringField('Job Number (Item)', validators=[Length(max=50)])  # Optional - will use part_number if not provided
    mpn = StringField('MPN (Manufacturing Part Number)', validators=[Length(max=50)])
    part_number = StringField('Part Number', validators=[DataRequired(), Length(min=1, max=50)])  # Now required - serves as job identifier
    po = StringField('PO (Purchase Order)', validators=[Length(max=50)])
    work_order = StringField('Work Order Number', validators=[Length(max=50)])
    pcb_type = StringField('Component Type', validators=[Length(max=50)], default='Bare')
    dc = StringField('Date Code (DC)', validators=[Length(max=50)])
    msd = StringField('Moisture Sensitive Device (MSD)', validators=[Length(max=50)])
    quantity = IntegerField('Quantity', validators=[DataRequired(), NumberRange(min=1)])
    location_from = StringField('Location From', validators=[DataRequired(), Length(min=1, max=50), validate_location_field], default='Receiving Area')
    location_to = StringField('Location To', validators=[DataRequired(), Length(min=1, max=50), validate_location_field])
    submit = SubmitField('Stock Parts')

class PickForm(FlaskForm):
    """Form for picking electronic parts."""
    pcn = IntegerField('PCN Number', validators=[Optional(), NumberRange(min=0)])  # Optional - when specified, pick from that specific PCN only
    job = StringField('Job Number (Item)', validators=[Length(max=50)])  # Optional - will use part_number if not provided
    mpn = StringField('MPN (Manufacturing Part Number)', validators=[Length(max=50)])
    part_number = StringField('Part Number', validators=[DataRequired(), Length(min=1, max=50)])  # Now required - serves as job identifier
    po = StringField('Job Number', validators=[Length(max=50)])
    work_order = StringField('Work Order Number', validators=[Length(max=50)])
    pcb_type = StringField('Component Type', validators=[Length(max=50)], default='Bare')
    dc = StringField('Date Code (DC)', validators=[Length(max=50)])
    msd = StringField('Moisture Sensitive Device (MSD)', validators=[Length(max=50)])
    quantity = IntegerField('Quantity', validators=[InputRequired(), NumberRange(min=0)])
    submit = SubmitField('Pick Parts')

class RestockForm(FlaskForm):
    """Form for restocking parts from Count Area to specified location."""
    pcn = StringField('PCN Number', validators=[Optional(), Length(max=50)])
    item = StringField('Item Number', validators=[Optional(), Length(max=50)])
    po = StringField('PO Number', validators=[Optional(), Length(max=50)])
    quantity = IntegerField('Quantity to Restock', validators=[DataRequired(), NumberRange(min=1)])
    location_from = StringField('Source Location', validators=[Optional(), Length(max=50), validate_location_field], default='Count Area')
    location_to = StringField('Destination Location (Optional)', validators=[Optional(), Length(max=50), validate_location_field])
    submit = SubmitField('Restock Parts')

    def validate(self, extra_validators=None):
        """Custom validation to ensure either PCN or Item is provided."""
        if not super().validate(extra_validators):
            return False

        if not self.pcn.data and not self.item.data:
            self.pcn.errors.append('Either PCN or Item Number is required')
            self.item.errors.append('Either PCN or Item Number is required')
            return False

        return True

# User authentication now handled by ACI Dashboard

_IS_VERCEL = bool(os.environ.get('VERCEL'))

class DatabaseManager:
    """Handle database operations using PostgreSQL with connection pooling and failover."""

    def __init__(self):
        self.db_config = DB_CONFIG
        self.pool = None

        if _IS_VERCEL:
            # Serverless: skip pool creation (avoids 2 SSL handshakes on cold start)
            logger.info("Serverless mode: using on-demand DB connections")
        else:
            # Long-running server (Docker): use connection pool
            self._init_pool()

    def _init_pool(self):
        try:
            self.pool = psycopg2.pool.ThreadedConnectionPool(
                minconn=2,
                maxconn=15,
                **self.db_config
            )
            logger.info(f"Database connection pool initialized (primary)")
        except Exception as e:
            logger.error(f"Failed to create connection pool: {e}")
            raise

    def get_connection(self):
        """Get a database connection from the pool or create one directly (serverless)."""
        try:
            if self.pool:
                return self.pool.getconn()
            return psycopg2.connect(**self.db_config)
        except Exception as e:
            logger.error(f"Failed to get connection: {e}")
            raise

    def return_connection(self, conn):
        """Return a connection to the pool or close it (serverless)."""
        try:
            if self.pool:
                self.pool.putconn(conn)
            else:
                conn.close()
        except Exception as e:
            logger.error(f"Failed to return/close connection: {e}")

    def get_pool_stats(self):
        """Get connection pool statistics for monitoring."""
        if not self.pool:
            return {'mode': 'serverless', 'pool': False}
        try:
            return {
                'minconn': self.pool.minconn,
                'maxconn': self.pool.maxconn,
                'closed': self.pool.closed
            }
        except Exception as e:
            logger.error(f"Failed to get pool stats: {e}")
            return {
                'error': str(e)
            }

    def execute_function(self, function_name: str, params: tuple) -> Dict[str, Any]:
        """Execute a PostgreSQL function and return the result."""
        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Build function call with proper parameter count
                param_placeholders = ', '.join(['%s'] * len(params))
                sql = f"SELECT {function_name}({param_placeholders})"
                cur.execute(sql, params)
                result = cur.fetchone()
                conn.commit()
                return dict(result[function_name.split('.')[-1]])
        except Exception as e:
            if conn:
                conn.rollback()
            error_msg = get_safe_error_message(e, "database function")
            return {'success': False, 'error': error_msg}
        finally:
            if conn:
                self.return_connection(conn)

    def validate_location(self, location: str) -> bool:
        """Check if a location exists in tblLoc table or is a valid text location.
        Location must be exactly 7 digits or a standard text location."""
        if not location or location.strip() == '':
            return False

        location = location.strip()

        # Allow standard text locations (these are used throughout the system)
        standard_locations = [
            'Receiving Area', 'Rec Area',
            'Count Area',
            'Stock Room',
            'MFG Floor'
        ]

        # Case-insensitive check for standard locations
        if location.lower() in [loc.lower() for loc in standard_locations]:
            return True

        # Must be exactly 7 digits
        if not re.match(r'^\d{7}$', location):
            return False

        # Check if location exists in tblLoc
        conn = None
        try:
            conn = self.get_connection()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT COUNT(*) FROM pcb_inventory."tblLoc"
                WHERE location::text = %s
            """, (location,))

            count = cursor.fetchone()[0]
            return count > 0
        except Exception as e:
            logger.error(f"Error validating location: {e}")
            return False
        finally:
            if conn:
                cursor.close()
                self.return_connection(conn)

    def stock_pcb(self, job: str, pcb_type: str, quantity: int, location_from: str, location_to: str,
                  itar_classification: str = 'NONE', user_role: str = 'USER',
                  itar_auth: bool = False, username: str = 'system', work_order: str = None,
                  dc: str = None, msd: str = None, pcn: int = None, mpn: str = None,
                  part_number: str = None) -> Dict[str, Any]:
        """Stock PCB - directly updates warehouse inventory and transaction tables."""
        # CRITICAL: Input validation
        if not isinstance(quantity, int) or quantity < 1 or quantity > 10000:
            return {
                'success': False,
                'error': f'Invalid quantity: {quantity}. Must be between 1 and 10,000.'
            }

        if not job or not isinstance(job, str) or len(job) > 50:
            return {
                'success': False,
                'error': 'Invalid job identifier. Must be 1-50 characters.'
            }

        if not pcn:
            return {
                'success': False,
                'error': 'PCN is required for stock operation'
            }

        try:
            pcn_int = int(pcn)
            if pcn_int < 0 or pcn_int > 99999:
                return {
                    'success': False,
                    'error': f'Invalid PCN: {pcn}. Must be between 1 and 99999.'
                }
        except (ValueError, TypeError):
            return {
                'success': False,
                'error': f'Invalid PCN format: {pcn}. Must be numeric.'
            }

        # Validate locations exist
        if not self.validate_location(location_to):
            return {
                'success': False,
                'error': f'Location "{location_to}" does not exist. Please verify the location code and try again.'
            }

        conn = None
        cursor = None
        try:
            conn = self.get_connection()
            # CRITICAL: Set SERIALIZABLE isolation to prevent race conditions
            conn.autocommit = False
            cursor = conn.cursor()

            # CRITICAL: Lock row with FOR UPDATE to prevent concurrent stock operations
            # Check if PCN and item combination exists in warehouse
            cursor.execute("""
                SELECT id, onhandqty, mpn, dc, msd, po, loc_to, loc_from
                FROM pcb_inventory."tblWhse_Inventory"
                WHERE item::text ILIKE %s AND pcn::text = %s
                FOR UPDATE  -- Lock row to prevent race conditions
                LIMIT 1
            """, (job, str(pcn)))

            existing = cursor.fetchone()

            if existing:
                # Update existing record - add to existing quantity and update location
                cursor.execute("""
                    UPDATE pcb_inventory."tblWhse_Inventory"
                    SET onhandqty = COALESCE(onhandqty, 0) + %s,
                        loc_to = %s,
                        loc_from = %s,
                        dc = COALESCE(%s, dc),
                        msd = COALESCE(%s, msd),
                        po = COALESCE(%s, po),
                        mpn = COALESCE(%s, mpn),
                        migrated_at = CURRENT_TIMESTAMP
                    WHERE item::text ILIKE %s AND pcn::text = %s
                """, (quantity, location_to, location_from, dc, msd, work_order, mpn, job, str(pcn)))
                logger.info(f"Updated warehouse inventory for item {job}, PCN {pcn} - added {quantity} units (moved from {location_from} to {location_to})")
            else:
                # PCN doesn't exist in warehouse yet - insert new record
                cursor.execute("""
                    INSERT INTO pcb_inventory."tblWhse_Inventory"
                    (item, pcn, mpn, dc, onhandqty, loc_from, loc_to, msd, po, migrated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
                """, (job, str(pcn), mpn or '', dc, quantity, location_from, location_to, msd, work_order))
                logger.info(f"Inserted new warehouse inventory for item {job}, PCN {pcn} at {location_to}")

            # Record the stock transaction in tblTransaction
            cursor.execute("""
                INSERT INTO pcb_inventory."tblTransaction"
                (trantype, item, pcn, mpn, dc, msd, tranqty, tran_time, loc_from, loc_to, wo, po, userid)
                VALUES ('STOCK', %s, %s, %s, %s, %s, %s, TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', 'MM/DD/YY HH24:MI:SS'), %s, %s, %s, %s, %s)
            """, (job, str(pcn) if pcn else None, mpn, dc, msd, quantity, location_from, location_to, work_order, work_order, username))

            # Get the updated quantity before commit (within same transaction)
            cursor.execute("""
                SELECT COALESCE(SUM(onhandqty), 0) as total_qty
                FROM pcb_inventory."tblWhse_Inventory"
                WHERE item::text ILIKE %s
            """, (job,))
            total_result = cursor.fetchone()
            new_qty = int(total_result[0]) if total_result and total_result[0] else 0

            conn.commit()
            logger.info(f"Stock operation completed: {quantity} units of {job} (PCN: {pcn}) moved from {location_from} to {location_to}")

            # Clear cache after successful update
            cache.delete_memoized(self.get_current_inventory)
            cache.delete('stats_summary')

            return {
                'success': True,
                'message': f'Successfully stocked {quantity} units from {location_from} to {location_to}',
                'pcn': pcn,
                'stocked_qty': quantity,
                'new_qty': new_qty,
                'job': job,
                'location_from': location_from,
                'location_to': location_to
            }

        except psycopg2.extensions.TransactionRollbackError as e:
            # Serialization failure - concurrent transaction conflict
            if conn:
                conn.rollback()
            logger.warning(f"Stock operation serialization conflict: {e}")
            return {
                'success': False,
                'error': 'This item is being accessed by another user. Please try again.'
            }
        except psycopg2.IntegrityError as e:
            if conn:
                conn.rollback()
            logger.error(f"Stock operation integrity error: {e}")
            return {
                'success': False,
                'error': 'Data integrity error. Please verify your inputs.'
            }
        except Exception as e:
            if conn:
                conn.rollback()
            logger.error(f"Stock operation failed: {e}", exc_info=True)
            error_msg = get_safe_error_message(e, "stock operation")
            return {'success': False, 'error': error_msg}
        finally:
            if cursor:
                try:
                    cursor.close()
                except Exception as e:
                    logger.error(f"Error closing cursor: {e}")
            if conn:
                try:
                    self.return_connection(conn)
                except Exception as e:
                    logger.error(f"Error returning connection: {e}")
    
    def pick_pcb(self, job: str, pcb_type: str, quantity: int,
                 user_role: str = 'USER', itar_auth: bool = False, username: str = 'system', work_order: str = None, pcn: int = None) -> Dict[str, Any]:
        """Pick PCB from warehouse inventory.
        If pcn is provided, picks from that specific PCN only.
        Otherwise, picks using FIFO across all PCNs for the item.
        """
        # CRITICAL: Input validation to prevent invalid data
        if not isinstance(quantity, int) or quantity < 0 or quantity > 10000:
            return {
                'success': False,
                'error': f'Invalid quantity: {quantity}. Must be between 0 and 10,000.'
            }

        if not job or not isinstance(job, str) or len(job) > 50:
            return {
                'success': False,
                'error': 'Invalid job identifier. Must be 1-50 characters.'
            }

        if pcn is not None:
            try:
                pcn_int = int(pcn)
                if pcn_int < 0 or pcn_int > 99999:
                    return {
                        'success': False,
                        'error': f'Invalid PCN: {pcn}. Must be between 1 and 99999.'
                    }
            except (ValueError, TypeError):
                return {
                    'success': False,
                    'error': f'Invalid PCN format: {pcn}. Must be numeric.'
                }

        conn = None
        cursor = None
        try:
            conn = self.get_connection()
            # CRITICAL: Set SERIALIZABLE isolation to prevent race conditions
            conn.autocommit = False
            cursor = conn.cursor()

            try:
                # PURGE OPERATION: When quantity is 0, delete zero-qty records from inventory
                if quantity == 0:
                    if pcn:
                        # Check that the PCN exists and has zero on-hand qty
                        cursor.execute("""
                            SELECT pcn, item, onhandqty
                            FROM pcb_inventory."tblWhse_Inventory"
                            WHERE pcn::text = %s AND item::text ILIKE %s
                            FOR UPDATE
                        """, (str(pcn), job))
                        row = cursor.fetchone()
                        if not row:
                            conn.rollback()
                            return {
                                'success': False,
                                'error': f'PCN {pcn} not found for item {job} in warehouse inventory.',
                                'job': job, 'pcb_type': pcb_type
                            }
                        if row[2] and int(row[2]) > 0:
                            conn.rollback()
                            return {
                                'success': False,
                                'error': f'Cannot purge PCN {pcn} — on-hand quantity is {int(row[2])}. Pick the remaining units first or edit quantity to 0.',
                                'job': job, 'pcb_type': pcb_type
                            }
                        # Delete the zero-qty warehouse inventory record
                        cursor.execute("""
                            DELETE FROM pcb_inventory."tblWhse_Inventory"
                            WHERE pcn::text = %s AND item::text ILIKE %s
                        """, (str(pcn), job))
                        deleted_count = cursor.rowcount
                        # Record a PURGE transaction
                        cursor.execute("""
                            INSERT INTO pcb_inventory."tblTransaction"
                            (trantype, item, pcn, tranqty, tran_time, loc_from, loc_to, wo, userid)
                            VALUES ('PURGE', %s, %s, 0,
                                    TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', 'MM/DD/YY HH24:MI:SS'),
                                    'Warehouse', 'Purged', %s, %s)
                        """, (job, str(pcn), work_order or '', username))
                    else:
                        # Purge all zero-qty records for this item
                        cursor.execute("""
                            SELECT COUNT(*) FROM pcb_inventory."tblWhse_Inventory"
                            WHERE item::text ILIKE %s AND onhandqty = 0
                            FOR UPDATE
                        """, (job,))
                        zero_count = cursor.fetchone()[0]
                        if zero_count == 0:
                            # Check if item exists at all
                            cursor.execute("""
                                SELECT COUNT(*) FROM pcb_inventory."tblWhse_Inventory"
                                WHERE item::text ILIKE %s
                            """, (job,))
                            total = cursor.fetchone()[0]
                            conn.rollback()
                            if total > 0:
                                return {
                                    'success': False,
                                    'error': f'No zero-quantity records found for {job}. All records have on-hand quantity > 0.',
                                    'job': job, 'pcb_type': pcb_type
                                }
                            else:
                                return {
                                    'success': False,
                                    'error': f'Item {job} not found in warehouse inventory.',
                                    'job': job, 'pcb_type': pcb_type
                                }
                        # Get PCNs being purged for transaction logging
                        cursor.execute("""
                            SELECT pcn FROM pcb_inventory."tblWhse_Inventory"
                            WHERE item::text ILIKE %s AND onhandqty = 0
                        """, (job,))
                        purged_pcns = [str(r[0]) for r in cursor.fetchall()]
                        # Delete zero-qty records
                        cursor.execute("""
                            DELETE FROM pcb_inventory."tblWhse_Inventory"
                            WHERE item::text ILIKE %s AND onhandqty = 0
                        """, (job,))
                        deleted_count = cursor.rowcount
                        # Record PURGE transactions
                        for p in purged_pcns:
                            cursor.execute("""
                                INSERT INTO pcb_inventory."tblTransaction"
                                (trantype, item, pcn, tranqty, tran_time, loc_from, loc_to, wo, userid)
                                VALUES ('PURGE', %s, %s, 0,
                                        TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', 'MM/DD/YY HH24:MI:SS'),
                                        'Warehouse', 'Purged', %s, %s)
                            """, (job, p, work_order or '', username))

                    conn.commit()
                    logger.info(f"Purge operation: Deleted {deleted_count} zero-qty records for item {job} by {username}")

                    # Clear cache after inventory change
                    cache.delete_memoized(self.get_current_inventory)
                    cache.delete('stats_summary')

                    return {
                        'success': True,
                        'message': f'Successfully purged {deleted_count} zero-quantity record(s) for {job}',
                        'job': job,
                        'pcb_type': pcb_type,
                        'quantity_picked': 0,
                        'new_qty': 0,
                        'purged': True,
                        'records_deleted': deleted_count
                    }

                # CRITICAL: Lock rows with FOR UPDATE to prevent concurrent picks
                # Check if item exists in warehouse inventory with sufficient quantity
                # Use ILIKE for flexible matching (consistent with search_inventory)
                # If PCN specified, check only that PCN
                if pcn:
                    cursor.execute("""
                        SELECT onhandqty as total_qty
                        FROM pcb_inventory."tblWhse_Inventory"
                        WHERE pcn::text = %s AND item::text ILIKE %s
                        AND onhandqty > 0
                        FOR UPDATE  -- Lock rows to prevent race conditions
                    """, (str(pcn), job))
                else:
                    cursor.execute("""
                        SELECT SUM(onhandqty) as total_qty
                        FROM pcb_inventory."tblWhse_Inventory"
                        WHERE item::text ILIKE %s
                        AND onhandqty > 0
                        FOR UPDATE  -- Lock rows to prevent race conditions
                    """, (job,))

                result = cursor.fetchone()
                available_qty = int(result[0]) if result and result[0] else 0

                if available_qty < quantity:
                    conn.rollback()  # Release locks
                    pcn_msg = f" from PCN {pcn}" if pcn else ""
                    return {
                        'success': False,
                        'error': f'Cannot pick {quantity} units{pcn_msg}. Only {available_qty} available.',
                        'available_qty': available_qty,
                        'requested_qty': quantity,
                        'job': job,
                        'pcb_type': pcb_type
                    }

                # Update warehouse inventory - pick from specific locations using FIFO
                # This ensures we only pick the exact quantity needed from specific rows
                # If PCN is specified, only pick from that PCN
                # SECURE: Use conditional query execution instead of f-string interpolation

                if pcn:
                    # PCN-specific pick query
                    query_params = [job, str(pcn), quantity, quantity, quantity, quantity, quantity]
                    pick_query = """
                        WITH inventory_ordered AS (
                            SELECT
                                pcn,
                                item,
                                onhandqty,
                                migrated_at,
                                SUM(onhandqty) OVER (ORDER BY migrated_at, pcn) as running_total
                            FROM pcb_inventory."tblWhse_Inventory"
                            WHERE item::text ILIKE %s
                            AND pcn::text = %s
                            AND onhandqty > 0
                        ),"""
                else:
                    # FIFO pick query (all PCNs for item)
                    query_params = [job, quantity, quantity, quantity, quantity, quantity]
                    pick_query = """
                        WITH inventory_ordered AS (
                            SELECT
                                pcn,
                                item,
                                onhandqty,
                                migrated_at,
                                SUM(onhandqty) OVER (ORDER BY migrated_at, pcn) as running_total
                            FROM pcb_inventory."tblWhse_Inventory"
                            WHERE item::text ILIKE %s
                            AND onhandqty > 0
                        ),"""

                # Complete the query (same for both cases)
                pick_query += """
                    pick_rows AS (
                        SELECT
                            pcn,
                            item,
                            onhandqty,
                            running_total,
                            LAG(running_total, 1, 0) OVER (ORDER BY migrated_at, pcn) as prev_total
                        FROM inventory_ordered
                        ORDER BY migrated_at, pcn
                    ),
                    rows_to_update AS (
                        SELECT
                            pcn,
                            item,
                            CASE
                                -- If this row completes the pick, take only what's needed
                                WHEN prev_total < %s AND running_total >= %s
                                THEN %s - prev_total
                                -- If this row is fully consumed, take all
                                WHEN running_total <= %s
                                THEN onhandqty
                                ELSE 0
                            END as qty_to_pick
                        FROM pick_rows
                        WHERE prev_total < %s
                    )
                    UPDATE pcb_inventory."tblWhse_Inventory" w
                    SET onhandqty = GREATEST(0, w.onhandqty - r.qty_to_pick),
                        mfg_qty = (CASE WHEN w.mfg_qty ~ '^\-?[0-9]+$' THEN w.mfg_qty::integer ELSE 0 END + r.qty_to_pick)::text,
                        loc_to = 'MFG Floor'
                    FROM rows_to_update r
                    WHERE w.pcn::text = r.pcn::text
                    AND w.item = r.item
                    AND r.qty_to_pick > 0
                """

                cursor.execute(pick_query, tuple(query_params))

                updated_rows = cursor.rowcount

                if updated_rows == 0:
                    conn.rollback()
                    return {
                        'success': False,
                        'error': f'Job not found in inventory. Item {job} not found in warehouse inventory.',
                        'job': job,
                        'pcb_type': pcb_type
                    }

                # Record the pick transaction (movement from Receiving Area to MFG Floor)
                # If PCN specified, record transaction for that specific PCN
                # Otherwise, record for each PCN that was picked from using the FIFO logic
                if pcn:
                    # Single PCN pick - insert one transaction record
                    # Use actual loc_to from warehouse inventory as loc_from (where the part really is)
                    cursor.execute("""
                        INSERT INTO pcb_inventory."tblTransaction"
                        (trantype, item, pcn, mpn, dc, msd, tranqty, tran_time, loc_from, loc_to, wo, userid)
                        SELECT
                            'PICK',
                            %s,
                            pcn,
                            mpn,
                            CASE WHEN dc ~ '^[0-9]+$' THEN dc::integer ELSE NULL END as dc,
                            msd,
                            %s,
                            TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', 'MM/DD/YY HH24:MI:SS'),
                            COALESCE(loc_to, 'Warehouse'),
                            'MFG Floor',
                            %s,
                            %s
                        FROM pcb_inventory."tblWhse_Inventory"
                        WHERE item::text ILIKE %s AND pcn::text = %s
                        LIMIT 1
                    """, (job, quantity, work_order, username, job, str(pcn)))
                else:
                    # Multi-PCN FIFO pick - insert transaction for each PCN picked from
                    # Use same FIFO logic as the UPDATE query to record transactions accurately
                    query_params = [job]  # No PCN filter for FIFO
                    query_params.extend([quantity, quantity, quantity, quantity, quantity])
                    query_params.extend([job, work_order, username])

                    cursor.execute("""
                        WITH inventory_ordered AS (
                            SELECT
                                pcn,
                                item,
                                mpn,
                                dc,
                                msd,
                                loc_to,
                                onhandqty,
                                migrated_at,
                                SUM(onhandqty) OVER (ORDER BY migrated_at, pcn) as running_total
                            FROM pcb_inventory."tblWhse_Inventory"
                            WHERE item::text ILIKE %s
                            AND onhandqty > 0
                        ),
                        pick_rows AS (
                            SELECT
                                pcn,
                                item,
                                mpn,
                                dc,
                                msd,
                                loc_to,
                                onhandqty,
                                running_total,
                                LAG(running_total, 1, 0) OVER (ORDER BY migrated_at, pcn) as prev_total
                            FROM inventory_ordered
                            ORDER BY migrated_at, pcn
                        ),
                        rows_to_pick AS (
                            SELECT
                                pcn,
                                item,
                                mpn,
                                dc,
                                msd,
                                loc_to,
                                CASE
                                    WHEN prev_total < %s AND running_total >= %s
                                    THEN %s - prev_total
                                    WHEN running_total <= %s
                                    THEN onhandqty
                                    ELSE 0
                                END as qty_picked
                            FROM pick_rows
                            WHERE prev_total < %s
                        )
                        INSERT INTO pcb_inventory."tblTransaction"
                        (trantype, item, pcn, mpn, dc, msd, tranqty, tran_time, loc_from, loc_to, wo, userid)
                        SELECT
                            'PICK',
                            %s,
                            pcn::text,
                            mpn,
                            CASE WHEN dc ~ '^[0-9]+$' THEN dc::integer ELSE NULL END as dc,
                            msd,
                            qty_picked,
                            TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', 'MM/DD/YY HH24:MI:SS'),
                            COALESCE(loc_to, 'Warehouse'),
                            'MFG Floor',
                            %s,
                            %s
                        FROM rows_to_pick
                        WHERE qty_picked > 0
                    """, tuple(query_params))

                # Get the new remaining quantity
                cursor.execute("""
                    SELECT COALESCE(SUM(onhandqty), 0) as remaining_qty
                    FROM pcb_inventory."tblWhse_Inventory"
                    WHERE item::text ILIKE %s
                """, (job,))
                remaining_result = cursor.fetchone()
                new_qty = int(remaining_result[0]) if remaining_result and remaining_result[0] else 0

                conn.commit()
                logger.info(f"Pick operation: Updated {updated_rows} warehouse inventory records for item {job}, picked {quantity}, remaining {new_qty}, moved to MFG Floor")

                # Clear cache after inventory change
                cache.delete_memoized(self.get_current_inventory)
                cache.delete('stats_summary')

                return {
                    'success': True,
                    'message': f'Successfully picked {quantity} units of {job}',
                    'picked_qty': quantity,
                    'new_qty': new_qty,
                    'job': job,
                    'pcb_type': pcb_type
                }

            except psycopg2.extensions.TransactionRollbackError as e:
                # Serialization failure - concurrent transaction conflict
                if conn:
                    conn.rollback()
                logger.warning(f"Pick operation serialization conflict: {e}")
                return {
                    'success': False,
                    'error': 'This item is being accessed by another user. Please try again.'
                }
            except psycopg2.IntegrityError as e:
                if conn:
                    conn.rollback()
                logger.error(f"Pick operation integrity error: {e}")
                return {
                    'success': False,
                    'error': 'Data integrity error. Please verify your inputs.'
                }
            except Exception as e:
                if conn:
                    conn.rollback()
                logger.error(f"Failed to pick from warehouse inventory: {e}")
                raise
            finally:
                if cursor:
                    try:
                        cursor.close()
                    except Exception as e:
                        logger.error(f"Error closing cursor: {e}")
                if conn:
                    try:
                        self.return_connection(conn)
                    except Exception as e:
                        logger.error(f"Error returning connection: {e}")

        except Exception as e:
            logger.error(f"Pick operation failed: {e}", exc_info=True)
            error_msg = get_safe_error_message(e, "pick operation")
            return {'success': False, 'error': error_msg}

    def restock_pcb(self, pcn: int = None, item: str = None, quantity: int = 0,
                    location_from: str = 'MFG Floor', location_to: str = 'Count Area',
                    username: str = 'system') -> Dict[str, Any]:
        """Restock parts from specified source location to destination location."""
        # CRITICAL: Input validation
        if not isinstance(quantity, int) or quantity < 1 or quantity > 10000:
            return {
                'success': False,
                'error': f'Invalid quantity: {quantity}. Must be between 1 and 10,000.'
            }

        if pcn is not None:
            try:
                pcn_int = int(pcn)
                if pcn_int < 0 or pcn_int > 99999:
                    return {
                        'success': False,
                        'error': f'Invalid PCN: {pcn}. Must be between 1 and 99999.'
                    }
            except (ValueError, TypeError):
                return {
                    'success': False,
                    'error': f'Invalid PCN format: {pcn}. Must be numeric.'
                }

        if item and (not isinstance(item, str) or len(item) > 50):
            return {
                'success': False,
                'error': 'Invalid item identifier. Must be 1-50 characters.'
            }

        if not pcn and not item:
            return {
                'success': False,
                'error': 'Either PCN or Item number is required'
            }

        # Validate destination location exists (if provided)
        if location_to and not self.validate_location(location_to):
            return {
                'success': False,
                'error': f'Location "{location_to}" does not exist. Please verify the location code and try again.'
            }

        conn = None
        cursor = None
        try:
            conn = self.get_connection()
            conn.autocommit = False
            cursor = conn.cursor()

            try:
                # Determine search criteria and build query (SECURE: conditional execution)
                if pcn and item:
                    search_param = (str(pcn), item)
                    select_query = """
                        SELECT pcn, item, mpn, dc, mfg_qty, onhandqty, loc_to
                        FROM pcb_inventory."tblWhse_Inventory"
                        WHERE pcn::text = %s AND item::text ILIKE %s
                        FOR UPDATE
                        LIMIT 1
                    """
                elif pcn:
                    search_param = (str(pcn),)
                    select_query = """
                        SELECT pcn, item, mpn, dc, mfg_qty, onhandqty, loc_to
                        FROM pcb_inventory."tblWhse_Inventory"
                        WHERE pcn::text = %s
                        FOR UPDATE
                        LIMIT 1
                    """
                elif item:
                    search_param = (item,)
                    select_query = """
                        SELECT pcn, item, mpn, dc, mfg_qty, onhandqty, loc_to
                        FROM pcb_inventory."tblWhse_Inventory"
                        WHERE item = %s
                        FOR UPDATE
                        LIMIT 1
                    """
                else:
                    return {
                        'success': False,
                        'error': 'Either PCN or Item number is required'
                    }

                # CRITICAL: Lock row with FOR UPDATE and check MFG quantity
                cursor.execute(select_query, search_param)

                result = cursor.fetchone()

                if not result:
                    conn.rollback()  # Release lock
                    return {
                        'success': False,
                        'error': f'No parts found for {"PCN " + str(pcn) if pcn else "Item " + item}'
                    }

                pcn_num, item_num, mpn, dc, mfg_qty, current_onhand, existing_loc_to = result

                # If no destination specified, fall back to the part's existing location
                if not location_to:
                    location_to = existing_loc_to or 'Count Area'

                # Handle NULL quantities and convert mfg_qty from text to int
                if current_onhand is None:
                    current_onhand = 0
                if mfg_qty is None or mfg_qty == '':
                    mfg_qty_int = 0
                else:
                    try:
                        mfg_qty_int = int(mfg_qty)
                    except (ValueError, TypeError):
                        mfg_qty_int = 0

                # Check if this part was picked from inventory before allowing restock
                pick_check_query = """
                    SELECT COUNT(*) FROM pcb_inventory."tblTransaction"
                    WHERE trantype = 'PICK' AND pcn::text = %s
                """
                cursor.execute(pick_check_query, (str(pcn_num),))
                pick_count = cursor.fetchone()[0]
                if pick_count == 0:
                    conn.rollback()
                    return {
                        'success': False,
                        'error': f'You have not picked this part (PCN: {pcn_num}) from inventory. Parts must be picked before they can be restocked.'
                    }

                # Log if restocking more than tracked MFG quantity (user may have physical count)
                if mfg_qty_int < quantity:
                    logger.info(f'Restock qty ({quantity}) exceeds tracked MFG qty ({mfg_qty_int}) for PCN {pcn}. User override.')

                # Update warehouse inventory - move from specified source to destination
                # Use COALESCE to handle NULL onhandqty
                # Cast mfg_qty to integer for arithmetic, then back to text
                # SECURE: Use conditional query execution
                if pcn and item:
                    update_query = """
                        UPDATE pcb_inventory."tblWhse_Inventory"
                        SET mfg_qty = GREATEST(0, CASE WHEN mfg_qty ~ '^\-?[0-9]+$' THEN mfg_qty::integer ELSE 0 END - %s)::text,
                            onhandqty = COALESCE(onhandqty, 0) + %s,
                            loc_from = %s,
                            loc_to = %s
                        WHERE pcn::text = %s AND item::text ILIKE %s
                    """
                    cursor.execute(update_query, (quantity, quantity, location_from, location_to, str(pcn), item))
                elif pcn:
                    update_query = """
                        UPDATE pcb_inventory."tblWhse_Inventory"
                        SET mfg_qty = GREATEST(0, CASE WHEN mfg_qty ~ '^\-?[0-9]+$' THEN mfg_qty::integer ELSE 0 END - %s)::text,
                            onhandqty = COALESCE(onhandqty, 0) + %s,
                            loc_from = %s,
                            loc_to = %s
                        WHERE pcn::text = %s
                    """
                    cursor.execute(update_query, (quantity, quantity, location_from, location_to, str(pcn)))
                else:
                    update_query = """
                        UPDATE pcb_inventory."tblWhse_Inventory"
                        SET mfg_qty = GREATEST(0, CASE WHEN mfg_qty ~ '^\-?[0-9]+$' THEN mfg_qty::integer ELSE 0 END - %s)::text,
                            onhandqty = COALESCE(onhandqty, 0) + %s,
                            loc_from = %s,
                            loc_to = %s
                        WHERE item = %s
                    """
                    cursor.execute(update_query, (quantity, quantity, location_from, location_to, item))

                updated_rows = cursor.rowcount

                if updated_rows == 0:
                    conn.rollback()
                    return {
                        'success': False,
                        'error': 'Failed to update warehouse inventory'
                    }

                # Record the restock transaction
                # Get MSD from warehouse inventory
                cursor.execute("SELECT msd FROM pcb_inventory.\"tblWhse_Inventory\" WHERE pcn::text = %s LIMIT 1", (str(pcn_num),))
                msd_result = cursor.fetchone()
                msd = msd_result[0] if msd_result else None

                cursor.execute("""
                    INSERT INTO pcb_inventory."tblTransaction"
                    (trantype, item, pcn, mpn, dc, msd, tranqty, tran_time, loc_from, loc_to, userid)
                    VALUES ('RESTOCK', %s, %s, %s, %s, %s, %s, TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', 'MM/DD/YY HH24:MI:SS'), %s, %s, %s)
                """, (item_num, pcn_num, mpn, dc, msd, quantity, location_from, location_to, username))

                conn.commit()
                logger.info(f"Restock operation: PCN {pcn_num}, Item {item_num}, restocked {quantity} units from {location_from} to {location_to}")

                # Clear cache after inventory change
                cache.delete_memoized(self.get_current_inventory)
                cache.delete('stats_summary')

                return {
                    'success': True,
                    'message': f'Successfully restocked {quantity} units to {location_to}',
                    'quantity': quantity,
                    'pcn': pcn_num,
                    'item': item_num,
                    'mpn': mpn,
                    'location_to': location_to,
                    'new_mfg_qty': mfg_qty_int - quantity,
                    'new_onhand_qty': current_onhand + quantity
                }

            except psycopg2.extensions.TransactionRollbackError as e:
                # Serialization failure - concurrent transaction conflict
                if conn:
                    conn.rollback()
                logger.warning(f"Restock operation serialization conflict: {e}")
                return {
                    'success': False,
                    'error': 'This item is being accessed by another user. Please try again.'
                }
            except psycopg2.IntegrityError as e:
                if conn:
                    conn.rollback()
                logger.error(f"Restock operation integrity error: {e}")
                return {
                    'success': False,
                    'error': 'Data integrity error. Please verify your inputs.'
                }
            except Exception as e:
                if conn:
                    conn.rollback()
                logger.error(f"Failed to restock: {e}")
                raise
            finally:
                if cursor:
                    try:
                        cursor.close()
                    except Exception as e:
                        logger.error(f"Error closing cursor: {e}")
                if conn:
                    try:
                        self.return_connection(conn)
                    except Exception as e:
                        logger.error(f"Error returning connection: {e}")

        except Exception as e:
            logger.error(f"Restock operation failed: {e}", exc_info=True)
            error_msg = get_safe_error_message(e, "restock operation")
            return {'success': False, 'error': error_msg}

    def reverse_pick(self, transaction_id: int, username: str = 'system') -> Dict[str, Any]:
        """Reverse a specific PICK transaction - restores qty and original location."""
        conn = None
        cursor = None
        try:
            conn = self.get_connection()
            conn.autocommit = False
            cursor = conn.cursor(cursor_factory=RealDictCursor)

            # 1. Get the original PICK transaction
            cursor.execute("""
                SELECT id, trantype, item, pcn, mpn, dc, msd, tranqty, loc_from, loc_to, wo, userid, reversed
                FROM pcb_inventory."tblTransaction"
                WHERE id = %s
            """, (transaction_id,))
            txn = cursor.fetchone()

            if not txn:
                return {'success': False, 'error': 'Transaction not found.'}

            if txn['trantype'] != 'PICK':
                return {'success': False, 'error': f'Can only reverse PICK transactions. This is a {txn["trantype"]}.'}

            if txn['reversed']:
                return {'success': False, 'error': 'This pick has already been reversed.'}

            pick_qty = int(txn['tranqty']) if txn['tranqty'] else 0
            if pick_qty <= 0:
                return {'success': False, 'error': 'Invalid pick quantity.'}

            original_location = txn['loc_from'] or 'Warehouse'
            pcn_val = txn['pcn']
            item_val = txn['item']

            # 2. Restore inventory: onhandqty += pick_qty, mfg_qty -= pick_qty, loc_to = original location
            cursor.execute("""
                UPDATE pcb_inventory."tblWhse_Inventory"
                SET onhandqty = COALESCE(onhandqty, 0) + %s,
                    mfg_qty = GREATEST(0, CASE WHEN mfg_qty ~ '^\-?[0-9]+$' THEN mfg_qty::integer ELSE 0 END - %s)::text,
                    loc_to = %s
                WHERE item::text ILIKE %s AND pcn::text = %s
            """, (pick_qty, pick_qty, original_location, item_val, pcn_val))

            updated_rows = cursor.rowcount
            if updated_rows == 0:
                conn.rollback()
                return {'success': False, 'error': f'Inventory record not found for item {item_val}, PCN {pcn_val}. It may have been purged.'}

            # 3. Mark original transaction as reversed
            cursor.execute("""
                UPDATE pcb_inventory."tblTransaction"
                SET reversed = TRUE
                WHERE id = %s
            """, (transaction_id,))

            # 4. Create REVERSE_PICK transaction with reference to original
            import datetime, pytz
            est = pytz.timezone('US/Eastern')
            now_est = datetime.datetime.now(est)
            tran_time = now_est.strftime('%m/%d/%y %H:%M:%S')

            cursor.execute("""
                INSERT INTO pcb_inventory."tblTransaction"
                (trantype, item, pcn, mpn, dc, msd, tranqty, tran_time, loc_from, loc_to, wo, userid, ref_transaction_id)
                VALUES ('REVERSE_PICK', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                item_val, pcn_val, txn['mpn'], txn['dc'], txn['msd'],
                str(pick_qty), tran_time,
                'MFG Floor', original_location,
                txn['wo'], username, transaction_id
            ))

            reverse_txn_id = cursor.lastrowid or None
            # Get the new transaction ID
            cursor.execute("SELECT lastval()")
            reverse_txn_id = cursor.fetchone()['lastval']

            # Update the original with reverse reference
            cursor.execute("""
                UPDATE pcb_inventory."tblTransaction"
                SET reversed_by_id = %s
                WHERE id = %s
            """, (reverse_txn_id, transaction_id))

            conn.commit()

            # Clear cache
            cache.delete_memoized(self.get_current_inventory)

            logger.info(f"Reversed PICK transaction #{transaction_id}: {pick_qty} units of {item_val} (PCN {pcn_val}) restored to {original_location} by {username}")

            return {
                'success': True,
                'item': item_val,
                'pcn': pcn_val,
                'quantity': pick_qty,
                'restored_location': original_location,
                'original_work_order': txn['wo'],
                'reverse_transaction_id': reverse_txn_id
            }

        except Exception as e:
            if conn:
                conn.rollback()
            logger.error(f"Reverse pick failed: {e}", exc_info=True)
            return {'success': False, 'error': get_safe_error_message(e, "reverse pick")}
        finally:
            if cursor:
                try:
                    cursor.close()
                except Exception:
                    pass
            if conn:
                try:
                    self.return_connection(conn)
                except Exception:
                    pass

    def get_recent_picks(self, limit: int = 50) -> List[Dict[str, Any]]:
        """Get recent PICK transactions for the reverse picks panel."""
        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cursor:
                cursor.execute("""
                    SELECT id, item, pcn, mpn, tranqty, tran_time, loc_from, loc_to, wo, userid, reversed
                    FROM pcb_inventory."tblTransaction"
                    WHERE trantype = 'PICK'
                    ORDER BY id DESC
                    LIMIT %s
                """, (limit,))
                return cursor.fetchall()
        except Exception as e:
            logger.error(f"Error fetching recent picks: {e}")
            return []
        finally:
            if conn:
                self.return_connection(conn)

    def get_current_inventory(self, user_role: str = 'USER', itar_auth: bool = False) -> List[Dict[str, Any]]:
        """Get current warehouse inventory - cached for performance."""
        cache_key = f"warehouse_inventory_{user_role}_{itar_auth}"
        cached = cache.get(cache_key)
        if cached:
            return cached

        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Read directly from tblWhse_Inventory table (warehouse inventory)
                cur.execute(
                    """
                    SELECT
                        id,
                        pcn,
                        item as job,
                        mpn as pcb_type,
                        onhandqty as qty,
                        loc_to as location,
                        migrated_at as checked_on,
                        migrated_at as updated_at
                    FROM pcb_inventory."tblWhse_Inventory"
                    WHERE onhandqty > 0
                    ORDER BY item, mpn
                    """
                )
                result = [dict(row) for row in cur.fetchall()]
                cache.set(cache_key, result, timeout=60)  # Cache for 1 minute
                return result
        except Exception as e:
            logger.error(f"Failed to get warehouse inventory: {e}")
            return []
        finally:
            if conn:
                self.return_connection(conn)
    
    def get_inventory_summary(self, limit: int = 100) -> List[Dict[str, Any]]:
        """Get warehouse inventory summary grouped by MPN and location with descriptions."""
        cache_key = f"inventory_summary_{limit}"
        cached = cache.get(cache_key)
        if cached:
            return cached

        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute('''
                    SELECT
                        w.mpn as pcb_type,
                        w.loc_to as location,
                        COUNT(DISTINCT w.item) as job_count,
                        SUM(w.onhandqty) as total_qty,
                        AVG(w.onhandqty) as avg_qty,
                        MAX(p."DESC") as description
                    FROM pcb_inventory."tblWhse_Inventory" w
                    LEFT JOIN pcb_inventory."tblPN_List" p ON w.item = p.item
                    WHERE w.onhandqty > 0
                    GROUP BY w.mpn, w.loc_to
                    ORDER BY total_qty DESC, w.mpn, w.loc_to
                    LIMIT %s
                ''', (limit,))
                result = [dict(row) for row in cur.fetchall()]
                cache.set(cache_key, result, timeout=300)  # Cache for 5 minutes
                return result
        except Exception as e:
            logger.error(f"Failed to get warehouse summary: {e}")
            return []
        finally:
            if conn:
                self.return_connection(conn)
    
    def get_inventory_stats(self) -> Dict[str, int]:
        """Get accurate inventory statistics efficiently - just aggregates, no data loading."""
        cache_key = "inventory_stats_fast"
        cached = cache.get(cache_key)
        if cached:
            return cached

        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute('''
                    SELECT
                        COUNT(DISTINCT item) as total_jobs,
                        SUM(onhandqty) as total_quantity,
                        COUNT(*) as total_items,
                        COUNT(DISTINCT mpn) as unique_mpns
                    FROM pcb_inventory."tblWhse_Inventory"
                    WHERE onhandqty > 0
                ''')
                result = dict(cur.fetchone())
                cache.set(cache_key, result, timeout=300)  # Cache for 5 minutes
                return result
        except Exception as e:
            logger.error(f"Failed to get inventory stats: {e}")
            return {'total_jobs': 0, 'total_quantity': 0, 'total_items': 0, 'unique_mpns': 0}
        finally:
            if conn:
                self.return_connection(conn)

    def get_low_stock_items(self, threshold: int = 10, limit: int = 50) -> List[Dict[str, Any]]:
        """Get low stock items from entire database."""
        cache_key = f"low_stock_{threshold}_{limit}"
        cached = cache.get(cache_key)
        if cached:
            return cached

        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute('''
                    SELECT
                        item as job,
                        pcn,
                        mpn as pcb_type,
                        onhandqty as qty,
                        loc_to as location,
                        migrated_at as updated_at
                    FROM pcb_inventory."tblWhse_Inventory"
                    WHERE onhandqty > 0 AND onhandqty < %s
                    ORDER BY onhandqty ASC
                    LIMIT %s
                ''', (threshold, limit))
                result = [dict(row) for row in cur.fetchall()]
                cache.set(cache_key, result, timeout=300)  # Cache for 5 minutes
                return result
        except Exception as e:
            logger.error(f"Failed to get low stock items: {e}")
            return []
        finally:
            if conn:
                self.return_connection(conn)

    def get_audit_log(self, limit: int = 50) -> List[Dict[str, Any]]:
        """Get recent warehouse transaction entries."""
        cache_key = f"audit_log_{limit}"
        cached = cache.get(cache_key)
        if cached:
            return cached

        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT
                        t.id,
                        t.trantype as operation,
                        t.item as job,
                        t.mpn as pcb_type,
                        CASE WHEN t.tranqty::text ~ '^\-?[0-9]+$' THEN t.tranqty::integer ELSE 0 END as quantity_change,
                        COALESCE(w.onhandqty, CASE WHEN t.tranqty::text ~ '^\-?[0-9]+$' THEN t.tranqty::integer ELSE 0 END) as new_quantity,
                        t.tran_time as timestamp,
                        t.loc_from,
                        t.loc_to,
                        t.userid as user_id
                    FROM pcb_inventory."tblTransaction" t
                    LEFT JOIN LATERAL (
                        SELECT onhandqty FROM pcb_inventory."tblWhse_Inventory" w
                        WHERE w.pcn::text = t.pcn::text LIMIT 1
                    ) w ON true
                    WHERE t.trantype IN ('GEN', 'STOCK', 'PICK', 'UPDATE')
                      AND t.tran_time IS NOT NULL
                    ORDER BY t.tran_time DESC
                    LIMIT %s
                    """,
                    (limit,)
                )
                result = [dict(row) for row in cur.fetchall()]
                cache.set(cache_key, result, timeout=120)  # Cache for 2 minutes
                return result
        except Exception as e:
            logger.error(f"Failed to get audit log from transactions: {e}")
            return []
        finally:
            if conn:
                self.return_connection(conn)
    
    def get_dashboard_data(self, summary_limit=100, activity_limit=10, low_stock_threshold=10, low_stock_limit=50):
        """Get all dashboard data in a single DB connection to minimize round trips."""
        conn = None
        try:
            conn = self.get_connection()
            cur = conn.cursor(cursor_factory=RealDictCursor)

            # 1) Inventory stats
            cur.execute('''
                SELECT
                    COUNT(DISTINCT item) as total_jobs,
                    SUM(onhandqty) as total_quantity,
                    COUNT(*) as total_items,
                    COUNT(DISTINCT mpn) as unique_mpns
                FROM pcb_inventory."tblWhse_Inventory"
                WHERE onhandqty > 0
            ''')
            stats = dict(cur.fetchone())

            # 2) Inventory summary
            cur.execute('''
                SELECT
                    w.mpn as pcb_type,
                    w.loc_to as location,
                    COUNT(DISTINCT w.item) as job_count,
                    SUM(w.onhandqty) as total_qty,
                    AVG(w.onhandqty) as avg_qty,
                    MAX(p."DESC") as description
                FROM pcb_inventory."tblWhse_Inventory" w
                LEFT JOIN pcb_inventory."tblPN_List" p ON w.item = p.item
                WHERE w.onhandqty > 0
                GROUP BY w.mpn, w.loc_to
                ORDER BY total_qty DESC, w.mpn, w.loc_to
                LIMIT %s
            ''', (summary_limit,))
            summary = [dict(row) for row in cur.fetchall()]

            # 3) Low stock items
            cur.execute('''
                SELECT
                    item as job, pcn, mpn as pcb_type,
                    onhandqty as qty, loc_to as location, migrated_at as updated_at
                FROM pcb_inventory."tblWhse_Inventory"
                WHERE onhandqty > 0 AND onhandqty < %s
                ORDER BY onhandqty ASC
                LIMIT %s
            ''', (low_stock_threshold, low_stock_limit))
            low_stock = [dict(row) for row in cur.fetchall()]

            # 4) Recent activity
            cur.execute("""
                SELECT
                    t.id,
                    t.trantype as operation,
                    t.item as job,
                    t.mpn as pcb_type,
                    CASE WHEN t.tranqty::text ~ '^\-?[0-9]+$' THEN t.tranqty::integer ELSE 0 END as quantity_change,
                    COALESCE(w.onhandqty, CASE WHEN t.tranqty::text ~ '^\-?[0-9]+$' THEN t.tranqty::integer ELSE 0 END) as new_quantity,
                    t.tran_time as timestamp,
                    t.loc_from, t.loc_to,
                    t.userid as user_id
                FROM pcb_inventory."tblTransaction" t
                LEFT JOIN LATERAL (
                    SELECT onhandqty FROM pcb_inventory."tblWhse_Inventory" w
                    WHERE w.pcn::text = t.pcn::text LIMIT 1
                ) w ON true
                WHERE t.trantype IN ('GEN', 'STOCK', 'PICK', 'UPDATE')
                  AND t.tran_time IS NOT NULL
                ORDER BY t.tran_time DESC
                LIMIT %s
            """, (activity_limit,))
            activity = [dict(row) for row in cur.fetchall()]

            return {
                'stats': stats,
                'summary': summary,
                'low_stock': low_stock,
                'activity': activity,
            }
        except Exception as e:
            logger.error(f"Failed to get dashboard data: {e}")
            return None
        finally:
            if conn:
                self.return_connection(conn)

    def search_inventory(self, job: str = None, pcb_type: str = None, pcn: str = None,
                        user_role: str = 'USER', itar_auth: bool = False) -> List[Dict[str, Any]]:
        """Search warehouse inventory with optional filters.
        If PCN is provided, returns that specific PCN's data.
        Otherwise, returns TOTAL quantity per item (aggregated across all PCNs) for accurate pick validation.
        If job search returns no results, also tries matching by PCN number.
        """
        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                params = []

                # If PCN is specified, return that specific PCN's data (not aggregated)
                if pcn:
                    query = """
                        SELECT
                            item as job,
                            'Bare' as pcb_type,
                            onhandqty as qty,
                            loc_to as location,
                            dc as date_code,
                            msd as msd_level,
                            mpn as part_number,
                            pcn
                        FROM pcb_inventory."tblWhse_Inventory"
                        WHERE pcn::text = %s
                    """
                    params.append(pcn)

                    if job:
                        query += " AND item::text ILIKE %s"
                        params.append(f'%{job}%')

                    query += " ORDER BY migrated_at, pcn"
                else:
                    # Query warehouse inventory - aggregate by ITEM ONLY to show total available
                    # This ensures pick validation uses the correct total quantity
                    # Include items with zero qty so they can be found for purge operations
                    query = """
                        SELECT
                            item as job,
                            'Bare' as pcb_type,
                            SUM(onhandqty) as qty,
                            MAX(loc_to) as location,
                            MAX(dc) as date_code,
                            MAX(msd) as msd_level,
                            MAX(mpn) as part_number,
                            COUNT(DISTINCT pcn) as pcn_count
                        FROM pcb_inventory."tblWhse_Inventory"
                        WHERE onhandqty >= 0
                    """

                    if job:
                        query += " AND item::text ILIKE %s"
                        params.append(f'%{job}%')

                    query += " GROUP BY item"
                    query += " ORDER BY item"

                cur.execute(query, params)
                results = [dict(row) for row in cur.fetchall()]

                # If no results found and job looks like a PCN (numeric), try searching by PCN
                if not results and job and not pcn and job.strip().isdigit():
                    cur.execute("""
                        SELECT
                            item as job,
                            'Bare' as pcb_type,
                            onhandqty as qty,
                            loc_to as location,
                            dc as date_code,
                            msd as msd_level,
                            mpn as part_number,
                            pcn
                        FROM pcb_inventory."tblWhse_Inventory"
                        WHERE pcn::text = %s
                        ORDER BY migrated_at, pcn
                    """, (job.strip(),))
                    results = [dict(row) for row in cur.fetchall()]

                return results
        except Exception as e:
            logger.error(f"Search failed: {e}")
            return []
        finally:
            if conn:
                self.return_connection(conn)
    
    def get_stats_summary(self) -> Dict[str, Any]:
        """Get comprehensive statistics summary for stats page - cached for performance."""
        cached = cache.get('stats_summary')
        if cached:
            return cached

        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Get basic counts
                cur.execute("""
                    SELECT
                        COUNT(*) as total_records,
                        COUNT(DISTINCT job) as unique_jobs,
                        SUM(qty) as total_quantity,
                        COUNT(DISTINCT pcb_type) as pcb_types,
                        MAX(updated_at) as last_updated
                    FROM pcb_inventory.tblpcb_inventory
                """)
                stats = dict(cur.fetchone())

                # Format last_updated
                if stats['last_updated']:
                    stats['last_updated'] = stats['last_updated'].strftime('%B %d, %Y %I:%M %p')
                else:
                    stats['last_updated'] = 'Never'

                cache.set('stats_summary', stats, timeout=120)  # Cache for 2 minutes
                return stats
        except Exception as e:
            logger.error(f"Failed to get stats summary: {e}")
            return {
                'total_records': 0, 'unique_jobs': 0, 'total_quantity': 0,
                'pcb_types': 0, 'last_updated': 'Unknown'
            }
        finally:
            if conn:
                self.return_connection(conn)
    
    def get_pcb_type_breakdown(self) -> List[Dict[str, Any]]:
        """Get PCB type breakdown for stats comparison."""
        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT 
                        pcb_type as name,
                        SUM(qty) as postgres_count,
                        SUM(qty) as source_count  -- Assuming same for now
                    FROM pcb_inventory.tblpcb_inventory
                    GROUP BY pcb_type
                    ORDER BY pcb_type
                """)
                return [dict(row) for row in cur.fetchall()]
        except Exception as e:
            logger.error(f"Failed to get PCB type breakdown: {e}")
            return []
        finally:
            if conn:
                self.return_connection(conn)
    
    def get_location_breakdown(self) -> List[Dict[str, Any]]:
        """Get location distribution for stats page."""
        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("""
                    SELECT
                        location as range,
                        COUNT(*) as item_count,
                        SUM(qty) as total_qty,
                        ROUND((COUNT(*) * 100.0 / (SELECT COUNT(*) FROM pcb_inventory.tblpcb_inventory)), 1) as usage_percent
                    FROM pcb_inventory.tblpcb_inventory
                    GROUP BY location
                    ORDER BY location
                """)
                return [dict(row) for row in cur.fetchall()]
        except Exception as e:
            logger.error(f"Failed to get location breakdown: {e}")
            return []
        finally:
            if conn:
                self.return_connection(conn)

    def assign_pcn_to_item(self, job: str, pcb_type: str, username: str = 'system') -> Dict[str, Any]:
        """Assign a PCN to an inventory item using the database function."""
        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Call the assign_pcn database function
                cur.execute(
                    "SELECT pcb_inventory.assign_pcn(%s, %s, %s) as result",
                    (job, pcb_type, username)
                )
                result = cur.fetchone()
                conn.commit()

                if result and result['result']:
                    return result['result']
                else:
                    return {'success': False, 'error': 'PCN assignment failed'}
        except Exception as e:
            if conn:
                conn.rollback()
            logger.error(f"Failed to assign PCN: {e}")
            return {'success': False, 'error': str(e)}
        finally:
            if conn:
                self.return_connection(conn)

    def get_pcn_history(self, limit: int = 100, filters: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """Get PCN transaction history with warehouse inventory data - shows ALL transactions."""
        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Query from tblTransaction with warehouse inventory data
                # Show ALL transactions (removed DISTINCT ON to show complete history)
                query = """
                    SELECT
                        t.record_no,
                        t.trantype as status,
                        t.item as job,
                        t.pcn,
                        t.id as transaction_id,
                        COALESCE(w.mpn, t.mpn) as mpn,
                        COALESCE(w.dc::text, t.dc::text) as dc,
                        COALESCE(t.msd, w.msd, '0') as msd,
                        CASE WHEN t.tranqty::text ~ '^\-?[0-9]+$' THEN t.tranqty::integer ELSE 0 END as quantity,
                        CASE WHEN w.mfg_qty ~ '^\-?[0-9]+$' THEN w.mfg_qty::integer ELSE 0 END as mfg_qty,
                        t.tran_time as generated_at,
                        t.loc_from,
                        COALESCE(w.loc_to, t.loc_to) as location,
                        t.wo,
                        COALESCE(w.po, t.po) as po,
                        t.userid as user_id
                    FROM pcb_inventory."tblTransaction" t
                    LEFT JOIN pcb_inventory."tblWhse_Inventory" w
                        ON t.pcn = w.pcn
                    WHERE t.pcn IS NOT NULL
                """
                params = []

                if filters:
                    if filters.get('pcn'):
                        query += " AND t.pcn::text LIKE %s"
                        params.append(f"%{filters['pcn']}%")
                    if filters.get('job'):
                        query += " AND t.item::text LIKE %s"
                        params.append(f"%{filters['job']}%")
                    if filters.get('status'):
                        query += " AND t.trantype = %s"
                        params.append(filters['status'])

                query += " ORDER BY t.id DESC LIMIT %s"
                params.append(limit)

                cur.execute(query, params)
                return [dict(row) for row in cur.fetchall()]
        except Exception as e:
            logger.error(f"Failed to get PCN history: {e}")
            return []
        finally:
            if conn:
                self.return_connection(conn)

    def search_pcn(self, pcn_number: str = None, job: str = None) -> List[Dict[str, Any]]:
        """Search for PCN records by PCN number or job number - returns unique PCNs only, newest first."""
        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                query = """
                    SELECT * FROM (
                        SELECT DISTINCT ON (t.pcn)
                            t.record_no,
                            t.trantype as status,
                            t.item as job,
                            t.pcn,
                            t.id as transaction_id,
                            COALESCE(w.mpn, t.mpn) as mpn,
                            COALESCE(w.dc::text, t.dc::text) as dc,
                            COALESCE(t.msd, w.msd, '0') as msd,
                            CASE WHEN t.tranqty::text ~ '^\-?[0-9]+$' THEN t.tranqty::integer ELSE 0 END as quantity,
                            CASE WHEN w.mfg_qty ~ '^\-?[0-9]+$' THEN w.mfg_qty::integer ELSE 0 END as mfg_qty,
                            t.tran_time as generated_at,
                            t.loc_from,
                            COALESCE(w.loc_to, t.loc_to) as location,
                            t.wo,
                            COALESCE(w.po, t.po) as po,
                            t.userid as user_id
                        FROM pcb_inventory."tblTransaction" t
                        LEFT JOIN pcb_inventory."tblWhse_Inventory" w
                            ON t.pcn = w.pcn
                        WHERE t.pcn IS NOT NULL
                """
                params = []

                if pcn_number:
                    query += " AND t.pcn::text LIKE %s"
                    params.append(f"%{pcn_number}%")

                if job:
                    query += " AND t.item::text LIKE %s"
                    params.append(f"%{job}%")

                query += " ORDER BY t.pcn, t.id DESC"
                query += " ) sub ORDER BY CAST(pcn AS INTEGER) DESC, transaction_id DESC"

                cur.execute(query, params)
                return [dict(row) for row in cur.fetchall()]
        except Exception as e:
            logger.error(f"PCN search failed: {e}")
            return []
        finally:
            if conn:
                self.return_connection(conn)

    def get_po_history(self, limit: int = 100, offset: int = 0, filters: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """Get PO history with optional filters and pagination."""
        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                query = """
                    SELECT id, po as po_number, item, pcn, mpn, dc as date_code,
                           CASE
                               WHEN tranqty ~ '^[0-9]+$' THEN CAST(tranqty AS INTEGER)
                               ELSE NULL
                           END as quantity,
                           trantype as transaction_type, tran_time as transaction_date,
                           loc_from as location_from, loc_to as location_to, userid as user_id
                    FROM pcb_inventory."tblTransaction"
                    WHERE po IS NOT NULL AND po <> ''
                """
                params = []

                if filters:
                    if filters.get('po_number'):
                        query += " AND po LIKE %s"
                        params.append(f"%{filters['po_number']}%")
                    if filters.get('item'):
                        query += " AND item LIKE %s"
                        params.append(f"%{filters['item']}%")
                    if filters.get('date_from'):
                        query += " AND tran_time >= %s"
                        params.append(filters['date_from'])
                    if filters.get('date_to'):
                        query += " AND tran_time <= %s"
                        params.append(filters['date_to'])

                query += " ORDER BY id DESC LIMIT %s OFFSET %s"
                params.append(limit)
                params.append(offset)

                cur.execute(query, params)
                results = [dict(row) for row in cur.fetchall()]

                # Convert quantity strings to integers
                for result in results:
                    if result.get('quantity') is not None:
                        try:
                            if isinstance(result['quantity'], str):
                                result['quantity'] = int(result['quantity']) if result['quantity'].strip() else 0
                        except (ValueError, AttributeError):
                            result['quantity'] = 0

                return results
        except Exception as e:
            logger.error(f"Failed to get PO history: {e}")
            return []

    def get_po_history_count(self, filters: Dict[str, Any] = None) -> int:
        """Get total count of PO history records with optional filters."""
        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor() as cur:
                query = """
                    SELECT COUNT(*) FROM pcb_inventory."tblTransaction"
                    WHERE po IS NOT NULL AND po <> ''
                """
                params = []

                if filters:
                    if filters.get('po_number'):
                        query += " AND po LIKE %s"
                        params.append(f"%{filters['po_number']}%")
                    if filters.get('item'):
                        query += " AND item LIKE %s"
                        params.append(f"%{filters['item']}%")
                    if filters.get('date_from'):
                        query += " AND tran_time >= %s"
                        params.append(filters['date_from'])
                    if filters.get('date_to'):
                        query += " AND tran_time <= %s"
                        params.append(filters['date_to'])

                cur.execute(query, params)
                return cur.fetchone()[0]
        except Exception as e:
            logger.error(f"Failed to get PO history count: {e}")
            return 0
        finally:
            if conn:
                self.return_connection(conn)

    def search_po(self, po_number: str = None, item: str = None) -> List[Dict[str, Any]]:
        """Search for PO records by PO number or item."""
        conn = None
        try:
            conn = self.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                query = """
                    SELECT id, po as po_number, item, pcn, mpn, dc as date_code,
                           CASE
                               WHEN tranqty ~ '^[0-9]+$' THEN CAST(tranqty AS INTEGER)
                               ELSE NULL
                           END as quantity,
                           trantype as transaction_type, tran_time as transaction_date,
                           loc_from as location_from, loc_to as location_to, userid as user_id
                    FROM pcb_inventory."tblTransaction"
                    WHERE po IS NOT NULL AND po <> ''
                """
                params = []

                if po_number:
                    query += " AND po LIKE %s"
                    params.append(f"%{po_number}%")

                if item:
                    query += " AND item LIKE %s"
                    params.append(f"%{item}%")

                query += " ORDER BY id DESC"

                cur.execute(query, params)
                results = [dict(row) for row in cur.fetchall()]

                # Convert quantity strings to integers
                for result in results:
                    if result.get('quantity') is not None:
                        try:
                            if isinstance(result['quantity'], str):
                                result['quantity'] = int(result['quantity']) if result['quantity'].strip() else 0
                        except (ValueError, AttributeError):
                            result['quantity'] = 0

                return results
        except Exception as e:
            logger.error(f"PO search failed: {e}")
            return []
        finally:
            if conn:
                self.return_connection(conn)

# User Authentication and Authorization Functions
class UserManager:
    """Handle user authentication and authorization."""
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
    
    def get_user_by_username(self, username: str) -> Dict[str, Any]:
        """Get user information by username."""
        conn = None
        try:
            conn = self.db_manager.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    "SELECT * FROM pcb_inventory.users WHERE username = %s AND active = TRUE",
                    (username,)
                )
                user = cur.fetchone()
                return dict(user) if user else None
        except Exception as e:
            logger.error(f"Failed to get user {username}: {e}")
            return None
        finally:
            if conn:
                self.db_manager.return_connection(conn)
    
    def get_all_users(self) -> List[Dict[str, Any]]:
        """Get all active users for the demo interface."""
        conn = None
        try:
            conn = self.db_manager.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    "SELECT username, role, itar_authorized FROM pcb_inventory.users WHERE active = TRUE ORDER BY username"
                )
                return [dict(row) for row in cur.fetchall()]
        except Exception as e:
            logger.error(f"Failed to get users: {e}")
            return []
        finally:
            if conn:
                self.db_manager.return_connection(conn)
    
    def can_access_itar(self, user_role: str, itar_authorized: bool) -> bool:
        """Check if user can access ITAR items."""
        return user_role == 'Super User' or user_role == 'ITAR'
    
    def simulate_aci_login(self, username: str) -> Dict[str, Any]:
        """Simulate login from ACI dashboard."""
        user = self.get_user_by_username(username)
        if not user:
            return {'success': False, 'error': 'User not found'}
        
        # Create session token
        session_token = secrets.token_urlsafe(32)
        
        # Update user's session info
        conn = None
        try:
            conn = self.db_manager.get_connection()
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE pcb_inventory.users SET session_token = %s, token_expires_at = %s, last_login = %s WHERE username = %s",
                    (session_token, datetime.now().replace(hour=23, minute=59, second=59), datetime.now(), username)
                )
                conn.commit()
        except Exception as e:
            logger.error(f"Failed to update session for {username}: {e}")
        finally:
            if conn:
                self.db_manager.return_connection(conn)
        
        return {
            'success': True,
            'user': user,
            'session_token': session_token
        }

def require_auth(f):
    """Decorator to require user authentication - NO GUEST ACCESS."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Check if user is logged in
        if 'user_id' not in session or 'username' not in session:
            # For API requests, return JSON error instead of redirect
            if request.path.startswith('/api/'):
                return jsonify({
                    'success': False,
                    'error': 'Authentication required. Please log in.'
                }), 401

            # For page requests, redirect to FORGE login with SSO redirect back to KOSH
            is_local = request.host and ('.local' in request.host or '192.168.' in request.host or 'localhost' in request.host)
            if is_local:
                forge_login_url = 'http://acidashboard.aci.local:2005/login?redirect=kosh'
            else:
                forge_login_url = 'https://aci-forge.vercel.app/login?redirect=kosh'
            return redirect(forge_login_url)

        # Check for ACI Dashboard SSO token in headers (optional)
        auth_token = request.headers.get('X-ACI-Auth-Token') or session.get('aci_auth_token')
        if auth_token:
            session['aci_auth_token'] = auth_token

        return f(*args, **kwargs)
    return decorated_function

def require_itar_access(f):
    """Decorator to require ITAR access."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user_role = session.get('role', 'USER')
        itar_auth = session.get('itar_authorized', False)
        
        if not user_manager.can_access_itar(user_role, itar_auth):
            flash('Access denied: ITAR authorization required', 'error')
            return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function

@app.before_request
def load_current_user():
    """Load current user information into g object."""
    g.current_user = {
        'username': session.get('full_name', session.get('username', 'anonymous')),
        'role': session.get('role', 'USER'),
        'itar_authorized': session.get('itar_authorized', False)
    }
    g.user_can_see_itar = user_manager.can_access_itar(
        g.current_user['role'],
        g.current_user['itar_authorized']
    ) if 'user_manager' in globals() else False

# Initialize database manager
db_manager = DatabaseManager()
user_manager = UserManager(db_manager)
expiration_manager = ExpirationManager()

import threading

# Ensure tblActivityLog exists (lightweight check, skips migration on Vercel cold starts)
def _ensure_activity_log_table():
    """Create tblActivityLog if needed. Runs once per process."""
    conn = None
    try:
        conn = db_manager.get_connection()
        cur = conn.cursor()
        # Quick check if table already exists — avoids expensive CREATE IF NOT EXISTS on every cold start
        cur.execute("""
            SELECT 1 FROM information_schema.tables
            WHERE table_schema = 'pcb_inventory' AND table_name = 'tblActivityLog'
        """)
        if cur.fetchone():
            db_manager.return_connection(conn)
            return  # Table exists, skip all migration work

        cur.execute('''
            CREATE TABLE IF NOT EXISTS pcb_inventory."tblActivityLog" (
                id SERIAL PRIMARY KEY,
                user_id INTEGER,
                username VARCHAR(100),
                full_name VARCHAR(200),
                action_type VARCHAR(50) NOT NULL,
                description TEXT,
                details TEXT,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York'),
                seen BOOLEAN DEFAULT FALSE,
                seen_at TIMESTAMP
            )
        ''')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_activity_log_created ON pcb_inventory."tblActivityLog" (created_at DESC)')
        cur.execute('CREATE INDEX IF NOT EXISTS idx_activity_log_seen ON pcb_inventory."tblActivityLog" (seen)')
        conn.commit()
        logger.info("tblActivityLog table created")
    except Exception as e:
        logger.error(f"Failed to ensure tblActivityLog: {e}")
    finally:
        if conn:
            try:
                db_manager.return_connection(conn)
            except:
                pass

# Run in background thread so it doesn't block app startup
threading.Thread(target=_ensure_activity_log_table, daemon=True).start()


def _ensure_aci_partnumbers_table():
    """Create tblACI_PartNumbers if needed for manual ACI number creation."""
    conn = None
    try:
        conn = db_manager.get_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT 1 FROM information_schema.tables
            WHERE table_schema = 'pcb_inventory' AND table_name = 'tblACI_PartNumbers'
        """)
        if cur.fetchone():
            db_manager.return_connection(conn)
            return

        cur.execute('''
            CREATE TABLE IF NOT EXISTS pcb_inventory."tblACI_PartNumbers" (
                id SERIAL PRIMARY KEY,
                aci_pn VARCHAR(20) NOT NULL UNIQUE,
                manufacturer VARCHAR(255),
                mpn VARCHAR(255),
                description TEXT,
                comment TEXT,
                loaded VARCHAR(1) DEFAULT 'N',
                created_by VARCHAR(100),
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York')
            )
        ''')
        cur.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_aci_pn_unique ON pcb_inventory."tblACI_PartNumbers" (aci_pn)')
        conn.commit()
        logger.info("tblACI_PartNumbers table created")
    except Exception as e:
        logger.error(f"Failed to ensure tblACI_PartNumbers: {e}")
    finally:
        if conn:
            try:
                db_manager.return_connection(conn)
            except:
                pass

threading.Thread(target=_ensure_aci_partnumbers_table, daemon=True).start()


def _do_log_activity(user_id, username, full_name, action_type, description, details):
    """Background worker that inserts a row into tblActivityLog."""
    conn = None
    try:
        conn = db_manager.get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO pcb_inventory."tblActivityLog"
            (user_id, username, full_name, action_type, description, details, seen)
            VALUES (%s, %s, %s, %s, %s, %s, FALSE)
        """, (user_id, username, full_name, action_type, description, details))
        conn.commit()
    except Exception as e:
        logger.error(f"Failed to log activity: {e}")
        if conn:
            try:
                conn.rollback()
            except:
                pass
    finally:
        if conn:
            try:
                db_manager.return_connection(conn)
            except:
                pass


def log_user_activity(action_type, description, details=None):
    """Log a user activity to tblActivityLog in a background thread (non-blocking).

    action_type: LOGIN, LOGOUT, STOCK, PICK, RESTOCK, PCN_GENERATE, SHORTAGE_REPORT, PART_NUMBER_CHANGE
    """
    user_id = session.get('user_id')
    username = session.get('username', 'system')
    full_name = session.get('full_name', '')

    # Skip logging for super admin
    if username and username.lower() == 'kanav':
        return

    t = threading.Thread(
        target=_do_log_activity,
        args=(user_id, username, full_name, action_type, description, details),
        daemon=True
    )
    t.start()


@app.route('/health')
def health_check():
    """Health check endpoint for Docker."""
    try:
        # Test database connection
        inventory = db_manager.get_current_inventory()
        pool_stats = db_manager.get_pool_stats()

        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'inventory_items': len(inventory),
            'connection_pool': pool_stats,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/health/database')
def database_health_check():
    """Detailed database health check endpoint."""
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor()

        # Test query execution
        cursor.execute("SELECT COUNT(*) FROM pcb_inventory.\"tblWhse_Inventory\"")
        inventory_count = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM pcb_inventory.\"tblTransaction\"")
        transaction_count = cursor.fetchone()[0]

        cursor.close()
        db_manager.return_connection(conn)

        pool_stats = db_manager.get_pool_stats()

        return jsonify({
            'status': 'healthy',
            'database': {
                'connected': True,
                'inventory_records': inventory_count,
                'transaction_records': transaction_count
            },
            'connection_pool': pool_stats,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        logger.error(f"Database health check failed: {e}")
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("50 per minute", methods=["POST"])  # Only rate-limit actual login attempts
def login():
    """Secure login page with bulletproof authentication."""
    # If already logged in, redirect to dashboard
    if 'user_id' in session:
        return redirect(url_for('index'))

    # For GET requests, redirect to FORGE login (centralized auth)
    if request.method == 'GET':
        is_local = request.host and ('.local' in request.host or '192.168.' in request.host or 'localhost' in request.host)
        if is_local:
            forge_login_url = 'http://acidashboard.aci.local:2005/login?redirect=kosh'
        else:
            forge_login_url = 'https://aci-forge.vercel.app/login?redirect=kosh'
        return redirect(forge_login_url)

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember = request.form.get('remember') == 'on'

        if not username or not password:
            flash('Please provide both username and password.', 'danger')
            return render_template('login.html')

        # Get user from database
        conn = None
        try:
            conn = db_manager.get_connection()
            cursor = conn.cursor(cursor_factory=RealDictCursor)

            cursor.execute("""
                SELECT id, userid, username, userlogin, password, usersecurity
                FROM pcb_inventory."tblUser"
                WHERE userlogin = %s
            """, (username,))

            user = cursor.fetchone()

            # Secure logging - NO PASSWORDS
            logger.info(f"Login attempt - username: '{username}', user found: {user is not None}")

            # Check password with bcrypt (secure hashing)
            if user:
                try:
                    # Check if password is already hashed (starts with $2b$)
                    if user['password'].startswith('$2b$'):
                        # Password is hashed, use bcrypt.checkpw()
                        password_match = bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8'))
                    else:
                        # Legacy plain text password - compare directly but log warning
                        logger.warning(f"User {username} has plain text password - should be migrated to bcrypt")
                        password_match = (user['password'] == password)
                except Exception as e:
                    logger.error(f"Password verification error: {e}")
                    password_match = False
            else:
                password_match = False

            if user and password_match:
                # Successful login
                session.clear()  # Clear any old session data
                session['user_id'] = user['id']
                session['username'] = user['userlogin']
                session['full_name'] = user['username']
                # Set role from DB, but override to Admin for authorized users
                db_role = user['usersecurity']
                if user['userlogin'].lower() in ADMIN_AUTHORIZED_USERS and db_role != 'Admin':
                    session['role'] = 'Admin'
                    logger.info(f"User {username} promoted to Admin (in ADMIN_AUTHORIZED_USERS, DB role: '{db_role}')")
                else:
                    session['role'] = db_role
                    logger.info(f"User {username} role from DB: '{db_role}' (type: {type(db_role).__name__})")
                session['itar_authorized'] = True if session['role'] == 'Admin' else False
                session.permanent = True  # Always use 14-hour session

                # Record login notification for all users except super admin (kanav)
                if user['userlogin'].lower() != 'kanav':
                    try:
                        cursor.execute("""
                            INSERT INTO pcb_inventory."tblLoginNotifications"
                            (user_id, username, full_name, login_time, seen)
                            VALUES (%s, %s, %s, CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', FALSE)
                        """, (user['id'], user['userlogin'], user['username']))
                        logger.info(f"Login notification recorded for user: {username}")
                    except Exception as notif_error:
                        logger.error(f"Failed to record login notification: {notif_error}")

                conn.commit()

                # Log activity
                log_user_activity('LOGIN', f'{user["username"] or username} logged in')

                logger.info(f"Successful login: {username}")
                flash(f'Welcome back, {user["username"] or username}!', 'success')

                # Redirect to next page or dashboard (with whitelist validation)
                next_page = request.args.get('next')
                # Whitelist of allowed redirect paths
                allowed_redirects = [
                    '/', '/index', '/dashboard',
                    '/stock', '/pick', '/restock',
                    '/generate_pcn', '/pcn_history',
                    '/bom_loader', '/warehouse_inventory'
                ]
                if next_page and next_page in allowed_redirects:
                    return redirect(next_page)
                elif next_page and next_page.startswith('/') and '/' in next_page[1:]:
                    # Allow paths like /warehouse_inventory/view
                    base_path = '/' + next_page.lstrip('/').split('/')[0]
                    if base_path in allowed_redirects:
                        return redirect(next_page)
                return redirect(url_for('index'))
            else:
                # Failed login
                logger.warning(f"Failed login attempt for username: {username}")
                flash('Invalid username or password. Please try again.', 'danger')
                return redirect(url_for('login'))

        except Exception as e:
            logger.error(f"Login error: {e}")
            flash('An error occurred. Please try again later.', 'danger')
            return redirect(url_for('login'))
        finally:
            if conn:
                db_manager.return_connection(conn)

    return render_template('login.html')

@app.route('/logout')
def logout():
    """Secure logout - clears all session data."""
    username = session.get('username', 'Unknown')
    full_name = session.get('full_name', '')
    log_user_activity('LOGOUT', f'{full_name or username} logged out')
    session.clear()
    logger.info(f"User logged out: {username}")
    # Redirect to FORGE login with redirect back to KOSH
    is_local = request.host and ('.local' in request.host or '192.168.' in request.host or 'localhost' in request.host)
    if is_local:
        return redirect('http://acidashboard.aci.local:2005/login?redirect=kosh')
    return redirect('https://aci-forge.vercel.app/login?redirect=kosh')

@app.route('/')
@require_auth
def index():
    """Main dashboard page - optimized for fast loading with accurate stats."""
    try:
        # Serve cached dashboard data if available (60s TTL)
        dashboard_cache_key = "dashboard_data_v1"
        cached_dashboard = cache.get(dashboard_cache_key)

        if cached_dashboard:
            stats = cached_dashboard['stats']
            enhanced_summary = cached_dashboard['enhanced_summary']
            recent_activity = cached_dashboard['recent_activity']
            low_stock_items = cached_dashboard['low_stock_items']
            most_active_jobs = cached_dashboard['most_active_jobs']
            pcb_type_data = cached_dashboard['pcb_type_data']
            LOW_STOCK_THRESHOLD = cached_dashboard['low_stock_threshold']
        else:
            LOW_STOCK_THRESHOLD = 10

            # Single DB connection for all dashboard queries
            dashboard = db_manager.get_dashboard_data(
                summary_limit=100, activity_limit=10,
                low_stock_threshold=LOW_STOCK_THRESHOLD, low_stock_limit=50
            )

            if dashboard:
                stats_data = dashboard['stats']
                summary = dashboard['summary']
                recent_activity = dashboard['activity']
                low_stock_items = dashboard['low_stock']
            else:
                stats_data = {'total_jobs': 0, 'total_quantity': 0, 'total_items': 0, 'unique_mpns': 0}
                summary = []
                recent_activity = []
                low_stock_items = []

            total_jobs = stats_data.get('total_jobs', 0)
            total_quantity = stats_data.get('total_quantity', 0) or 0
            total_items = stats_data.get('total_items', 0)

            # Most active jobs from summary (top 5)
            most_active_jobs = sorted(
                [(item.get('pcb_type', 'Unknown'), item.get('total_qty', 0)) for item in summary],
                key=lambda x: x[1],
                reverse=True
            )[:5]

            # PCB type distribution for chart
            pcb_type_data = {}
            for item in summary:
                pcb_type = item.get('pcb_type') or 'Unknown'
                qty = item.get('total_qty') or 0
                pcb_type_data[pcb_type] = pcb_type_data.get(pcb_type, 0) + qty

            stats = {
                'total_jobs': total_jobs,
                'total_quantity': total_quantity,
                'total_items': total_items,
                'pcb_types': stats_data.get('unique_mpns', 0),
                'low_stock_count': len(low_stock_items)
            }

            # Enhanced summary with safe formatting
            enhanced_summary = []
            if summary:
                for item in summary:
                    enhanced_item = dict(item)
                    enhanced_item['job_count'] = enhanced_item.get('job_count', 1)
                    enhanced_item['total_quantity'] = enhanced_item.get('total_qty', 0)
                    enhanced_item['average_quantity'] = enhanced_item.get('total_qty', 0) / max(enhanced_item.get('job_count', 1), 1)
                    enhanced_summary.append(enhanced_item)

            # Cache all dashboard data for 60 seconds
            cache.set(dashboard_cache_key, {
                'stats': stats,
                'enhanced_summary': enhanced_summary,
                'recent_activity': recent_activity,
                'low_stock_items': low_stock_items,
                'low_stock_threshold': LOW_STOCK_THRESHOLD,
                'most_active_jobs': most_active_jobs,
                'pcb_type_data': pcb_type_data,
            }, timeout=60)

        return render_template('index.html',
                             stats=stats,
                             summary=enhanced_summary,
                             recent_activity=recent_activity,
                             low_stock_items=low_stock_items,
                             low_stock_threshold=LOW_STOCK_THRESHOLD,
                             most_active_jobs=most_active_jobs,
                             pcb_type_data=pcb_type_data)
    except Exception as e:
        import traceback
        logger.error(f"Error loading dashboard: {e}")
        logger.error(traceback.format_exc())
        # Provide safe default values on error
        safe_stats = {
            'total_jobs': 0,
            'total_quantity': 0,
            'total_items': 0,
            'pcb_types': len(PCB_TYPES),
            'low_stock_count': 0
        }
        flash('Error loading dashboard. Please try again.', 'error')
        return render_template('index.html', stats=safe_stats, summary=[], recent_activity=[],
                             low_stock_items=[], low_stock_threshold=10, most_active_jobs=[], pcb_type_data={}, inventory_with_trends=[])

@app.route('/stock', methods=['GET', 'POST'])
@require_auth
def stock():
    """Stock PCB page."""
    form = StockForm()

    if form.validate_on_submit():
        logger.info(f"Stock form validation passed - Form data: job={form.job.data}, part_number={form.part_number.data}, quantity={form.quantity.data}, location_from={form.location_from.data}, location_to={form.location_to.data}")

        try:
            # Convert PCN to integer if provided
            pcn_value = None
            if hasattr(form, 'pcn_number') and form.pcn_number.data:
                try:
                    pcn_value = int(form.pcn_number.data) if form.pcn_number.data else None
                except (ValueError, TypeError):
                    pcn_value = None

            # Use part_number as job identifier if job not provided
            job_value = form.job.data if form.job.data else form.part_number.data

            logger.info(f"Calling stock_pcb with: job={job_value}, quantity={form.quantity.data}, location_from={form.location_from.data}, location_to={form.location_to.data}, pcn={pcn_value}")
            result = db_manager.stock_pcb(
                job=job_value,
                pcb_type='Bare',  # Default value since field was removed
                quantity=form.quantity.data,
                location_from=form.location_from.data,
                location_to=form.location_to.data,
                itar_classification='NONE',  # ITAR removed from form
                user_role=session.get('role', 'USER'),
                itar_auth=session.get('itar_authorized', False),
                username=session.get('username', 'system'),
                work_order=form.po.data if hasattr(form, 'po') and form.po.data else None,
                dc=form.dc.data if hasattr(form, 'dc') and form.dc.data else None,
                msd=form.msd.data if hasattr(form, 'msd') and form.msd.data else None,
                pcn=pcn_value,
                mpn=form.mpn.data if hasattr(form, 'mpn') and form.mpn.data else None,
                part_number=form.part_number.data if hasattr(form, 'part_number') and form.part_number.data else None
            )
            logger.info(f"stock_pcb returned: {result}")

            if result.get('success'):
                log_user_activity('STOCK', f"Stocked {result['stocked_qty']} units of {result['job']}", f"New total: {result['new_qty']}")
                flash(f"Successfully stocked {result['stocked_qty']} units of {result['job']}. "
                      f"New total: {result['new_qty']}", 'success')
                return redirect(url_for('stock'))
            else:
                flash(f"Stock operation failed: {result.get('error', 'Unknown error')}", 'error')
                
        except Exception as e:
            logger.error(f"Stock operation error: {e}")
            flash('Stock operation failed. Please try again.', 'error')
    else:
        if form.errors:
            logger.error(f"Stock form validation failed - Errors: {form.errors}")
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f"{field}: {error}", 'error')

    return render_template('stock.html', form=form)

@app.route('/pick', methods=['GET', 'POST'])
@require_auth
def pick():
    """Pick PCB page."""
    form = PickForm()
    
    if form.validate_on_submit():
        logger.info(f"Pick form validation passed - Form data: job={form.job.data}, part_number={form.part_number.data}, quantity={form.quantity.data}")

        try:
            user_role = session.get('role', 'USER')
            itar_auth = session.get('itar_authorized', False)

            # Use part_number as job identifier if job not provided
            job_value = form.job.data if form.job.data else form.part_number.data
            pcn_value = form.pcn.data if form.pcn.data else None

            logger.info(f"Calling pick_pcb with: job={job_value}, pcn={pcn_value}, quantity={form.quantity.data}")
            result = db_manager.pick_pcb(
                job=job_value,
                pcb_type='Bare',  # Default value since field was removed
                quantity=form.quantity.data,
                user_role=user_role,
                itar_auth=itar_auth,
                username=session.get('username', 'system'),
                work_order=form.work_order.data if form.work_order.data else None,
                pcn=pcn_value  # Pass PCN if specified - picks from that specific PCN only
            )
            logger.info(f"pick_pcb returned: {result}")

            if result.get('success'):
                if result.get('purged'):
                    log_user_activity('PICK', f"Purged {result.get('records_deleted', 0)} zero-quantity record(s) for {result['job']}")
                    flash(f"Successfully purged {result.get('records_deleted', 0)} zero-quantity record(s) for {result['job']}.", 'success')
                else:
                    log_user_activity('PICK', f"Picked {result['picked_qty']} units of {result['job']}", f"Remaining: {result['new_qty']}")
                    flash(f"Successfully picked {result['picked_qty']} units of {result['job']}. "
                          f"Remaining: {result['new_qty']}", 'success')
                return redirect(url_for('pick'))
            else:
                error_msg = result.get('error', 'Unknown error')
                if 'Insufficient quantity' in error_msg:
                    flash(f"Insufficient quantity! Available: {result.get('available_qty', 0)}, "
                          f"Requested: {result.get('requested_qty', 0)}", 'error')
                elif 'Job not found' in error_msg:
                    flash(f"Item {result['job']} not found in warehouse inventory", 'error')
                else:
                    flash(f"Pick operation failed: {error_msg}", 'error')
                
        except Exception as e:
            logger.error(f"Pick operation error: {e}")
            flash('Pick operation failed. Please try again.', 'error')
    else:
        if form.errors:
            logger.error(f"Pick form validation failed - Errors: {form.errors}")
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f"{field}: {error}", 'error')

    recent_picks = db_manager.get_recent_picks(limit=50)
    return render_template('pick.html', form=form, recent_picks=recent_picks)


@app.route('/api/reverse-pick/<int:transaction_id>', methods=['POST'])
@require_auth
def api_reverse_pick(transaction_id):
    """API endpoint to reverse a specific PICK transaction. Available to all users."""
    try:
        username = session.get('username', 'system')
        result = db_manager.reverse_pick(transaction_id=transaction_id, username=username)

        if result.get('success'):
            log_user_activity('REVERSE_PICK',
                f"Reversed PICK #{transaction_id}: {result['quantity']} units of {result['item']} (PCN {result['pcn']}) restored to {result['restored_location']}",
                f"WO: {result.get('original_work_order', 'N/A')}")

        return jsonify(result)
    except Exception as e:
        logger.error(f"Reverse pick error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/recent-picks')
@require_auth
def api_recent_picks():
    """API endpoint to get recent pick transactions."""
    try:
        picks = db_manager.get_recent_picks(limit=50)
        return jsonify({'success': True, 'picks': picks})
    except Exception as e:
        logger.error(f"Error fetching recent picks: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/restock', methods=['GET', 'POST'])
@require_auth
def restock():
    """Restock parts from Count Area to specified location."""
    form = RestockForm()

    # Set default source location on GET request (destination is left blank)
    if not form.location_from.data:
        form.location_from.data = 'Count Area'

    if form.validate_on_submit():
        logger.info(f"Restock form validation passed - PCN={form.pcn.data}, Item={form.item.data}, Quantity={form.quantity.data}, Location={form.location_to.data}")

        try:
            username = session.get('username', 'system')

            pcn_value = None
            if form.pcn.data:
                try:
                    pcn_value = int(form.pcn.data.strip())
                except (ValueError, AttributeError):
                    pcn_value = None

            # Default location_to to the part's current location if not specified
            loc_to_value = form.location_to.data.strip() if form.location_to.data and form.location_to.data.strip() else None

            result = db_manager.restock_pcb(
                pcn=pcn_value,
                item=form.item.data.strip() if form.item.data else None,
                quantity=form.quantity.data,
                location_from=form.location_from.data or 'Count Area',
                location_to=loc_to_value,
                username=username
            )
            logger.info(f"restock_pcb returned: {result}")

            if result.get('success'):
                location_to = result.get('location_to', '')
                log_user_activity('RESTOCK', f"Restocked {result['quantity']} units of {result['item']} (PCN: {result['pcn']}) to {location_to}", f"MFG Qty: {result['new_mfg_qty']}, On Hand: {result['new_onhand_qty']}")
                flash(f"Successfully restocked {result['quantity']} units of {result['item']} (PCN: {result['pcn']}) to {location_to}. "
                      f"MFG Qty: {result['new_mfg_qty']}, On Hand: {result['new_onhand_qty']}", 'success')
                # Pass PCN to show print label button
                return redirect(url_for('restock', restocked_pcn=result['pcn']))
            else:
                error_msg = result.get('error', 'Unknown error')
                flash(f"Restock operation failed: {error_msg}", 'error')

        except Exception as e:
            logger.error(f"Restock operation error: {e}")
            flash('Restock operation failed. Please try again.', 'error')
    else:
        if form.errors:
            logger.error(f"Restock form validation failed - Errors: {form.errors}")
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f"{field}: {error}", 'error')

    return render_template('restock.html', form=form)


@app.route('/api/restock', methods=['POST'])
@csrf.exempt
@require_auth
def api_restock():
    """AJAX endpoint for restock — returns JSON instead of redirect."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400

        pcn_raw = (data.get('pcn') or '').strip()
        item_raw = (data.get('item') or '').strip()
        quantity_raw = data.get('quantity')
        location_to_raw = (data.get('location_to') or '').strip()

        if not pcn_raw and not item_raw:
            return jsonify({'success': False, 'error': 'Either PCN or Item Number is required'}), 400

        try:
            quantity = int(quantity_raw)
            if quantity < 1:
                raise ValueError
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Please enter a valid quantity'}), 400

        pcn_value = None
        if pcn_raw:
            try:
                pcn_value = int(pcn_raw)
            except (ValueError, AttributeError):
                pcn_value = None

        username = session.get('username', 'system')

        result = db_manager.restock_pcb(
            pcn=pcn_value,
            item=item_raw or None,
            quantity=quantity,
            location_from=data.get('location_from') or 'Count Area',
            location_to=location_to_raw or None,
            username=username
        )

        if result.get('success'):
            location_to = result.get('location_to', '')
            log_user_activity(
                'RESTOCK',
                f"Restocked {result['quantity']} units of {result['item']} (PCN: {result['pcn']}) to {location_to}",
                f"MFG Qty: {result['new_mfg_qty']}, On Hand: {result['new_onhand_qty']}"
            )
            return jsonify({
                'success': True,
                'pcn': str(result['pcn']),
                'item': str(result['item']),
                'mpn': str(result.get('mpn', '')),
                'quantity': result['quantity'],
                'location_to': location_to,
                'new_mfg_qty': result['new_mfg_qty'],
                'new_onhand_qty': result['new_onhand_qty']
            })
        else:
            return jsonify({'success': False, 'error': result.get('error', 'Unknown error')}), 400

    except Exception as e:
        logger.error(f"API restock error: {e}")
        return jsonify({'success': False, 'error': 'Restock operation failed. Please try again.'}), 500


@app.route('/part-number-change', methods=['GET', 'POST'])
@require_auth
def part_number_change():
    """Change part number (item) for a PCN."""
    if request.method == 'POST':
        pcn = request.form.get('pcn', '').strip()
        new_part_number = request.form.get('new_part_number', '').strip()
        username = session.get('username', 'unknown')

        if not pcn or not new_part_number:
            flash('PCN and new part number are required.', 'danger')
            return render_template('part_number_change.html')

        conn = None
        try:
            conn = db_manager.get_connection()
            cursor = conn.cursor(cursor_factory=RealDictCursor)

            # Check if PCN exists
            cursor.execute('''
                SELECT pcn, item, mpn, onhandqty, loc_to
                FROM pcb_inventory."tblWhse_Inventory"
                WHERE pcn::text = %s
            ''', (pcn,))

            item = cursor.fetchone()

            if not item:
                flash(f'PCN {pcn} not found in inventory.', 'danger')
                return render_template('part_number_change.html')

            old_part_number = item['item']

            # Check if new part number is the same
            if old_part_number == new_part_number:
                flash(f'New part number is the same as current part number ({old_part_number}).', 'warning')
                return render_template('part_number_change.html', item=item)

            # Update part number in inventory
            cursor.execute('''
                UPDATE pcb_inventory."tblWhse_Inventory"
                SET item = %s
                WHERE pcn::text = %s
            ''', (new_part_number, pcn))

            # Check if update succeeded
            rows_updated = cursor.rowcount
            if rows_updated == 0:
                conn.rollback()
                flash(f'Failed to update PCN {pcn}. No rows were modified.', 'danger')
                return render_template('part_number_change.html')

            # Log the change in transaction table
            cursor.execute('''
                INSERT INTO pcb_inventory."tblTransaction"
                (trantype, item, pcn, mpn, tranqty, tran_time, loc_to, userid, created_at)
                VALUES (%s, %s, %s, %s, %s, TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', 'MM/DD/YY HH24:MI:SS'), %s, %s, CURRENT_TIMESTAMP)
            ''', ('PN_CHANGE', new_part_number, pcn, item['mpn'], 0, item['loc_to'], username))

            conn.commit()

            log_user_activity('PART_NUMBER_CHANGE', f"Changed part number for PCN {pcn}: '{old_part_number}' → '{new_part_number}'")
            logger.info(f"Part number changed by {username}: PCN {pcn} from '{old_part_number}' to '{new_part_number}'")
            flash(f'Successfully changed part number for PCN {pcn} from "{old_part_number}" to "{new_part_number}".', 'success')

            # Fetch updated item
            cursor.execute('''
                SELECT pcn, item, mpn, onhandqty, loc_to
                FROM pcb_inventory."tblWhse_Inventory"
                WHERE pcn::text = %s
            ''', (pcn,))
            updated_item = cursor.fetchone()

            return render_template('part_number_change.html', item=updated_item, show_print=True)

        except Exception as e:
            if conn:
                conn.rollback()
            logger.error(f"Error changing part number: {e}")
            flash(f'Error changing part number: {str(e)}', 'danger')
            return render_template('part_number_change.html')
        finally:
            if conn:
                db_manager.return_connection(conn)

    return render_template('part_number_change.html')

@app.route('/api/search-inventory', methods=['GET'])
@require_auth
def api_search_inventory():
    """API endpoint to search full inventory database with improved MPN matching."""
    description = request.args.get('description', '').strip()
    mpn = request.args.get('mpn', '').strip()
    location = request.args.get('location', '').strip()

    if not description and not mpn and not location:
        return jsonify({'success': False, 'results': []})

    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        logger.info(f"Search request - Description: '{description}', MPN: '{mpn}', Location: '{location}'")

        # Build dynamic query with better matching
        query = '''
            SELECT
                w.mpn as pcb_type,
                w.loc_to as location,
                COUNT(DISTINCT w.item) as job_count,
                SUM(w.onhandqty) as total_qty,
                AVG(w.onhandqty) as avg_qty,
                MAX(p."DESC") as description,
                MAX(w.pcn) as pcn,
                MAX(w.item) as sample_item,
                -- Exact match score for sorting
                CASE
                    WHEN LOWER(w.mpn) = LOWER(%s) THEN 1
                    ELSE 2
                END as match_priority
            FROM pcb_inventory."tblWhse_Inventory" w
            LEFT JOIN pcb_inventory."tblPN_List" p ON w.item = p.item
            WHERE w.onhandqty > 0
        '''
        params = [mpn if mpn else '']

        if description:
            query += ' AND (LOWER(p."DESC") LIKE %s OR LOWER(w.item) LIKE %s)'
            params.extend([f'%{description.lower()}%', f'%{description.lower()}%'])

        if mpn:
            # Search for exact match OR partial match (handles hyphens, spaces, case)
            query += ' AND (LOWER(w.mpn) = %s OR LOWER(w.mpn) LIKE %s OR LOWER(REPLACE(w.mpn, \'-\', \'\')) LIKE %s)'
            mpn_clean = mpn.lower().replace('-', '').replace(' ', '')
            params.extend([mpn.lower(), f'%{mpn.lower()}%', f'%{mpn_clean}%'])

        if location:
            query += ' AND LOWER(w.loc_to) LIKE %s'
            params.append(f'%{location.lower()}%')

        query += ' GROUP BY w.mpn, w.loc_to ORDER BY match_priority, total_qty DESC LIMIT 200'

        cursor.execute(query, params)
        results = [dict(row) for row in cursor.fetchall()]

        # Remove match_priority from results (internal use only)
        for result in results:
            result.pop('match_priority', None)

        logger.info(f"Search results: {len(results)} items found")
        if results and mpn:
            logger.info(f"First result for MPN '{mpn}': {results[0].get('pcb_type')} at {results[0].get('location')}")

        return jsonify({'success': True, 'results': results, 'count': len(results)})

    except Exception as e:
        logger.error(f"Error searching inventory: {e}")
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/api/get-part-details', methods=['GET'])
@require_auth
def get_part_details():
    """API endpoint to get part details for autofill in restock form."""
    pcn = request.args.get('pcn', '').strip()
    item = request.args.get('item', '').strip()

    if not pcn and not item:
        return jsonify({'success': False, 'error': 'PCN or Item number is required'})

    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Determine search criteria - use both PCN and item when available
        if pcn and item:
            where_clause = "pcn::text = %s AND item::text ILIKE %s"
            search_params = (str(pcn), item)
        elif pcn:
            where_clause = "pcn::text = %s"
            search_params = (str(pcn),)
        else:
            where_clause = "item = %s"
            search_params = (item,)

        # Fetch part details from warehouse inventory
        cursor.execute(f"""
            SELECT
                pcn,
                item,
                mpn,
                dc,
                CASE WHEN mfg_qty ~ '^\-?[0-9]+$' THEN mfg_qty::integer ELSE 0 END as mfg_qty,
                COALESCE(onhandqty, 0) as onhandqty,
                loc_from,
                loc_to,
                msd,
                po
            FROM pcb_inventory."tblWhse_Inventory"
            WHERE {where_clause}
            ORDER BY id DESC
            LIMIT 1
        """, search_params)

        result = cursor.fetchone()

        if result:
            mfg_qty_int = int(result['mfg_qty']) if result['mfg_qty'] else 0

            return jsonify({
                'success': True,
                'data': {
                    'pcn': str(result['pcn']),
                    'item': str(result['item']),
                    'mpn': str(result['mpn']) if result['mpn'] else '',
                    'dc': str(result['dc']) if result['dc'] else '',
                    'mfg_qty': mfg_qty_int,
                    'onhandqty': int(result['onhandqty']) if result['onhandqty'] else 0,
                    'location_from': str(result['loc_from']) if result['loc_from'] and result['loc_from'] != 'Stock' else '-',
                    'location_to': str(result['loc_to']) if result['loc_to'] else '-',
                    'msd': str(result['msd']) if result['msd'] else '',
                    'po': str(result['po']) if result['po'] else '',
                    'has_mfg_qty': mfg_qty_int > 0
                }
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Part not found for {"PCN " + pcn if pcn else "Item " + item}'
            })

    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid PCN format. Must be a number.'})
    except Exception as e:
        logger.error(f"Error fetching part details: {e}")
        return jsonify({'success': False, 'error': str(e)})
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/pcb-inventory')
@require_auth
def pcb_inventory():
    """PCB Inventory listing page with pagination and advanced filters."""
    # Get search and pagination parameters
    search_job = request.args.get('job', '').strip()
    search_pcb_type = request.args.get('pcb_type', '').strip()
    search_location = request.args.get('location', '').strip()
    search_pcn = request.args.get('pcn', '').strip()
    search_date_from = request.args.get('date_from', '').strip()
    search_date_to = request.args.get('date_to', '').strip()
    search_min_qty = request.args.get('min_qty', '').strip()
    search_max_qty = request.args.get('max_qty', '').strip()

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    sort_by = request.args.get('sort', 'job')
    sort_order = request.args.get('order', 'asc')

    # Limit per_page to reasonable values
    per_page = min(max(per_page, 10), 200)

    user_role = session.get('role', 'USER')
    itar_auth = session.get('itar_authorized', False)

    try:
        # Get all inventory first
        inventory_data = db_manager.get_current_inventory(user_role, itar_auth)

        # Apply filters
        if search_job:
            # Support comma-separated job numbers
            job_list = [j.strip() for j in search_job.split(',') if j.strip()]
            inventory_data = [item for item in inventory_data if item.get('job') in job_list]

        if search_pcb_type:
            inventory_data = [item for item in inventory_data if item.get('pcb_type') == search_pcb_type]

        if search_location:
            inventory_data = [item for item in inventory_data if item.get('location') == search_location]

        if search_pcn:
            inventory_data = [item for item in inventory_data if item.get('pcn') and search_pcn.lower() in item.get('pcn', '').lower()]

        # Date range filter
        if search_date_from:
            from datetime import datetime
            date_from = datetime.strptime(search_date_from, '%Y-%m-%d')
            inventory_data = [item for item in inventory_data
                            if item.get('updated_at') and item.get('updated_at').replace(tzinfo=None) >= date_from]

        if search_date_to:
            from datetime import datetime
            date_to = datetime.strptime(search_date_to, '%Y-%m-%d')
            date_to = date_to.replace(hour=23, minute=59, second=59)
            inventory_data = [item for item in inventory_data
                            if item.get('updated_at') and item.get('updated_at').replace(tzinfo=None) <= date_to]

        # Quantity range filter
        if search_min_qty:
            try:
                min_qty = int(search_min_qty)
                inventory_data = [item for item in inventory_data if (item.get('qty') or 0) >= min_qty]
            except ValueError:
                pass

        if search_max_qty:
            try:
                max_qty = int(search_max_qty)
                inventory_data = [item for item in inventory_data if (item.get('qty') or 0) <= max_qty]
            except ValueError:
                pass

        # Sort the data - handle None values properly
        reverse_sort = sort_order == 'desc'
        if sort_by == 'job':
            inventory_data.sort(key=lambda x: (x.get('job') or ''), reverse=reverse_sort)
        elif sort_by == 'pcb_type':
            inventory_data.sort(key=lambda x: (x.get('pcb_type') or ''), reverse=reverse_sort)
        elif sort_by == 'qty':
            inventory_data.sort(key=lambda x: (x.get('qty') or 0), reverse=reverse_sort)
        elif sort_by == 'location':
            inventory_data.sort(key=lambda x: (x.get('location') or ''), reverse=reverse_sort)
        elif sort_by == 'updated_at':
            inventory_data.sort(key=lambda x: (x.get('updated_at') or ''), reverse=reverse_sort)

        # Get unique locations for dropdown
        all_inventory = db_manager.get_current_inventory(user_role, itar_auth)
        locations = sorted(list(set(item.get('location') for item in all_inventory if item.get('location'))))

        # Calculate pagination
        total_items = len(inventory_data)
        total_pages = (total_items + per_page - 1) // per_page
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page

        paginated_inventory = inventory_data[start_idx:end_idx]

        # Calculate pagination info
        pagination = {
            'page': page,
            'per_page': per_page,
            'total': total_items,
            'total_pages': total_pages,
            'has_prev': page > 1,
            'has_next': page < total_pages,
            'prev_num': page - 1 if page > 1 else None,
            'next_num': page + 1 if page < total_pages else None,
            'pages': list(range(max(1, page - 2), min(total_pages + 1, page + 3)))
        }

        return render_template('inventory.html',
                             inventory=paginated_inventory,
                             pagination=pagination,
                             pcb_types=PCB_TYPES,
                             locations=locations,
                             search_job=search_job,
                             search_pcb_type=search_pcb_type,
                             search_location=search_location,
                             search_pcn=search_pcn,
                             search_date_from=search_date_from,
                             search_date_to=search_date_to,
                             search_min_qty=search_min_qty,
                             search_max_qty=search_max_qty,
                             sort_by=sort_by,
                             sort_order=sort_order)
    except Exception as e:
        import traceback
        logger.error(f"Error loading inventory: {e}")
        logger.error(traceback.format_exc())
        flash('Error loading inventory. Please try again.', 'error')
        return render_template('inventory.html', inventory=[], pagination={'total': 0},
                             pcb_types=PCB_TYPES, locations=[], search_job='', search_pcb_type='',
                             search_location='', search_pcn='', search_date_from='', search_date_to='',
                             search_min_qty='', search_max_qty='', sort_by='job', sort_order='asc')

@app.route('/warehouse-inventory')
@require_auth
def warehouse_inventory():
    """Warehouse Inventory listing page - reads from PostgreSQL database."""
    conn = None
    cursor = None
    try:
        # Get search parameters
        search_item = request.args.get('search_item', '').strip()
        search_pcn = request.args.get('search_pcn', '').strip()
        search_mpn = request.args.get('search_mpn', '').strip()
        search_location = request.args.get('search_location', '').strip()

        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        per_page = min(max(per_page, 10), 200)

        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Build query with filters
        query = """
            SELECT id, item, pcn, mpn, dc, onhandqty, loc_from, loc_to,
                   mfg_qty, qty_old, msd, po, cost, vendor, migrated_at
            FROM pcb_inventory."tblWhse_Inventory"
            WHERE 1=1
        """
        params = []

        if search_item:
            query += " AND LOWER(item::text) LIKE %s"
            params.append(f"%{search_item.lower()}%")

        if search_pcn:
            query += " AND pcn::text LIKE %s"
            params.append(f"{search_pcn}%")

        if search_mpn:
            query += " AND LOWER(mpn::text) LIKE %s"
            params.append(f"%{search_mpn.lower()}%")

        if search_location:
            query += " AND LOWER(loc_to::text) LIKE %s"
            params.append(f"%{search_location.lower()}%")

        # Get total count for pagination
        count_query = f"SELECT COUNT(*) as total FROM ({query}) AS filtered"
        cursor.execute(count_query, params)
        total_records = cursor.fetchone()['total']

        # Add sorting and pagination (newest entries first for efficiency)
        query += " ORDER BY id DESC LIMIT %s OFFSET %s"
        params.extend([per_page, (page - 1) * per_page])

        # Execute main query
        cursor.execute(query, params)
        rows = cursor.fetchall()

        # Convert to list of dicts with consistent naming (matching .mdb format)
        inventory = []
        for row in rows:
            inventory.append({
                'PCN': row['pcn'],
                'Item': row['item'],
                'MPN': row['mpn'],
                'DC': row['dc'],
                'OnHandQty': row['onhandqty'],
                'Loc_To': row['loc_to'],
                'MFG_Qty': row['mfg_qty'],
                'Qty_Old': row['qty_old'],
                'MSD': row['msd'],
                'PO': row['po'],
                'Cost': row['cost'],
                'Vendor': row['vendor'],
                'Loc_From': row['loc_from']
            })

        # Calculate pagination
        total_pages = (total_records + per_page - 1) // per_page if total_records > 0 else 1

        pagination = {
            'page': page,
            'per_page': per_page,
            'total': total_records,
            'total_pages': total_pages,
            'has_prev': page > 1,
            'has_next': page < total_pages,
            'prev_num': page - 1 if page > 1 else None,
            'next_num': page + 1 if page < total_pages else None,
            'pages': list(range(max(1, page - 2), min(total_pages + 1, page + 3)))
        }

        return render_template('warehouse_inventory.html',
                             inventory=inventory,
                             pagination=pagination,
                             total_records=total_records,
                             search_item=search_item,
                             search_pcn=search_pcn,
                             search_mpn=search_mpn,
                             search_location=search_location)

    except Exception as e:
        logger.error(f"Error loading warehouse inventory: {e}")
        flash('Error loading warehouse inventory. Please try again.', 'error')
        return render_template('warehouse_inventory.html', inventory=[],
                             pagination={'total': 0, 'page': 1, 'total_pages': 1, 'per_page': 10},
                             total_records=0)
    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass
        if conn:
            try:
                db_manager.return_connection(conn)
            except Exception:
                pass

@app.route('/api/warehouse-inventory/item')
@require_auth
def get_warehouse_item():
    """API endpoint to get a single warehouse inventory item."""
    try:
        item_id = request.args.get('item', '').strip()
        pcn = request.args.get('pcn', '').strip()

        if not item_id or not pcn:
            return jsonify({'success': False, 'message': 'Item and PCN are required'}), 400

        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        try:
            # Query for specific item
            cursor.execute("""
                SELECT id, item, pcn, mpn, dc, onhandqty, loc_from, loc_to,
                       mfg_qty, qty_old, msd, po, cost
                FROM pcb_inventory."tblWhse_Inventory"
                WHERE item::text = %s AND pcn::text = %s
                LIMIT 1
            """, (item_id, pcn))

            row = cursor.fetchone()

            if row:
                # Convert to dict with consistent naming (matching .mdb format)
                item_data = {
                    'PCN': row['pcn'],
                    'Item': row['item'],
                    'MPN': row['mpn'],
                    'DC': row['dc'],
                    'OnHandQty': row['onhandqty'],
                    'Loc_From': row['loc_from'],
                    'Loc_To': row['loc_to'],
                    'MFG_Qty': row['mfg_qty'],
                    'Qty_Old': row['qty_old'],
                    'MSD': row['msd'],
                    'PO': row['po'],
                    'Cost': row['cost']
                }
                return jsonify({'success': True, 'item': item_data})
            else:
                return jsonify({'success': False, 'message': 'Item not found'}), 404

        finally:


            if cursor:


                cursor.close()


            if conn:


                db_manager.return_connection(conn)

    except Exception as e:
        logger.error(f"Error fetching warehouse item: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/warehouse-inventory/recent')
@require_auth
def get_recent_warehouse_inventory():
    """API endpoint to get recent warehouse inventory items for stock page."""
    try:
        limit = request.args.get('limit', 10, type=int)

        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        try:
            # Get recent warehouse inventory items with quantity > 0
            cursor.execute("""
                SELECT id, item, pcn, mpn, dc, onhandqty, loc_to as location,
                       msd, po, migrated_at as updated_at
                FROM pcb_inventory."tblWhse_Inventory"
                WHERE onhandqty > 0
                ORDER BY id DESC
                LIMIT %s
            """, (limit,))

            rows = cursor.fetchall()

            data = []
            for row in rows:
                data.append({
                    'id': row['id'],
                    'item': row['item'],
                    'pcn': row['pcn'],
                    'mpn': row['mpn'],
                    'dc': row['dc'],
                    'onhandqty': row['onhandqty'],
                    'location': row['location'],
                    'msd': row['msd'],
                    'po': row['po'],
                    'updated_at': row['updated_at'].isoformat() if row['updated_at'] else None
                })

            return jsonify({'success': True, 'data': data})

        finally:
            if cursor:
                cursor.close()
            if conn:
                db_manager.return_connection(conn)

    except Exception as e:
        logger.error(f"Error fetching recent warehouse inventory: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/locations', methods=['GET'])
@require_auth
def get_locations():
    """API endpoint to fetch valid locations from tblLoc for autocomplete/dropdown."""
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT location, area, shelf, loc
            FROM pcb_inventory."tblLoc"
            ORDER BY location
        """)
        rows = cursor.fetchall()
        cursor.close()
        db_manager.return_connection(conn)

        locations = [{'location': r[0], 'area': r[1], 'shelf': r[2], 'loc': r[3]} for r in rows]
        return jsonify({'success': True, 'locations': locations, 'count': len(locations)})
    except Exception as e:
        logger.error(f"Error fetching locations: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/warehouse-inventory/update', methods=['POST'])
@require_auth
def update_warehouse_item():
    """API endpoint to update warehouse inventory item."""
    try:
        data = request.get_json()

        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400

        # Validate required fields
        required_fields = ['item', 'pcn']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'{field} is required'}), 400

        conn = db_manager.get_connection()
        conn.autocommit = False
        cursor = conn.cursor()

        try:
            # Helper function to convert empty strings to None for numeric fields
            def to_int_or_none(value):
                if value == '' or value is None:
                    return None
                try:
                    return int(value)
                except (ValueError, TypeError):
                    return None

            def to_float_or_none(value):
                if value == '' or value is None:
                    return None
                try:
                    return float(value)
                except (ValueError, TypeError):
                    return None

            # Validate location fields - must be 7 digits or standard location
            loc_from_val = data.get('loc_from', '').strip() if data.get('loc_from') else ''
            loc_to_val = data.get('loc_to', '').strip() if data.get('loc_to') else ''

            if loc_to_val and not validate_location(loc_to_val):
                return jsonify({'success': False, 'message': 'Location To must be exactly 7 digits (e.g. 1101101) or a standard location.'}), 400

            if loc_from_val and not validate_location(loc_from_val):
                return jsonify({'success': False, 'message': 'Location From must be exactly 7 digits (e.g. 1101101) or a standard location.'}), 400

            # Validate quantities are not negative
            onhand_qty = to_int_or_none(data.get('onhandqty'))
            mfg_qty = to_int_or_none(data.get('mfg_qty'))

            if onhand_qty is not None and onhand_qty < 0:
                return jsonify({'success': False, 'message': 'On-hand quantity cannot be negative'}), 400

            if mfg_qty is not None and mfg_qty < 0:
                return jsonify({'success': False, 'message': 'MFG quantity cannot be negative'}), 400

            # Lock row before updating to prevent concurrent overwrites
            cursor.execute("""
                SELECT id FROM pcb_inventory."tblWhse_Inventory"
                WHERE item::text = %s AND pcn::text = %s
                  AND (mpn::text = %s OR (mpn IS NULL AND %s IS NULL))
                FOR UPDATE
            """, (data.get('item'), data.get('pcn'), data.get('mpn') or None, data.get('mpn') or None))

            if not cursor.fetchone():
                conn.rollback()
                return jsonify({'success': False, 'message': 'Item not found'}), 404

            # Update warehouse inventory record
            cursor.execute("""
                UPDATE pcb_inventory."tblWhse_Inventory"
                SET dc = %s,
                    onhandqty = %s,
                    loc_from = %s,
                    loc_to = %s,
                    mfg_qty = %s,
                    msd = %s,
                    po = %s,
                    cost = %s
                WHERE item::text = %s AND pcn::text = %s
                  AND (mpn::text = %s OR (mpn IS NULL AND %s IS NULL))
            """, (
                data.get('dc') or None,
                onhand_qty,
                data.get('loc_from') or None,
                data.get('loc_to') or None,
                mfg_qty,
                data.get('msd') or None,
                data.get('po') or None,
                to_float_or_none(data.get('cost')),
                data.get('item'),
                data.get('pcn'),
                data.get('mpn') or None,
                data.get('mpn') or None
            ))

            if cursor.rowcount == 0:
                conn.rollback()
                return jsonify({'success': False, 'message': 'Item not found'}), 404

            conn.commit()
            logger.info(f"Updated warehouse inventory item: {data.get('item')}, PCN: {data.get('pcn')}")

            return jsonify({
                'success': True,
                'message': 'Item updated successfully'
            })

        except Exception as e:


            if conn:


                conn.rollback()
            logger.error(f"Database error updating warehouse item: {e}")
            return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500
        finally:

            if cursor:

                cursor.close()

            if conn:

                db_manager.return_connection(conn)

    except Exception as e:
        logger.error(f"Error updating warehouse item: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/reports')
@require_auth
def reports():
    """Reports page."""
    try:
        # Get current inventory data for reports
        user_role = session.get('role', 'USER')
        itar_auth = session.get('itar_authorized', False)
        inventory = db_manager.get_current_inventory(user_role, itar_auth)

        # Create summary data matching template expectations
        summary = []
        location_type_summary = {}

        # Group by location and PCB type
        for item in inventory:
            location = item.get('location', 'Unknown')
            pcb_type = item.get('pcb_type', 'Unknown')
            key = f"{location}|{pcb_type}"

            if key not in location_type_summary:
                location_type_summary[key] = {
                    'location': location,
                    'pcb_type': pcb_type,
                    'job_count': 0,
                    'total_quantity': 0,
                    'jobs': set()
                }

            location_type_summary[key]['total_quantity'] += item.get('qty', 0)
            if item.get('job'):
                location_type_summary[key]['jobs'].add(item.get('job'))

        # Convert to list format expected by template
        total_all_qty = sum(item.get('qty', 0) for item in inventory)
        for data in location_type_summary.values():
            data['job_count'] = len(data['jobs'])
            data['average_quantity'] = data['total_quantity'] / max(data['job_count'], 1)
            data['percentage'] = (data['total_quantity'] / max(total_all_qty, 1)) * 100
            del data['jobs']  # Remove set object
            summary.append(data)

        # Sort by total quantity descending
        summary.sort(key=lambda x: x['total_quantity'], reverse=True)

        # Get audit log
        audit_log = db_manager.get_audit_log(100)

        return render_template('reports.html',
                             summary=summary,
                             audit_log=audit_log)
    except Exception as e:
        logger.error(f"Error loading reports: {e}")
        flash('Error loading reports. Please try again.', 'error')
        return render_template('reports.html', summary=[], audit_log=[])

# ============================================================================
# SHORTAGE REPORT ROUTES
# ============================================================================

@app.route('/shortage_report')
@require_auth
def shortage_report():
    """Shortage Report page - list saved reports and generate new ones."""
    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Get list of saved reports
        cursor.execute("""
            SELECT id, job, report_name, total_lines, shortage_lines,
                   total_cost, shortage_cost, created_by, created_at, notes, order_qty
            FROM pcb_inventory."tblShortageReport"
            ORDER BY created_at DESC
            LIMIT 50
        """)
        saved_reports = cursor.fetchall()

        # Get list of jobs that have BOMs loaded
        cursor.execute("""
            SELECT DISTINCT job FROM pcb_inventory."tblBOM"
            WHERE job IS NOT NULL AND job != ''
            ORDER BY job
        """)
        available_jobs = [row['job'] for row in cursor.fetchall()]

        return render_template('shortage_report.html',
                             saved_reports=saved_reports,
                             available_jobs=available_jobs,
                             column_definitions=SHORTAGE_EXPORT_COLUMNS)
    except Exception as e:
        logger.error(f"Error loading shortage report page: {e}")
        flash('Error loading page. Please try again.', 'error')
        return render_template('shortage_report.html', saved_reports=[], available_jobs=[],
                             column_definitions=SHORTAGE_EXPORT_COLUMNS)
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/shortage_report/generate', methods=['POST'])
@require_auth
def generate_shortage_report():
    """Generate a new shortage report for a job."""
    job = request.form.get('job', '').strip()
    report_name = request.form.get('report_name', '').strip()
    notes = request.form.get('notes', '').strip()
    order_qty_input = request.form.get('order_qty', '1').strip()

    if not job:
        flash('Please enter a job number.', 'danger')
        return redirect(url_for('shortage_report'))

    try:
        order_qty = int(order_qty_input)
        if order_qty < 1:
            raise ValueError
    except (ValueError, TypeError):
        flash('Order Qty must be a positive whole number.', 'danger')
        return redirect(url_for('shortage_report'))

    if not report_name:
        report_name = f"Shortage Report - {job} - {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Check if job has BOM data
        cursor.execute("SELECT COUNT(*) as count FROM pcb_inventory.\"tblBOM\" WHERE job = %s", (job,))
        if cursor.fetchone()['count'] == 0:
            flash(f'No BOM data found for job {job}. Please load BOM first.', 'warning')
            return redirect(url_for('shortage_report'))

        # Get total BOM line count
        cursor.execute("SELECT COUNT(*) as count FROM pcb_inventory.\"tblBOM\" WHERE job = %s", (job,))
        total_bom_lines = cursor.fetchone()['count']

        # Get the latest revision for this job
        cursor.execute("""
            SELECT job_rev FROM pcb_inventory."tblBOM"
            WHERE job = %s AND job_rev IS NOT NULL AND job_rev != ''
            ORDER BY created_at DESC LIMIT 1
        """, (job,))
        rev_row = cursor.fetchone()
        job_rev = rev_row['job_rev'] if rev_row else None

        # Generate report data by comparing BOM vs Inventory
        # First deduplicate BOM lines per aci_pn (alternate parts), then match inventory
        # Uses warehouse MPN (w.mpn) so each PCN shows its actual MPN, not BOM alternates
        cursor.execute("""
            WITH bom_lines AS (
                SELECT DISTINCT ON (b.aci_pn)
                    b.line,
                    b.aci_pn,
                    b.mpn as bom_mpn,
                    b.man,
                    b."DESC",
                    b.qty,
                    b.cost
                FROM pcb_inventory."tblBOM" b
                WHERE b.job = %s
                    AND (b.job_rev = (SELECT job_rev FROM pcb_inventory."tblBOM" WHERE job = %s AND job_rev IS NOT NULL AND job_rev != '' ORDER BY created_at DESC LIMIT 1)
                         OR NOT EXISTS (SELECT 1 FROM pcb_inventory."tblBOM" WHERE job = %s AND job_rev IS NOT NULL AND job_rev != ''))
                ORDER BY b.aci_pn, b.line
            ),
            inventory_match AS (
                SELECT DISTINCT ON (COALESCE(w.pcn, bl.aci_pn || '_nopcn'), bl.aci_pn)
                    bl.line,
                    bl.aci_pn,
                    COALESCE(w.mpn, bl.bom_mpn) as mpn,
                    bl.man,
                    bl."DESC",
                    bl.qty,
                    bl.cost,
                    w.pcn,
                    COALESCE(w.item, bl.aci_pn) as item,
                    COALESCE(w.onhandqty, 0) as onhandqty,
                    w.loc_to,
                    CASE WHEN bl.aci_pn = w.item THEN 1 WHEN w.item IS NOT NULL THEN 2 ELSE 3 END as match_priority
                FROM bom_lines bl
                LEFT JOIN pcb_inventory."tblWhse_Inventory" w
                    ON (bl.aci_pn = w.item OR bl.bom_mpn = w.mpn)
                    AND COALESCE(w.loc_to, '') != 'MFG Floor'
                ORDER BY COALESCE(w.pcn, bl.aci_pn || '_nopcn'), bl.aci_pn, match_priority
            )
            SELECT
                line as line_no,
                aci_pn,
                pcn,
                mpn,
                CAST(COALESCE(NULLIF(qty, ''), '0') AS INTEGER) as qty,
                COALESCE(SUM(onhandqty), 0) as qty_on_hand,
                item,
                COALESCE(loc_to, '') as location,
                CAST(COALESCE(NULLIF(cost, ''), '0') AS DECIMAL(10,4)) as unit_cost,
                man as manufacturer,
                "DESC" as description
            FROM inventory_match
            GROUP BY line, aci_pn, mpn, man, "DESC", qty, cost, pcn, item, loc_to
            ORDER BY
                CASE WHEN line ~ '^[0-9]+$' THEN CAST(line AS INTEGER) ELSE 999999 END,
                line
        """, (job, job, job))
        matched_items = cursor.fetchall()

        if not matched_items:
            flash(f'No BOM data found for job {job}.', 'warning')
            return redirect(url_for('shortage_report'))

        # Calculate REQ for each item and only keep actual shortages (on_hand < req)
        report_items = []
        shortage_count = 0
        for item in matched_items:
            qty = int(item['qty'] or 0)
            req = qty * order_qty
            on_hand = int(item['qty_on_hand'] or 0)
            item['req'] = req
            item['order_qty'] = order_qty
            if on_hand < req:
                shortage_count += 1
                report_items.append(item)

        # Calculate costs
        total_cost = sum(
            float(item['qty'] or 0) * float(item['unit_cost'] or 0)
            for item in report_items
        )
        shortage_cost = sum(
            float(item['req'] or 0) * float(item['unit_cost'] or 0)
            for item in report_items if item['qty_on_hand'] < item['req']
        )

        # Create the report header
        username = session.get('username', 'Unknown')
        cursor.execute("""
            INSERT INTO pcb_inventory."tblShortageReport"
            (job, report_name, total_lines, shortage_lines, total_cost, shortage_cost, created_by, notes, order_qty, job_rev)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (job, report_name, total_bom_lines, shortage_count, total_cost, shortage_cost, username, notes, order_qty, job_rev))
        report_id = cursor.fetchone()['id']

        # Insert all BOM line items (including those with no inventory match)
        for item in report_items:
            cursor.execute("""
                INSERT INTO pcb_inventory."tblShortageReportItems"
                (report_id, line_no, aci_pn, pcn, mpn, qty_required, qty_on_hand, order_qty,
                 item, location, unit_cost, line_cost, manufacturer, description, req)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                report_id, item['line_no'], item['aci_pn'], item['pcn'], item['mpn'],
                item['qty'], item['qty_on_hand'], order_qty,
                item['item'], item['location'], float(item['unit_cost'] or 0),
                float(item['qty'] or 0) * float(item['unit_cost'] or 0),
                item['manufacturer'], item['description'], item['req']
            ))

        conn.commit()
        log_user_activity('SHORTAGE_REPORT', f"Generated shortage report for job {job}", f"{len(report_items)} items, {shortage_count} shortages")
        flash(f'Shortage report generated! {len(report_items)} items matched inventory, {shortage_count} have shortages.', 'success')
        return redirect(url_for('view_shortage_report', report_id=report_id))

    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error generating shortage report: {e}")
        flash('Error generating report. Please try again.', 'danger')
        return redirect(url_for('shortage_report'))
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/shortage_report/view/<int:report_id>')
@require_auth
def view_shortage_report(report_id):
    """View a saved shortage report."""
    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Get report header
        cursor.execute("""
            SELECT id, job, report_name, total_lines, shortage_lines,
                   total_cost, shortage_cost, created_by, created_at, notes, order_qty, job_rev
            FROM pcb_inventory."tblShortageReport"
            WHERE id = %s
        """, (report_id,))
        report = cursor.fetchone()

        if not report:
            flash('Report not found.', 'danger')
            return redirect(url_for('shortage_report'))

        # Get report line items (exclude MFG Floor items)
        cursor.execute("""
            SELECT line_no, aci_pn, pcn, mpn, qty_required, qty_on_hand, order_qty,
                   item, location, unit_cost, line_cost, manufacturer, description, req
            FROM pcb_inventory."tblShortageReportItems"
            WHERE report_id = %s AND COALESCE(location, '') != 'MFG Floor'
            ORDER BY
                CASE WHEN line_no ~ '^[0-9]+$' THEN CAST(line_no AS INTEGER) ELSE 999999 END,
                line_no
        """, (report_id,))
        items = cursor.fetchall()

        return render_template('shortage_report_view.html', report=report, items=items,
                                       column_definitions=SHORTAGE_EXPORT_COLUMNS)

    except Exception as e:
        logger.error(f"Error viewing shortage report: {e}")
        flash('Error loading report. Please try again.', 'danger')
        return redirect(url_for('shortage_report'))
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/shortage_report/export/<int:report_id>', methods=['GET', 'POST'])
@require_auth
def export_shortage_report(report_id):
    """Export shortage report to Excel with optional column customization."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    conn = None
    try:
        # Parse column config from POST body, or use defaults
        if request.method == 'POST' and request.is_json:
            config = request.get_json()
            selected_columns = config.get('columns', [])
            highlighted_columns = set(config.get('highlighted', []))
            export_filter = config.get('filter', 'all')
        else:
            selected_columns = [c['key'] for c in SHORTAGE_EXPORT_COLUMNS if c['default']]
            highlighted_columns = set()
            export_filter = 'all'

        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Get report header
        cursor.execute("""
            SELECT job, report_name, total_lines, shortage_lines,
                   total_cost, shortage_cost, created_by, created_at, order_qty, job_rev
            FROM pcb_inventory."tblShortageReport"
            WHERE id = %s
        """, (report_id,))
        report = cursor.fetchone()

        if not report:
            return jsonify({'success': False, 'error': 'Report not found.'}), 404

        # Get report line items (exclude MFG Floor)
        cursor.execute("""
            SELECT line_no, aci_pn, pcn, mpn, qty_required as qty, order_qty,
                   req, item, qty_on_hand, location,
                   unit_cost, line_cost, manufacturer, description
            FROM pcb_inventory."tblShortageReportItems"
            WHERE report_id = %s AND COALESCE(location, '') != 'MFG Floor'
            ORDER BY
                CASE WHEN line_no ~ '^[0-9]+$' THEN CAST(line_no AS INTEGER) ELSE 999999 END,
                line_no
        """, (report_id,))
        items = cursor.fetchall()

        # Apply filter: shortages only
        if export_filter == 'shortages_only':
            items = [i for i in items if (i.get('qty_on_hand') or 0) < (i.get('req') or 0)]

        # Hide zero on-hand rows (default: true, matches UI toggle)
        hide_zero = True
        if request.method == 'POST' and request.is_json:
            hide_zero = config.get('hide_zero', True)
        if hide_zero:
            items = [i for i in items if (i.get('qty_on_hand') or 0) != 0]

        # Build active column list from selection
        col_registry = {c['key']: c for c in SHORTAGE_EXPORT_COLUMNS}
        active_cols = [col_registry[k] for k in selected_columns if k in col_registry]
        if not active_cols:
            active_cols = [col_registry[k] for k in [c['key'] for c in SHORTAGE_EXPORT_COLUMNS if c['default']]]
        num_cols = len(active_cols)
        last_col = get_column_letter(num_cols)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Shortage Report"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        shortage_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        highlight_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Title section
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'] = f"Shortage Report - Job: {report['job']}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'A2:{last_col}2')
        ws['A2'] = f"Generated: {report['created_at'].strftime('%Y-%m-%d %H:%M') if report['created_at'] else 'N/A'} by {report['created_by']} | Rev: {report.get('job_rev', 'N/A')} | Order Qty: {report.get('order_qty', 'N/A')}"
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'A3:{last_col}3')
        ws['A3'] = f"Shortage Items: {report['shortage_lines']} of {report['total_lines']} BOM lines"
        ws['A3'].alignment = Alignment(horizontal='center')

        # Headers (row 5)
        for col_idx, col_def in enumerate(active_cols, 1):
            cell = ws.cell(row=5, column=col_idx, value=col_def['label'])
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, item in enumerate(items, 6):
            is_shortage = (item.get('qty_on_hand') or 0) < (item.get('req') or 0)
            for col_idx, col_def in enumerate(active_cols, 1):
                value = get_export_cell_value(item, col_def['key'])
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if is_shortage:
                    cell.fill = shortage_fill
                elif col_def['key'] in highlighted_columns:
                    cell.fill = highlight_fill

        # Column widths
        for col_idx, col_def in enumerate(active_cols, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def['width']

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        filename = f"Shortage_Report_{report['job']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        logger.error(f"Error exporting shortage report: {e}")
        if request.method == 'POST':
            return jsonify({'success': False, 'error': str(e)}), 500
        flash('Error exporting report. Please try again.', 'danger')
        return redirect(url_for('view_shortage_report', report_id=report_id))
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/shortage_report/delete/<int:report_id>', methods=['POST'])
@require_auth
def delete_shortage_report(report_id):
    """Delete a saved shortage report."""
    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor()

        # Delete report (cascade will delete items)
        cursor.execute('DELETE FROM pcb_inventory."tblShortageReport" WHERE id = %s', (report_id,))
        conn.commit()

        flash('Report deleted successfully.', 'success')
        return redirect(url_for('shortage_report'))

    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error deleting shortage report: {e}")
        flash('Error deleting report. Please try again.', 'danger')
        return redirect(url_for('shortage_report'))
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/api/bom/jobs')
@require_auth
def get_bom_jobs():
    """API endpoint to get list of jobs with BOMs."""
    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        cursor.execute("""
            SELECT DISTINCT job FROM pcb_inventory."tblBOM"
            WHERE job IS NOT NULL AND job != ''
            ORDER BY job
        """)
        jobs = [row['job'] for row in cursor.fetchall()]

        return jsonify({'success': True, 'jobs': jobs})

    except Exception as e:
        logger.error(f"Error getting BOM jobs: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)

# ============================================================================
# END SHORTAGE REPORT ROUTES
# ============================================================================

@app.route('/sources')
@require_auth
def sources():
    """Sources page - shows all migrated Access tables (super users only)."""
    user_role = session.get('role', 'USER')
    
    # Only super users can access sources
    if user_role != 'ADMIN':
        flash('Access denied: Super user privileges required', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Get list of all migrated tables
        conn = psycopg2.connect(
            host='aci-database',
            port=5432,
            database='pcb_inventory',
            user='stockpick_user',
            password='stockpick_pass'
        )
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get all tables in the pcb_inventory schema
        cursor.execute("""
            SELECT table_name, 
                   (SELECT COUNT(*) FROM pcb_inventory."" || table_name || "") as record_count
            FROM information_schema.tables 
            WHERE table_schema = 'pcb_inventory' 
            AND table_type = 'BASE TABLE'
            AND table_name NOT IN ('inventory_audit')
            ORDER BY table_name
        """)
        
        # Alternative approach - get tables manually
        cursor.execute("""
            SELECT schemaname, tablename 
            FROM pg_tables 
            WHERE schemaname = 'pcb_inventory'
            AND tablename NOT IN ('inventory_audit')
            ORDER BY tablename
        """)
        
        table_info = []
        for row in cursor.fetchall():
            table_name = row['tablename']
            try:
                # Get record count for each table
                if '"' in table_name:
                    count_sql = f'SELECT COUNT(*) as count FROM pcb_inventory.{table_name}'
                else:
                    count_sql = f'SELECT COUNT(*) as count FROM pcb_inventory."{table_name}"'
                
                cursor.execute(count_sql)
                count_result = cursor.fetchone()
                record_count = count_result['count'] if count_result else 0
                
                # Get column info
                cursor.execute(f"""
                    SELECT column_name, data_type 
                    FROM information_schema.columns 
                    WHERE table_schema = 'pcb_inventory' 
                    AND table_name = '{table_name}'
                    AND column_name NOT IN ('id', 'created_at')
                    ORDER BY ordinal_position
                """)
                columns = cursor.fetchall()
                
                table_info.append({
                    'name': table_name,
                    'record_count': record_count,
                    'column_count': len(columns),
                    'columns': [col['column_name'] for col in columns[:5]]  # Show first 5 columns
                })
                
            except Exception as e:
                logger.error(f"Error getting info for table {table_name}: {e}")
                table_info.append({
                    'name': table_name,
                    'record_count': 0,
                    'column_count': 0,
                    'columns': []
                })
        
        cursor.close()
        db_manager.return_connection(conn)

        return render_template('sources.html', tables=table_info)
        
    except Exception as e:
        logger.error(f"Error loading sources: {e}")
        flash('Error loading sources. Please try again.', 'error')
        return render_template('sources.html', tables=[])

@app.route('/sources/<table_name>')
@require_auth
def view_source_table(table_name):
    """View data from a specific source table."""
    user_role = session.get('role', 'USER')
    
    # Only super users can access sources
    if user_role != 'ADMIN':
        flash('Access denied: Super user privileges required', 'error')
        return redirect(url_for('dashboard'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 25
    
    try:
        conn = psycopg2.connect(
            host='aci-database',
            port=5432,
            database='pcb_inventory',
            user='stockpick_user',
            password='stockpick_pass'
        )
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # Get total count
        count_sql = f'SELECT COUNT(*) as count FROM pcb_inventory."{table_name}"'
        cursor.execute(count_sql)
        total_records = cursor.fetchone()['count']
        
        # Get paginated data
        offset = (page - 1) * per_page
        data_sql = f'SELECT * FROM pcb_inventory."{table_name}" ORDER BY id LIMIT {per_page} OFFSET {offset}'
        cursor.execute(data_sql)
        records = cursor.fetchall()
        
        # Get column names
        if records:
            columns = [col for col in records[0].keys() if col not in ['id', 'created_at']]
        else:
            columns = []
        
        # Calculate pagination
        total_pages = (total_records + per_page - 1) // per_page
        pagination = {
            'page': page,
            'per_page': per_page,
            'total': total_records,
            'total_pages': total_pages,
            'has_prev': page > 1,
            'has_next': page < total_pages,
            'prev_num': page - 1 if page > 1 else None,
            'next_num': page + 1 if page < total_pages else None,
        }
        
        cursor.close()
        db_manager.return_connection(conn)

        return render_template('source_table.html',
                             table_name=table_name,
                             records=records,
                             columns=columns,
                             pagination=pagination)
        
    except Exception as e:
        logger.error(f"Error viewing table {table_name}: {e}")
        flash('Error viewing table. Please try again.', 'error')
        return redirect(url_for('sources'))

@app.route('/stats')
@require_auth  
def stats():
    """Data migration statistics and comparison page."""
    try:
        # Get current PostgreSQL statistics
        postgres_stats = db_manager.get_stats_summary()
        
        # Source database statistics (actual Access database data)
        source_stats = {
            'total_records': 836,  # Actual records in Access tblPCB_Inventory table  
            'unique_jobs': 750,    # Actual unique jobs from migration
            'total_quantity': 211679,  # Actual total PCBs from migration
            'pcb_types': 3,        # Actual PCB types found (Bare, Partial, Completed)
            'migration_date': 'August 19, 2025'
        }
        
        # Calculate integrity check
        integrity_check = {
            'records_match': abs(postgres_stats['total_records'] - source_stats['total_records']) <= 5,
            'jobs_match': abs(postgres_stats['unique_jobs'] - source_stats['unique_jobs']) <= 2,
            'quantity_match': abs(postgres_stats['total_quantity'] - source_stats['total_quantity']) <= 1000,
            'record_difference': postgres_stats['total_records'] - source_stats['total_records'],
            'job_difference': postgres_stats['unique_jobs'] - source_stats['unique_jobs'],
            'quantity_difference': postgres_stats['total_quantity'] - source_stats['total_quantity']
        }
        
        # Get PCB type breakdown
        pcb_breakdown = db_manager.get_pcb_type_breakdown()
        
        # Get location breakdown  
        location_breakdown = db_manager.get_location_breakdown()
        
        return render_template('stats.html',
                             source_stats=source_stats,
                             postgres_stats=postgres_stats,
                             integrity_check=integrity_check,
                             pcb_breakdown=pcb_breakdown,
                             location_breakdown=location_breakdown)
    except Exception as e:
        logger.error(f"Error loading stats: {e}")
        flash("Error loading statistics page", 'error')
        return redirect(url_for('index'))

# SSO Integration with ACI Dashboard
@app.route('/sso/login', methods=['POST'])
def sso_login():
    """Handle SSO login from ACI Dashboard."""
    try:
        data = request.get_json() or {}

        # Extract user data from ACI Dashboard
        username = data.get('username')
        role = data.get('role', 'USER')
        itar_authorized = data.get('itar_authorized', False)
        auth_token = data.get('token')

        if username:
            session['username'] = username
            session['full_name'] = data.get('full_name', username.split('@')[0].capitalize())
            session['role'] = role
            session['itar_authorized'] = itar_authorized
            session['aci_auth_token'] = auth_token

            return jsonify({
                'success': True,
                'message': f'User {username} logged in successfully',
                'redirect_url': url_for('index')
            })
        else:
            return jsonify({'success': False, 'error': 'Missing username'}), 400

    except Exception as e:
        logger.error(f"SSO login error: {e}")
        return jsonify({'success': False, 'error': 'SSO login failed'}), 500

@app.route('/sso/callback')
@csrf.exempt
def sso_callback():
    """Handle SSO callback redirect from ACI FORGE.
    Validates the SSO JWT token and creates a KOSH session."""
    from jose import jwt as jose_jwt, JWTError as JoseJWTError

    sso_secret = os.environ.get('SSO_SECRET_KEY', '')
    token = request.args.get('token', '')

    if not token:
        flash('SSO login failed: No token provided.', 'danger')
        return redirect(url_for('login'))

    if not sso_secret:
        flash('SSO not configured on this server.', 'danger')
        return redirect(url_for('login'))

    try:
        payload = jose_jwt.decode(token, sso_secret, algorithms=['HS256'])
    except JoseJWTError:
        flash('SSO login failed: Invalid or expired token.', 'danger')
        return redirect(url_for('login'))

    if payload.get('type') != 'sso':
        flash('SSO login failed: Invalid token type.', 'danger')
        return redirect(url_for('login'))

    if payload.get('target_app') != 'kosh':
        flash('SSO login failed: Token not intended for this application.', 'danger')
        return redirect(url_for('login'))

    username = payload.get('sub', '')

    # Look up user in KOSH database
    # Try exact match first, then email match, then email prefix match (case-insensitive)
    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Try exact match on userlogin
        cursor.execute("""
            SELECT id, userid, username, userlogin, usersecurity
            FROM pcb_inventory."tblUser"
            WHERE userlogin = %s
        """, (username,))
        user = cursor.fetchone()

        # Try case-insensitive match
        if not user:
            cursor.execute("""
                SELECT id, userid, username, userlogin, usersecurity
                FROM pcb_inventory."tblUser"
                WHERE LOWER(userlogin) = LOWER(%s)
            """, (username,))
            user = cursor.fetchone()

        # If username is an email, try matching by email prefix (e.g., adam@... matches adam or AdamJ)
        if not user and '@' in username:
            email_prefix = username.split('@')[0].lower()
            cursor.execute("""
                SELECT id, userid, username, userlogin, usersecurity
                FROM pcb_inventory."tblUser"
                WHERE LOWER(userlogin) LIKE %s
                ORDER BY id LIMIT 1
            """, (email_prefix + '%',))
            user = cursor.fetchone()

        # Auto-create user if not found (FORGE is the source of truth)
        if not user:
            from passlib.hash import bcrypt as passlib_bcrypt
            display_name = username.split('@')[0].capitalize() if '@' in username else username
            default_password = passlib_bcrypt.hash('Welcome1!')
            cursor.execute("""
                INSERT INTO pcb_inventory."tblUser" (username, userlogin, password, usersecurity)
                VALUES (%s, %s, %s, %s)
                RETURNING id, username, userlogin, usersecurity
            """, (display_name, username, default_password, 'user'))
            conn.commit()
            user = cursor.fetchone()
            if user:
                user['userid'] = None
                logger.info(f"SSO auto-created KOSH user: {username}")

        if not user:
            flash(f'SSO login failed: Could not create user "{username}" in KOSH. Contact your administrator.', 'danger')
            return redirect(url_for('login'))

        # Create KOSH session (same as regular login)
        session.clear()
        session['user_id'] = user['id']
        session['username'] = user['userlogin']
        session['full_name'] = user['username']
        session['role'] = user['usersecurity']
        session['itar_authorized'] = True if user['usersecurity'] == 'Admin' else False
        session['sso_login'] = True
        session.permanent = True
        conn.commit()

        logger.info(f"SSO login successful for user: {username}")
        flash(f'Welcome, {user["username"] or username}! (Signed in via ACI FORGE)', 'success')
        return redirect(url_for('index'))

    except Exception as e:
        logger.error(f"SSO callback error: {e}")
        flash('SSO login failed: Internal error.', 'danger')
        return redirect(url_for('login'))
    finally:
        if conn:
            db_manager.return_connection(conn)


# API Endpoints
@app.route('/api/inventory')
@require_auth
def api_inventory():
    """API endpoint for inventory data."""
    try:
        user_role = session.get('role', 'USER')
        itar_auth = session.get('itar_authorized', False)
        inventory = db_manager.get_current_inventory(user_role, itar_auth)
        return jsonify({'success': True, 'data': inventory})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/stock', methods=['POST'])
@validate_api_request(['pcb_type', 'quantity', 'location'])
@require_auth
def api_stock():
    """API endpoint for stocking PCBs."""
    try:
        data = request.get_json()

        # Use part_number as job identifier (they're the same thing)
        job = data.get('part_number') or data.get('job')
        if not job:
            return jsonify({'success': False, 'error': 'Part number is required'}), 400

        user_role = session.get('role', 'USER')
        itar_auth = session.get('itar_authorized', False)
        itar_classification = data.get('itar_classification', 'NONE')

        # Check ITAR access
        if itar_classification == 'ITAR' and not user_manager.can_access_itar(user_role, itar_auth):
            return jsonify({'success': False, 'error': 'Access denied: ITAR authorization required'}), 403

        result = db_manager.stock_pcb(
            job=job,
            pcb_type=data['pcb_type'],
            quantity=data['quantity'],  # Already validated and converted to int
            location=data['location'],
            itar_classification=itar_classification,
            user_role=user_role,
            itar_auth=itar_auth,
            username=session.get('username', 'system')
        )
        return jsonify(result)
    except Exception as e:
        logger.error(f"API stock error: {e}")
        return jsonify({'success': False, 'error': 'Stock operation failed'}), 500

@app.route('/api/pick', methods=['POST'])
@validate_api_request(['pcb_type', 'quantity'])
@require_auth
def api_pick():
    """API endpoint for picking PCBs."""
    try:
        data = request.get_json()

        # Use part_number as job identifier (they're the same thing)
        job = data.get('part_number') or data.get('job')
        if not job:
            return jsonify({'success': False, 'error': 'Part number is required'}), 400

        user_role = session.get('role', 'USER')
        itar_auth = session.get('itar_authorized', False)

        result = db_manager.pick_pcb(
            job=job,
            pcb_type=data['pcb_type'],
            quantity=data['quantity'],  # Already validated and converted to int
            user_role=user_role,
            itar_auth=itar_auth,
            username=session.get('username', 'system')
        )
        return jsonify(result)
    except Exception as e:
        logger.error(f"API pick error: {e}")
        return jsonify({'success': False, 'error': 'Pick operation failed'}), 500

@app.route('/api/search')
@require_auth
def api_search():
    """API endpoint for searching inventory."""
    try:
        job = request.args.get('job')
        pcb_type = request.args.get('pcb_type')
        pcn = request.args.get('pcn')  # Optional PCN filter
        user_role = session.get('role', 'USER')
        itar_auth = session.get('itar_authorized', False)

        inventory = db_manager.search_inventory(
            job=job,
            pcb_type=pcb_type,
            pcn=pcn,  # Pass PCN to search_inventory
            user_role=user_role,
            itar_auth=itar_auth
        )

        # Skip expiration info for search results to avoid serialization issues
        return jsonify({'success': True, 'data': inventory})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/expiration-check')
@require_auth
def api_expiration_check():
    """API endpoint for checking expiration status of specific item."""
    try:
        dc = request.args.get('dc')
        pcb_type = request.args.get('pcb_type', 'Bare')
        msd = request.args.get('msd')

        expiration_info = expiration_manager.calculate_expiration_status(dc, pcb_type, msd)

        return jsonify({
            'success': True,
            'expiration': expiration_info,
            'display_text': expiration_manager.format_expiration_display(expiration_info),
            'badge_class': expiration_manager.get_expiration_badge_class(expiration_info['status']),
            'icon_class': expiration_manager.get_expiration_icon(expiration_info['status'])
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/bom/mpns/<part_number>', methods=['GET'])
@require_auth
def api_get_mpns_for_part(part_number):
    """Get MPNs from BOM table for a specific part number"""
    conn = None
    try:
        logger.info(f"Fetching MPNs from BOM for part_number={part_number}")

        conn = db_manager.get_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Query BOM table to get all MPNs for this ACI part number
        cur.execute("""
            SELECT DISTINCT mpn
            FROM pcb_inventory."tblBOM"
            WHERE aci_pn = %s AND mpn IS NOT NULL AND mpn != ''
            ORDER BY mpn
        """, (part_number,))

        results = cur.fetchall()
        cur.close()

        if results:
            mpns = [{'mpn': row['mpn']} for row in results]
            logger.info(f"Found {len(mpns)} MPN(s) for part {part_number} in BOM")
            return jsonify({'success': True, 'mpns': mpns, 'count': len(mpns), 'part_number': part_number})
        else:
            logger.info(f"No MPNs found in BOM for part {part_number}")
            return jsonify({'success': True, 'mpns': [], 'count': 0, 'part_number': part_number})

    except Exception as e:
        logger.error(f"Error fetching MPNs for part {part_number}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)


# Access Database Routes
@app.route('/source')
def source_access():
    """Source (Access) database browser main page."""
    try:
        from access_db_manager import AccessDBManager
        
        # Path to Access database (mounted in container)
        access_db_path = "/app/INVENTORY TABLE.mdb"
        
        with AccessDBManager(access_db_path) as access_db:
            db_info = access_db.get_database_info()
            
        return render_template('source_access.html', 
                             db_info=db_info,
                             page_title="Source (Access) Database")
    except Exception as e:
        flash(f'Error accessing Access database: {str(e)}', 'error')
        return render_template('source_access.html', 
                             db_info=None,
                             error=str(e),
                             page_title="Source (Access) Database")

@app.route('/source/table/<table_name>')
def source_table_view(table_name):
    """View data from a specific Access database table."""
    try:
        from access_db_manager import AccessDBManager
        
        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        offset = (page - 1) * per_page
        
        # Path to Access database (mounted in container)
        access_db_path = "/app/INVENTORY TABLE.mdb"
        
        with AccessDBManager(access_db_path) as access_db:
            # Get table schema
            schema = access_db.get_table_schema(table_name)
            
            # Get table data
            data, total_records = access_db.get_table_data(table_name, limit=per_page, offset=offset)
            
            # Calculate pagination info with safety checks
            total_records = max(0, total_records)  # Ensure non-negative
            total_pages = max(1, (total_records + per_page - 1) // per_page) if total_records > 0 else 1
            has_prev = page > 1
            has_next = page < total_pages
            
            pagination_info = {
                'page': page,
                'per_page': per_page,
                'total_records': total_records,
                'total_pages': total_pages,
                'has_prev': has_prev,
                'has_next': has_next,
                'prev_page': page - 1 if has_prev else None,
                'next_page': page + 1 if has_next else None
            }
            
        return render_template('source_table_view.html',
                             table_name=table_name,
                             schema=schema,
                             data=data,
                             pagination=pagination_info,
                             page_title=f"Table: {table_name}")
    except Exception as e:
        # Log the full exception for debugging
        app.logger.error(f'Error viewing table {table_name}: {str(e)}', exc_info=True)
        flash(f'Error viewing table {table_name}: {str(e)}', 'error')
        return redirect(url_for('source_access'))

@app.route('/source/query', methods=['GET', 'POST'])
def source_query():
    """DISABLED: Raw SQL query interface removed for security reasons."""
    flash("Raw SQL query interface has been disabled for security reasons. Use the table view instead.", "warning")
    return redirect(url_for('source_access'))

@app.route('/api/source/tables')
def api_source_tables():
    """API endpoint to get Access database table list."""
    try:
        from access_db_manager import AccessDBManager
        
        # Path to Access database (mounted in container)
        access_db_path = "/app/INVENTORY TABLE.mdb"
        
        with AccessDBManager(access_db_path) as access_db:
            tables = access_db.get_table_list()
            
        return jsonify({'success': True, 'data': tables})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/source/table-data/<table_name>')
def api_source_table_data(table_name):
    """API endpoint to get actual data from Access database table."""
    try:
        from access_db_manager import AccessDBManager
        
        # Get query parameters
        limit = request.args.get('limit', 100, type=int)
        offset = request.args.get('offset', 0, type=int)
        
        # Path to Access database (mounted in container)
        access_db_path = "/app/INVENTORY TABLE.mdb"
        
        with AccessDBManager(access_db_path) as access_db:
            data, total_records = access_db.get_table_data(table_name, limit=limit, offset=offset)
            
            # Check if we got actual data or fallback message
            if data and len(data) > 0:
                first_row = data[0]
                # Check if this is our fallback data (contains 'Message' key)
                if 'Message' in first_row and 'requires mdb-tools' in str(first_row.get('Message', '')):
                    return jsonify({
                        'success': False, 
                        'message': first_row.get('Message', 'Data access limited'),
                        'note': first_row.get('Note', ''),
                        'alternative': first_row.get('Alternative', '')
                    })
                else:
                    # This is actual data
                    return jsonify({
                        'success': True, 
                        'data': data, 
                        'total_records': total_records,
                        'table_name': table_name
                    })
            else:
                return jsonify({
                    'success': False, 
                    'message': 'No data available',
                    'total_records': 0
                })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/generate-pcn')
@require_auth
def generate_pcn():
    """Generate PCN page"""
    return render_template('generate_pcn.html')

@app.route('/po-history')
@require_auth
def po_history():
    """PO History lookup page"""
    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    # Get filter parameters
    search_po = request.args.get('po', '').strip()
    search_item = request.args.get('item', '').strip()
    search_mpn = request.args.get('mpn', '').strip()
    search_pcn = request.args.get('pcn', '').strip()
    search_date_from = request.args.get('date_from', '').strip()
    search_date_to = request.args.get('date_to', '').strip()

    conn = None
    try:
        conn = db_manager.get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Build query from tblTransaction (actual table that stores PO data)
            query = """
                SELECT id, po as po_number, item, pcn, mpn, dc as date_code,
                       CASE
                           WHEN tranqty ~ '^[0-9]+$' THEN CAST(tranqty AS INTEGER)
                           ELSE NULL
                       END as quantity,
                       trantype as transaction_type, tran_time as transaction_date,
                       loc_from as location_from, loc_to as location_to, userid as user_id,
                       vendor as vendor_name
                FROM pcb_inventory."tblTransaction"
                WHERE po IS NOT NULL AND po <> ''
            """
            params = []

            if search_po:
                query += " AND po ILIKE %s"
                params.append(f'%{search_po}%')

            if search_item:
                query += " AND item ILIKE %s"
                params.append(f'%{search_item}%')

            if search_mpn:
                query += " AND mpn ILIKE %s"
                params.append(f'%{search_mpn}%')

            if search_pcn:
                query += " AND pcn = %s"
                params.append(search_pcn)

            if search_date_from:
                query += " AND tran_time >= %s"
                params.append(search_date_from)

            if search_date_to:
                query += " AND tran_time <= %s"
                params.append(f'{search_date_to} 23:59:59')

            # Get total count
            count_query = f"SELECT COUNT(*) FROM ({query}) AS count_query"
            cur.execute(count_query, params)
            total_count = cur.fetchone()['count']

            # Add sorting and pagination
            query += " ORDER BY id DESC LIMIT %s OFFSET %s"
            params.extend([per_page, (page - 1) * per_page])

            # Execute query
            cur.execute(query, params)
            receipts = [dict(row) for row in cur.fetchall()]

            # Convert quantity strings to integers for proper summing in template
            for receipt in receipts:
                if receipt.get('quantity') is not None:
                    try:
                        # If it's already an int, leave it; if it's a string, convert it
                        if isinstance(receipt['quantity'], str):
                            receipt['quantity'] = int(receipt['quantity']) if receipt['quantity'].strip() else 0
                    except (ValueError, AttributeError):
                        receipt['quantity'] = 0

            # Calculate pagination
            total_pages = (total_count + per_page - 1) // per_page
            pagination = {
                'page': page,
                'per_page': per_page,
                'total': total_count,
                'total_pages': total_pages,
                'has_prev': page > 1,
                'has_next': page < total_pages,
                'prev_num': page - 1 if page > 1 else None,
                'next_num': page + 1 if page < total_pages else None,
                'pages': list(range(max(1, page - 2), min(total_pages + 1, page + 3)))
            }

            return render_template('po_history.html',
                                 receipts=receipts,
                                 pagination=pagination,
                                 search_po=search_po,
                                 search_item=search_item,
                                 search_mpn=search_mpn,
                                 search_pcn=search_pcn,
                                 search_date_from=search_date_from,
                                 search_date_to=search_date_to)
    except Exception as e:
        logger.error(f"Error loading PO history: {e}")
        flash('Error loading PO history. Please try again.', 'error')
        return render_template('po_history.html', receipts=[], pagination={'total': 0})
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/pcn-history')
@require_auth
def pcn_history():
    """PCN transaction history page - focused on efficiency"""
    # Get PCN parameter only
    search_pcn = request.args.get('pcn', '').strip()

    conn = None
    transactions = []
    pcn_info = None

    try:
        if search_pcn:
            conn = db_manager.get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Get all transactions for the PCN (no pagination, show everything)
                # Format tran_time consistently as MM/DD/YYYY HH:MI:SS AM/PM for ALL date formats
                query = """
                    SELECT trantype, item, mpn, tranqty,
                           CASE
                               -- Handle ISO format timestamps (YYYY-MM-DD HH:MM:SS...) - convert from UTC to EST
                               WHEN tran_time ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}' THEN
                                   TO_CHAR(timezone('America/New_York', tran_time::timestamptz), 'MM/DD/YYYY HH12:MI:SS AM')
                               -- Handle old short format (MM/DD/YY HH:MI:SS) - convert to full year
                               WHEN tran_time ~ '^[0-9]{2}/[0-9]{2}/[0-9]{2}\\s+[0-9]{2}:[0-9]{2}' THEN
                                   TO_CHAR(TO_TIMESTAMP(tran_time, 'MM/DD/YY HH24:MI:SS'), 'MM/DD/YYYY HH12:MI:SS AM')
                               -- If empty, NULL or other format, return as-is
                               ELSE
                                   tran_time
                           END as tran_time,
                           loc_from, loc_to, wo, po,
                           -- Create sortable timestamp for ORDER BY
                           CASE
                               WHEN tran_time ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}' THEN tran_time::timestamptz
                               WHEN tran_time ~ '^[0-9]{2}/[0-9]{2}/[0-9]{2}\\s+[0-9]{2}:[0-9]{2}' THEN TO_TIMESTAMP(tran_time, 'MM/DD/YY HH24:MI:SS')
                               ELSE NULL
                           END as sort_time
                    FROM pcb_inventory."tblTransaction"
                    WHERE pcn::text = %s
                    ORDER BY sort_time DESC NULLS LAST, id DESC
                    LIMIT 500
                """
                cur.execute(query, (search_pcn,))
                transactions = [dict(row) for row in cur.fetchall()]

                # Get PCN info from warehouse inventory
                cur.execute("""
                    SELECT item, mpn, dc, onhandqty, mfg_qty, loc_to, msd, po
                    FROM pcb_inventory."tblWhse_Inventory"
                    WHERE pcn::text = %s
                    LIMIT 1
                """, (search_pcn,))
                result = cur.fetchone()
                if result:
                    pcn_info = dict(result)

            return render_template('pcn_history.html',
                                 transactions=transactions,
                                 pcn_info=pcn_info,
                                 search_pcn=search_pcn)
        else:
            # No PCN provided, just show the search form
            return render_template('pcn_history.html',
                                 transactions=[],
                                 pcn_info=None,
                                 search_pcn='')

    except Exception as e:
        logger.error(f"Error loading PCN history: {e}")
        flash('Error loading PCN history. Please try again.', 'error')
        return render_template('pcn_history.html', transactions=[], pcn_info=None, search_pcn=search_pcn)
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/stock-alerts')
@require_auth
def stock_alerts():
    """Stock Alerts page - shows all items below threshold."""
    conn = None
    cursor = None
    try:
        LOW_STOCK_THRESHOLD = 10

        # Get pagination parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 25, type=int)
        per_page = min(max(per_page, 10), 100)

        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Get low stock items with pagination
        query = """
            SELECT pcn, item, mpn, dc, onhandqty, loc_to, msd, po
            FROM pcb_inventory."tblWhse_Inventory"
            WHERE onhandqty < %s AND onhandqty >= 0
            ORDER BY onhandqty ASC, item ASC
        """

        # Get total count
        count_query = """
            SELECT COUNT(*) as total
            FROM pcb_inventory."tblWhse_Inventory"
            WHERE onhandqty < %s AND onhandqty >= 0
        """
        cursor.execute(count_query, (LOW_STOCK_THRESHOLD,))
        total_records = cursor.fetchone()['total']

        # Get paginated results
        query += " LIMIT %s OFFSET %s"
        cursor.execute(query, (LOW_STOCK_THRESHOLD, per_page, (page - 1) * per_page))
        low_stock_items = [dict(row) for row in cursor.fetchall()]

        # Calculate pagination
        total_pages = (total_records + per_page - 1) // per_page if total_records > 0 else 1

        pagination = {
            'page': page,
            'per_page': per_page,
            'total': total_records,
            'total_pages': total_pages,
            'has_prev': page > 1,
            'has_next': page < total_pages,
            'prev_num': page - 1 if page > 1 else None,
            'next_num': page + 1 if page < total_pages else None,
            'pages': list(range(max(1, page - 2), min(total_pages + 1, page + 3)))
        }

        return render_template('stock_alerts.html',
                             low_stock_items=low_stock_items,
                             low_stock_threshold=LOW_STOCK_THRESHOLD,
                             pagination=pagination,
                             total_records=total_records)

    except Exception as e:
        logger.error(f"Error loading stock alerts: {e}")
        flash('Error loading stock alerts. Please try again.', 'error')
        return render_template('stock_alerts.html',
                             low_stock_items=[],
                             low_stock_threshold=10,
                             pagination={'total': 0, 'page': 1, 'total_pages': 1, 'per_page': 25},
                             total_records=0)
    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass
        if conn:
            try:
                db_manager.return_connection(conn)
            except Exception:
                pass

@app.route('/api/pcn/generate', methods=['POST'])
@require_auth
def api_generate_pcn():
    """API endpoint to generate new PCN"""
    try:
        data = request.get_json()

        # Validate required fields
        if not data.get('item'):
            return jsonify({'error': 'Item (Job Number) is required'}), 400

        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        try:
            # Validate quantity input
            quantity = data.get('quantity', 0)
            if quantity:
                try:
                    quantity = int(quantity)
                    if quantity < 0 or quantity > 10000:
                        return jsonify({'error': 'Quantity must be between 0 and 10000'}), 400
                except (ValueError, TypeError):
                    return jsonify({'error': 'Quantity must be a number'}), 400

            # Generate new PCN using MAX+1 with advisory lock to prevent duplicates
            # Lock key 73746 = arbitrary constant for PCN generation
            cursor.execute("SELECT pg_advisory_xact_lock(73746)")
            cursor.execute("""
                SELECT COALESCE(MAX(pcn::integer), 0) + 1 as next_pcn
                FROM pcb_inventory."tblTransaction"
                WHERE pcn ~ '^[0-9]+$'
            """)
            result = cursor.fetchone()
            pcn_number = str(result['next_pcn'])

            logger.info(f"Generated new PCN: {pcn_number} (using MAX+1 with advisory lock)")

            # Insert into tblTransaction
            cursor.execute("""
                INSERT INTO pcb_inventory."tblTransaction"
                (record_no, trantype, item, pcn, mpn, dc, msd, tranqty, tran_time, loc_from, loc_to, wo, po, userid, vendor)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', 'MM/DD/YY HH24:MI:SS'), %s, %s, %s, %s, %s, %s)
                RETURNING id, pcn, item, mpn, dc, msd, tranqty, created_at
            """, (
                None,  # record_no
                'PCN Generation',  # trantype
                data.get('item'),  # item (job number)
                pcn_number,  # pcn
                data.get('mpn'),  # mpn
                data.get('date_code'),  # dc
                data.get('msd'),  # msd
                data.get('quantity', '0'),  # tranqty
                '-',  # loc_from
                data.get('location', 'Inventory'),  # loc_to
                data.get('wo'),  # wo (work order)
                data.get('po_number'),  # po
                session.get('username', 'system'),  # userid
                data.get('vendor_name')  # vendor
            ))

            transaction_record = cursor.fetchone()
            logger.info(f"Created PCN {pcn_number} in tblTransaction (ID: {transaction_record['id']})")

            # Also insert into warehouse inventory
            cursor.execute("""
                INSERT INTO pcb_inventory."tblWhse_Inventory"
                (item, pcn, mpn, dc, onhandqty, loc_from, loc_to, msd, po, vendor)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data.get('item'),
                pcn_number,
                data.get('mpn') or '',
                data.get('date_code'),
                data.get('quantity', 0),  # Set initial quantity from user input
                data.get('location_from', '-'),
                data.get('location', 'Receiving Area'),
                data.get('msd'),
                data.get('po_number'),
                data.get('vendor_name')
            ))
            logger.info(f"Added PCN {pcn_number} to warehouse inventory")

            conn.commit()

            logger.info(f"Generated PCN: {pcn_number} for item: {data.get('item')}")
            log_user_activity('PCN_GENERATE', f"Generated PCN {pcn_number} for item {data.get('item')}", f"Qty: {data.get('quantity', 0)}, MPN: {data.get('mpn', '')}")

            return jsonify({
                'success': True,
                'pcn_number': transaction_record['pcn'],
                'pcn_id': transaction_record['id'],
                'item': transaction_record['item'],
                'po_number': data.get('po_number'),
                'part_number': data.get('part_number'),
                'mpn': transaction_record['mpn'],
                'quantity': transaction_record['tranqty'],
                'date_code': transaction_record['dc'],
                'msd': data.get('msd'),
                'created_at': transaction_record['created_at'].isoformat() if transaction_record['created_at'] else None
            })

        except Exception as e:


            if conn:


                conn.rollback()
            logger.error(f"Error generating PCN: {e}")
            return jsonify({'error': str(e)}), 500
        finally:

            if cursor:

                cursor.close()

            if conn:

                db_manager.return_connection(conn)

    except Exception as e:
        logger.error(f"Error in PCN generation endpoint: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/pcn/details/<pcn_number>', methods=['GET'])
def api_get_pcn_details(pcn_number):
    """API endpoint to get PCN details by PCN number - for auto-populating fields on scan"""
    try:
        conn = None
        cursor = None
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        try:
            # Check if item_number filter was provided (for resolving duplicate PCNs)
            item_filter = request.args.get('item')

            # Query from tblWhse_Inventory (main warehouse table)
            if item_filter:
                cursor.execute("""
                    SELECT pcn, item, mpn, onhandqty, dc, msd, loc_from, loc_to, po
                    FROM pcb_inventory."tblWhse_Inventory"
                    WHERE pcn::text = %s AND item::text = %s
                    LIMIT 1
                """, (pcn_number, item_filter))
            else:
                cursor.execute("""
                    SELECT pcn, item, mpn, onhandqty, dc, msd, loc_from, loc_to, po
                    FROM pcb_inventory."tblWhse_Inventory"
                    WHERE pcn::text = %s
                    ORDER BY id DESC
                """, (pcn_number,))

            whse_records = cursor.fetchall()

            if whse_records:
                # If multiple records found for the same PCN, return them all
                # so the frontend can ask the user to choose
                if len(whse_records) > 1:
                    matches = []
                    for rec in whse_records:
                        matches.append({
                            'pcn_number': str(rec['pcn']),
                            'part_number': rec['item'],
                            'job': rec['item'],
                            'mpn': rec['mpn'],
                            'quantity': rec['onhandqty'],
                            'date_code': rec['dc'],
                            'msd': rec['msd'],
                            'location': rec['loc_to'],
                            'location_from': rec['loc_from'],
                            'po_number': rec['po']
                        })
                    return jsonify({
                        'success': True,
                        'multiple': True,
                        'count': len(matches),
                        'matches': matches,
                        'message': f'Multiple items found for PCN {pcn_number}. Please select the correct one.'
                    })

                # Single record - return as before
                whse_record = whse_records[0]
                return jsonify({
                    'success': True,
                    'pcn_number': str(whse_record['pcn']),
                    'part_number': whse_record['item'],
                    'job': whse_record['item'],
                    'mpn': whse_record['mpn'],
                    'quantity': whse_record['onhandqty'],
                    'date_code': whse_record['dc'],
                    'msd': whse_record['msd'],
                    'location': whse_record['loc_to'],
                    'location_from': whse_record['loc_from'],
                    'po_number': whse_record['po']
                })

            # If not in tblWhse_Inventory, try tblTransaction
            cursor.execute("""
                SELECT pcn, item, mpn, tranqty, dc, msd, loc_from, loc_to, wo, po, userid, created_at
                FROM pcb_inventory."tblTransaction"
                WHERE pcn::text = %s
                ORDER BY created_at DESC
                LIMIT 1
            """, (pcn_number,))

            transaction_record = cursor.fetchone()

            if transaction_record:
                return jsonify({
                    'success': True,
                    'pcn_number': str(transaction_record['pcn']),
                    'part_number': transaction_record['item'],
                    'job': transaction_record['item'],
                    'mpn': transaction_record['mpn'],
                    'quantity': transaction_record['tranqty'],
                    'date_code': transaction_record['dc'],
                    'msd': transaction_record['msd'],
                    'location': transaction_record['loc_to'],
                    'location_from': transaction_record['loc_from'],
                    'work_order': transaction_record['wo'],
                    'po_number': transaction_record['po'],
                    'created_at': transaction_record['created_at'].isoformat() if transaction_record['created_at'] else None,
                    'created_by': transaction_record['userid']
                })

            return jsonify({'success': False, 'error': 'PCN not found'}), 404

        finally:


            if cursor:


                cursor.close()


            if conn:


                db_manager.return_connection(conn)

    except Exception as e:
        logger.error(f"Error fetching PCN details: {e}")
        return jsonify({'success': False, 'error': 'Internal server error'}), 500

@app.route('/api/pcn/list', methods=['GET'])
def api_list_pcn():
    """API endpoint to list PCN records"""
    try:
        conn = None
        cursor = None
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        try:
            # Query from tblWhse_Inventory for recent PCN records
            cursor.execute("""
                SELECT id as pcn_id, pcn as pcn_number, item, po as po_number,
                       item as part_number, mpn, onhandqty as quantity,
                       dc as date_code, msd, created_at, NULL as created_by
                FROM pcb_inventory."tblWhse_Inventory"
                WHERE pcn IS NOT NULL AND pcn <> ''
                ORDER BY created_at DESC
                LIMIT 100
            """)

            records = cursor.fetchall()

            return jsonify({
                'success': True,
                'records': [{
                    'pcn_id': r['pcn_id'],
                    'pcn_number': str(r['pcn_number']),
                    'item': r['item'],
                    'po_number': r['po_number'],
                    'part_number': r['part_number'],
                    'mpn': r['mpn'],
                    'quantity': r['quantity'],
                    'date_code': r['date_code'],
                    'msd': r['msd'],
                    'created_at': r['created_at'].isoformat() if r['created_at'] else None,
                    'created_by': r['created_by']
                } for r in records]
            })

        finally:


            if cursor:


                cursor.close()


            if conn:


                db_manager.return_connection(conn)

    except Exception as e:
        logger.error(f"Error listing PCN records: {e}")
        return jsonify({'error': 'Internal server error'}), 500

@app.route('/api/pcn/delete/<pcn_number>', methods=['DELETE'])
@require_auth
def api_delete_pcn(pcn_number):
    """API endpoint to delete a PCN record"""
    logger.info(f"=== INSIDE api_delete_pcn FUNCTION - PCN: {pcn_number} ===")

    try:
        conn = None
        cursor = None
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        try:
            # Check if PCN exists in tblTransaction
            cursor.execute("""
                SELECT pcn, item
                FROM pcb_inventory."tblTransaction"
                WHERE pcn::text = %s
                LIMIT 1
            """, (pcn_number,))

            transaction_record = cursor.fetchone()

            if not transaction_record:
                return jsonify({'success': False, 'error': 'PCN not found'}), 404

            item_name = transaction_record['item']

            # Delete from tblWhse_Inventory (warehouse inventory)
            cursor.execute("""
                DELETE FROM pcb_inventory."tblWhse_Inventory"
                WHERE pcn::text = %s
            """, (pcn_number,))

            deleted_warehouse = cursor.rowcount
            logger.info(f"Deleted {deleted_warehouse} records from tblWhse_Inventory for PCN {pcn_number}")

            # Delete from tblTransaction
            cursor.execute("""
                DELETE FROM pcb_inventory."tblTransaction"
                WHERE pcn::text = %s
            """, (pcn_number,))

            deleted_transaction = cursor.rowcount
            logger.info(f"Deleted {deleted_transaction} records from tblTransaction for PCN {pcn_number}")

            conn.commit()

            logger.info(f"Deleted PCN {pcn_number} (Item: {item_name}) by user: {session.get('username', 'system')}")

            return jsonify({
                'success': True,
                'message': f'PCN {pcn_number} deleted successfully',
                'pcn_number': pcn_number
            })

        except Exception as e:


            if conn:


                conn.rollback()
            logger.error(f"Error deleting PCN {pcn_number}: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
        finally:

            if cursor:

                cursor.close()

            if conn:

                db_manager.return_connection(conn)

    except Exception as e:
        logger.error(f"Error in PCN delete endpoint: {e}")
        return jsonify({'success': False, 'error': 'Internal server error'}), 500

@app.route('/api/pcn/assign', methods=['POST'])
@require_auth
def api_assign_pcn():
    """API endpoint to assign PCN to inventory item"""
    try:
        data = request.get_json()

        # Validate required fields
        if not data.get('job') or not data.get('pcb_type'):
            return jsonify({'success': False, 'error': 'Job and PCB type are required'}), 400

        username = session.get('username', 'system')
        result = db_manager.assign_pcn_to_item(
            job=data['job'],
            pcb_type=data['pcb_type'],
            username=username
        )

        return jsonify(result)
    except Exception as e:
        logger.error(f"Error assigning PCN: {e}")
        return jsonify({'success': False, 'error': 'Failed to assign PCN'}), 500

@app.route('/api/pcn/history', methods=['GET'])
def api_pcn_history():
    """API endpoint to get PCN history - NO AUTH REQUIRED for public access"""
    try:
        limit = request.args.get('limit', 100, type=int)
        pcn = request.args.get('pcn', None)
        job = request.args.get('job', None)
        pcb_type = request.args.get('pcb_type', None)
        status = request.args.get('status', None)

        filters = {}
        if pcn:
            filters['pcn'] = pcn
        if job:
            filters['job'] = job
        if pcb_type:
            filters['pcb_type'] = pcb_type
        if status:
            filters['status'] = status

        history = db_manager.get_pcn_history(limit=limit, filters=filters if filters else None)

        # Format dates for JSON serialization
        for record in history:
            if record.get('generated_at'):
                # Handle both datetime objects and string dates
                if hasattr(record['generated_at'], 'isoformat'):
                    record['generated_at'] = record['generated_at'].isoformat()
                # else: leave as string

        return jsonify({'success': True, 'data': history})
    except Exception as e:
        logger.error(f"Error getting PCN history: {e}")
        return jsonify({'success': False, 'error': 'Failed to get PCN history'}), 500

@app.route('/api/pcn/search', methods=['GET'])
@require_auth
def api_pcn_search():
    """API endpoint to search PCN records"""
    try:
        pcn_number = request.args.get('pcn', None)
        job = request.args.get('job', None)

        if not pcn_number and not job:
            return jsonify({'success': False, 'error': 'PCN number or job is required'}), 400

        results = db_manager.search_pcn(pcn_number=pcn_number, job=job)

        # Format dates for JSON serialization
        for record in results:
            if record.get('generated_at'):
                # Handle both datetime objects and string dates
                if hasattr(record['generated_at'], 'isoformat'):
                    record['generated_at'] = record['generated_at'].isoformat()
                # else: leave as string

        return jsonify({'success': True, 'data': results})
    except Exception as e:
        logger.error(f"Error searching PCN: {e}")
        return jsonify({'success': False, 'error': 'Failed to search PCN'}), 500

@app.route('/api/po/history', methods=['GET'])
def api_po_history():
    """API endpoint to get PO history - NO AUTH REQUIRED for public access"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        limit = request.args.get('limit', per_page, type=int)  # For backwards compatibility
        po_number = request.args.get('po_number', None)
        item = request.args.get('item', None)
        date_from = request.args.get('date_from', None)
        date_to = request.args.get('date_to', None)

        filters = {}
        if po_number:
            filters['po_number'] = po_number
        if item:
            filters['item'] = item
        if date_from:
            filters['date_from'] = date_from
        if date_to:
            filters['date_to'] = date_to

        # Calculate offset for pagination
        offset = (page - 1) * per_page

        # Get total count first
        total_count = db_manager.get_po_history_count(filters if filters else None)

        # Get paginated results
        history = db_manager.get_po_history(limit=per_page, offset=offset, filters=filters if filters else None)

        # Format dates for JSON serialization
        for record in history:
            if record.get('transaction_date'):
                # Handle both datetime objects and string dates
                if hasattr(record['transaction_date'], 'isoformat'):
                    record['transaction_date'] = record['transaction_date'].isoformat()
                # else: leave as string
            if record.get('created_at'):
                if hasattr(record['created_at'], 'isoformat'):
                    record['created_at'] = record['created_at'].isoformat()
                # else: leave as string

        return jsonify({
            'success': True,
            'data': history,
            'total': total_count,
            'page': page,
            'per_page': per_page,
            'total_pages': (total_count + per_page - 1) // per_page
        })
    except Exception as e:
        logger.error(f"Error getting PO history: {e}")
        return jsonify({'success': False, 'error': 'Failed to get PO history'}), 500

@app.route('/api/po/search', methods=['GET'])
def api_po_search():
    """API endpoint to search PO records - NO AUTH REQUIRED for public access"""
    try:
        po_number = request.args.get('po_number', None)
        item = request.args.get('item', None)

        if not po_number and not item:
            return jsonify({'success': False, 'error': 'PO number or item is required'}), 400

        results = db_manager.search_po(po_number=po_number, item=item)

        # Format dates for JSON serialization
        for record in results:
            if record.get('transaction_date'):
                # Handle both datetime objects and string dates
                if hasattr(record['transaction_date'], 'isoformat'):
                    record['transaction_date'] = record['transaction_date'].isoformat()
                # else: leave as string
            if record.get('created_at'):
                if hasattr(record['created_at'], 'isoformat'):
                    record['created_at'] = record['created_at'].isoformat()
                # else: leave as string

        return jsonify({'success': True, 'data': results, 'total': len(results)})
    except Exception as e:
        logger.error(f"Error searching PO: {e}")
        return jsonify({'success': False, 'error': 'Failed to search PO'}), 500

@app.route('/print-label/<pcn_number>')
def print_label(pcn_number):
    """Dedicated print page for barcode label"""
    try:
        conn = None
        cursor = None
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        try:
            # First check tblWhse_Inventory for most current data (updated by restock/stock/pick)
            cursor.execute("""
                SELECT pcn::varchar as pcn_number,
                       item,
                       po as po_number,
                       item as part_number,
                       mpn,
                       onhandqty as quantity,
                       dc as date_code,
                       msd,
                       NULL as barcode_data,
                       loc_to as location,
                       NULL as pcb_type
                FROM pcb_inventory."tblWhse_Inventory"
                WHERE pcn::text = %s
                ORDER BY created_at DESC LIMIT 1
            """, (pcn_number,))

            pcn_data = cursor.fetchone()

            if not pcn_data:
                return "PCN not found", 404

            log_user_activity('PRINT_LABEL', f"Printed label for PCN {pcn_number}", f"Item: {pcn_data.get('item', '')}")
            response = make_response(render_template('print_label.html', data=dict(pcn_data)))
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response

        finally:
            if cursor:
                cursor.close()
            if conn:
                db_manager.return_connection(conn)

    except Exception as e:
        logger.error(f"Error loading print label: {e}")
        return "Error loading label", 500

@app.route('/print-label/<pcn_number>/zpl')
def generate_zpl_label(pcn_number):
    """Generate ZPL code for Zebra ZP450 printer (3x1 inch label)"""
    try:
        conn = None
        cursor = None
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        try:
            # Get PCN data from tblWhse_Inventory
            cursor.execute("""
                SELECT pcn::varchar as pcn_number,
                       item,
                       po as po_number,
                       item as part_number,
                       mpn,
                       onhandqty as quantity,
                       dc as date_code,
                       msd
                FROM pcb_inventory."tblWhse_Inventory"
                WHERE pcn::text = %s
                ORDER BY created_at DESC LIMIT 1
            """, (pcn_number,))

            pcn_data = cursor.fetchone()

            if not pcn_data:
                return "PCN not found", 404

            # Convert to dict
            data = dict(pcn_data)

            # Generate ZPL code for 3x1 inch label (Zebra ZP450) - v5.0 compact scannable
            # Label dimensions: 3 inches wide (609 dots @ 203dpi), 1 inch tall (203 dots @ 203dpi)
            zpl = f"""^XA
^FO30,5^A0N,24,24^FDPCN: {data['pcn_number']}^FS
^FO30,28^A0N,24,24^FDQTY: {data.get('quantity', 0)}^FS

^FO210,2^BY3,2,55^BCN,55,N,N,N^FD{data['pcn_number']}^FS

^FO15,58^GB579,0,2^FS

^FO35,65^A0N,22,22^FDItem No: {data.get('item', 'N/A')}^FS
^FO320,65^A0N,22,22^FDDCC: {data.get('date_code', 'N/A')}^FS

^FO35,88^A0N,22,22^FDMPN: {data.get('mpn', 'N/A')}^FS
^FO320,88^A0N,22,22^FDMSD: {data.get('msd', 'N/A')}^FS

^FO35,111^A0N,22,22^FDPO: {data.get('po_number', 'N/A')}^FS

^XZ"""

            # Return ZPL as downloadable file
            response = make_response(zpl)
            response.headers['Content-Type'] = 'application/zpl'
            response.headers['Content-Disposition'] = f'attachment; filename="PCN_{pcn_number}.zpl"'
            return response

        finally:


            if cursor:


                cursor.close()


            if conn:


                db_manager.return_connection(conn)

    except Exception as e:
        logger.error(f"Error generating ZPL: {e}")
        return "Error generating ZPL", 500

@app.route('/api/valuation/snapshots', methods=['GET'])
def api_get_valuation_snapshots():
    """Get list of available pricing snapshots - simplified to return empty list"""
    try:
        # Pricing snapshot feature not available - return empty list
        return jsonify({'success': True, 'snapshots': []})

    except Exception as e:
        logger.error(f"Error fetching valuation snapshots: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/valuation/<snapshot_date>', methods=['GET'])
def api_get_valuation_by_date(snapshot_date):
    """Get inventory valuation - simplified to return current inventory summary"""
    conn = None
    try:
        # Validate date format
        from datetime import datetime, date
        try:
            parsed_date = datetime.strptime(snapshot_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'success': False,
                'error': f'Invalid date format. Use YYYY-MM-DD (e.g., 2025-08-31)'
            }), 400

        conn = db_manager.get_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Calculate inventory value by joining with BOM cost data
        cur.execute("""
            SELECT
                COUNT(DISTINCT inv.job) as item_count,
                SUM(COALESCE(inv.qty, 0)) as total_quantity,
                SUM(COALESCE(inv.qty, 0) * COALESCE(bom.cost, 0)) as total_value,
                COUNT(CASE WHEN bom.cost IS NOT NULL AND bom.cost > 0 THEN 1 END) as items_with_cost
            FROM pcb_inventory."tblPCB_Inventory" inv
            LEFT JOIN LATERAL (
                SELECT AVG(cost) as cost
                FROM pcb_inventory."tblBOM"
                WHERE job::text = inv.job
                  AND cost IS NOT NULL
                  AND cost > 0
            ) bom ON true
            WHERE inv.qty > 0
        """)
        result_row = cur.fetchone()

        cur.close()

        total_value = float(result_row['total_value']) if result_row and result_row['total_value'] else 0
        item_count = int(result_row['item_count']) if result_row and result_row['item_count'] else 0
        items_with_cost = int(result_row['items_with_cost']) if result_row and result_row['items_with_cost'] else 0

        cost_coverage = (items_with_cost / item_count * 100) if item_count > 0 else 0

        result = {
            'success': True,
            'snapshot': {
                'date': snapshot_date,
                'total_value': round(total_value, 2),
                'total_quantity': int(result_row['total_quantity']) if result_row and result_row['total_quantity'] else 0,
                'item_count': item_count,
                'notes': f'Calculated from BOM cost data ({items_with_cost}/{item_count} jobs have pricing - {cost_coverage:.1f}% coverage)'
            }
        }

        return jsonify(result)

    except Exception as e:
        logger.error(f"Error fetching valuation: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/api/inventory/history', methods=['GET'])
@require_auth
def api_inventory_history():
    """API endpoint to get inventory change history"""
    conn = None
    try:
        conn = db_manager.get_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Get query parameters
        limit = request.args.get('limit', 100, type=int)
        inventory_id = request.args.get('inventory_id', type=int)
        job = request.args.get('job')
        change_type = request.args.get('change_type')
        changed_by = request.args.get('changed_by')

        # Build query
        query = "SELECT * FROM pcb_inventory.v_inventory_full_history WHERE 1=1"
        params = []

        if inventory_id:
            query += " AND inventory_id = %s"
            params.append(inventory_id)

        if job:
            query += " AND job = %s"
            params.append(job)

        if change_type:
            query += " AND change_type = %s"
            params.append(change_type)

        if changed_by:
            query += " AND changed_by = %s"
            params.append(changed_by)

        query += " ORDER BY change_timestamp DESC LIMIT %s"
        params.append(limit)

        cur.execute(query, tuple(params))
        history = cur.fetchall()

        # Format dates for JSON serialization
        for record in history:
            if record.get('change_timestamp'):
                record['change_timestamp'] = record['change_timestamp'].isoformat()
            if record.get('inventory_created_at'):
                record['inventory_created_at'] = record['inventory_created_at'].isoformat()
            if record.get('inventory_updated_at'):
                record['inventory_updated_at'] = record['inventory_updated_at'].isoformat()

        cur.close()
        return jsonify({'success': True, 'data': history, 'total': len(history)})

    except Exception as e:
        logger.error(f"Error fetching inventory history: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/api/inventory/history/job/<job_number>', methods=['GET'])
@require_auth
def api_job_history(job_number):
    """API endpoint to get complete history for a specific job"""
    conn = None
    try:
        conn = db_manager.get_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute(
            "SELECT * FROM pcb_inventory.get_job_history(%s)",
            (job_number,)
        )
        history = cur.fetchall()

        # Format dates for JSON serialization
        for record in history:
            if record.get('change_timestamp'):
                record['change_timestamp'] = record['change_timestamp'].isoformat()

        cur.close()
        return jsonify({'success': True, 'job': job_number, 'history': history})

    except Exception as e:
        logger.error(f"Error fetching job history: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/api/inventory/history/pcn-assignments', methods=['GET'])
@require_auth
def api_pcn_assignment_history():
    """API endpoint to get all PCN assignments from inventory history"""
    conn = None
    try:
        conn = db_manager.get_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Query PCN assignments directly from tblTransaction
        cur.execute("""
            SELECT
                pcn,
                item,
                mpn,
                trantype as assignment_type,
                created_at as assigned_at,
                userid as assigned_by
            FROM pcb_inventory."tblTransaction"
            WHERE trantype = 'PCN Generation'
            ORDER BY created_at DESC
        """)
        assignments = cur.fetchall()

        # Format dates for JSON serialization
        for record in assignments:
            if record.get('assigned_at'):
                record['assigned_at'] = record['assigned_at'].isoformat()

        cur.close()
        return jsonify({'success': True, 'assignments': assignments})

    except Exception as e:
        logger.error(f"Error fetching PCN assignments: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)

# ============================================================================
# BOM LOADER - Clean Implementation
# ============================================================================

@app.route('/bom-loader')
@require_auth
def bom_loader():
    """BOM Loader page"""
    response = make_response(render_template('bom_loader.html'))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/api/bom/parse', methods=['POST'])
@require_auth
def api_bom_parse():
    """Parse Excel BOM file and return preview data"""
    conn = None
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'Only Excel files are supported'}), 400

        # Parse Excel file
        import openpyxl
        import io

        # Read file content and validate
        file_content = file.read()

        # Validate file size
        if len(file_content) == 0:
            return jsonify({'success': False, 'error': 'File is empty'}), 400
        if len(file_content) > 100 * 1024 * 1024:
            return jsonify({'success': False, 'error': 'File exceeds 100MB limit'}), 413

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        except Exception as e:
            logger.error(f"Failed to parse Excel file: {e}")
            return jsonify({'success': False, 'error': 'Invalid or corrupted Excel file. Please check the file and try again.'}), 400

        # Check for "BOM to Load" sheet
        if "BOM to Load" not in wb.sheetnames:
            return jsonify({'success': False, 'error': 'File must contain "BOM to Load" sheet'}), 400

        ws = wb["BOM to Load"]

        # Validate sheet has enough rows (header + at least 1 data row)
        if ws.max_row < 2:
            return jsonify({'success': False, 'error': 'BOM sheet appears empty or incomplete.'}), 400

        # Row 1 has column headers
        col_map = {}
        raw_headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
        logger.info(f"BOM Excel raw headers: {raw_headers}")
        for cell in ws[1]:
            if not cell.value:
                continue
            header = str(cell.value).strip().upper()
            col_idx = cell.column - 1

            # Normalize header: collapse whitespace
            header_norm = ' '.join(header.split())

            if header_norm == 'LINE' or header_norm == 'LINE #' or header_norm == 'LINE NO':
                col_map['line'] = col_idx
            elif 'DESC' in header_norm:
                col_map['desc'] = col_idx
            elif header_norm in ('MAN', 'MANUFACTURER', 'MFG', 'MFR'):
                col_map['man'] = col_idx
            elif header_norm in ('MPN', 'MFG PN', 'MFG P/N', 'MANUFACTURER PN', 'MFR PN',
                                  'MFR P/N', 'MFGPN', 'MANUFACTURER PART NUMBER',
                                  'MFG PART NUMBER', 'MFR PART NUMBER', 'MANUFACTURER PART NO',
                                  'MFG PART NO', 'MFG PART', 'MFRPN') or (
                    'MPN' in header_norm or 'MFG P' in header_norm or 'MFR P' in header_norm or
                    ('MANUFACTURER' in header_norm and 'PART' in header_norm)):
                col_map['mpn'] = col_idx
            elif 'ACI' in header_norm:
                col_map['aci_pn'] = col_idx
            elif 'QTY' in header_norm or header_norm == 'QUANTITY':
                col_map['qty'] = col_idx
            elif 'POU' in header_norm:
                col_map['pou'] = col_idx
            elif 'LOC' in header_norm and 'LAST' not in header_norm:
                col_map['loc'] = col_idx
            elif 'COST' in header_norm or 'PRICE' in header_norm:
                col_map['cost'] = col_idx
            elif 'JOB REV' in header_norm or header_norm == 'JOBREV' or header_norm == 'JOB REVISION':
                col_map['job_rev'] = col_idx
            elif 'LAST REV' in header_norm or header_norm == 'LASTREV' or header_norm == 'LAST REVISION':
                col_map['last_rev'] = col_idx
            elif 'CUST REV' in header_norm or header_norm == 'CUSTREV' or header_norm == 'CUSTOMER REV':
                col_map['cust_rev'] = col_idx
            elif 'CUST PN' in header_norm or 'CUST P/N' in header_norm or header_norm == 'CUSTOMER PN' or header_norm == 'CUSTOMER P/N':
                col_map['cust_pn'] = col_idx
            elif header_norm in ('CUST', 'CUSTOMER') and 'cust' not in col_map:
                col_map['cust'] = col_idx
            elif header_norm == 'JOB' and 'job' not in col_map:
                col_map['job'] = col_idx
            elif header_norm == 'REV' and 'job_rev' not in col_map:
                col_map['job_rev'] = col_idx

        logger.info(f"BOM to Load column map: {col_map}")

        # Parse BOM items and extract metadata from data rows
        # Keep scanning rows until all metadata fields are found
        metadata = {}
        bom_items = []
        metadata_fields = {
            'job': 'job', 'job_rev': 'job_rev', 'last_rev': 'last_rev',
            'cust': 'customer', 'cust_pn': 'cust_pn', 'cust_rev': 'cust_rev'
        }
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            line_val = row[col_map.get('line', 0)] if 'line' in col_map else None
            if not line_val:
                continue

            try:
                line_num = int(line_val)
            except (ValueError, TypeError):
                logger.debug(f"Skipping row with invalid line number: {line_val}")
                continue

            mpn = row[col_map.get('mpn', 3)] if 'mpn' in col_map else None
            if mpn:
                mpn = str(mpn).strip()
            else:
                mpn = ''

            # Extract metadata — fill in any missing fields from each row
            for col_key, meta_key in metadata_fields.items():
                if meta_key not in metadata and col_key in col_map:
                    cell_val = row[col_map[col_key]]
                    if cell_val:
                        metadata[meta_key] = str(cell_val).strip()

            # Handle qty with proper validation
            qty_val = row[col_map.get('qty', 5)] if 'qty' in col_map else None
            try:
                qty = int(qty_val) if qty_val and str(qty_val).strip() else 0
            except (ValueError, TypeError):
                qty = 0

            # Handle cost with proper validation
            cost_val = row[col_map.get('cost', 8)] if 'cost' in col_map else None
            try:
                cost = float(cost_val) if cost_val and str(cost_val).strip() else 0.0
            except (ValueError, TypeError):
                cost = 0.0

            # Extract per-row metadata fields
            row_job = str(row[col_map['job']] or '').strip() if 'job' in col_map and row[col_map['job']] else ''
            row_job_rev = str(row[col_map['job_rev']] or '').strip() if 'job_rev' in col_map and row[col_map['job_rev']] else ''
            row_last_rev = str(row[col_map['last_rev']] or '').strip() if 'last_rev' in col_map and row[col_map['last_rev']] else ''
            row_cust = str(row[col_map['cust']] or '').strip() if 'cust' in col_map and row[col_map['cust']] else ''
            row_cust_pn = str(row[col_map['cust_pn']] or '').strip() if 'cust_pn' in col_map and row[col_map['cust_pn']] else ''
            row_cust_rev = str(row[col_map['cust_rev']] or '').strip() if 'cust_rev' in col_map and row[col_map['cust_rev']] else ''

            bom_items.append({
                'line': line_num,
                'desc': str(row[col_map.get('desc', 1)] or '').strip() if 'desc' in col_map else '',
                'man': str(row[col_map.get('man', 2)] or '').strip() if 'man' in col_map else '',
                'mpn': str(mpn).strip(),
                'aci_pn': str(row[col_map.get('aci_pn', 4)] or '').strip() if 'aci_pn' in col_map else '',
                'qty': qty,
                'pou': str(row[col_map.get('pou', 6)] or '').strip() if 'pou' in col_map else '',
                'loc': str(row[col_map.get('loc', 7)] or '').strip() if 'loc' in col_map else '',
                'cost': cost,
                'job': row_job,
                'job_rev': row_job_rev,
                'last_rev': row_last_rev,
                'cust': row_cust,
                'cust_pn': row_cust_pn,
                'cust_rev': row_cust_rev
            })

        # Ensure job number is string (Excel may return int for numeric jobs)
        if 'job' in metadata:
            metadata['job'] = str(metadata['job']).strip()
            # Strip trailing .0 from numeric job numbers (e.g. 7942.0 -> 7942)
            if metadata['job'].endswith('.0'):
                metadata['job'] = metadata['job'][:-2]

        logger.info(f"BOM Header metadata extracted: {metadata}")
        logger.info(f"Parsed BOM: Job={metadata.get('job')}, JobRev={metadata.get('job_rev')}, Items={len(bom_items)}")

        # Log first item's rev for debugging
        if bom_items:
            logger.info(f"First BOM item job_rev='{bom_items[0].get('job_rev')}', last_rev='{bom_items[0].get('last_rev')}'")

        return jsonify({
            'success': True,
            'metadata': metadata,
            'bom_items': bom_items,
            'total_items': len(bom_items)
        })

    except Exception as e:
        logger.error(f"Error parsing BOM: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/bom/load', methods=['POST'])
@require_auth
def api_bom_load():
    """Load parsed BOM data to database"""
    conn = None
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400

        metadata = data.get('metadata', {})
        bom_items = data.get('bom_items', [])

        job = str(metadata.get('job', ''))
        if not job:
            return jsonify({'success': False, 'error': 'Job number is required'}), 400

        if not bom_items:
            return jsonify({'success': False, 'error': 'No BOM items to load'}), 400

        # Debug: log what rev values were received
        meta_rev = metadata.get('job_rev', '')
        meta_cust_rev = metadata.get('cust_rev', '')
        first_item_rev = bom_items[0].get('job_rev', '') if bom_items else ''
        logger.info(f"BOM LOAD DEBUG: job={job}, metadata.job_rev='{meta_rev}', metadata.cust_rev='{meta_cust_rev}', first_item.job_rev='{first_item_rev}', total_items={len(bom_items)}")
        logger.info(f"BOM LOAD DEBUG: full metadata keys={list(metadata.keys())}, first item keys={list(bom_items[0].keys()) if bom_items else 'none'}")

        conn = db_manager.get_connection()
        cur = conn.cursor()

        try:
            # Delete existing BOM records for this job
            cur.execute('DELETE FROM pcb_inventory."tblBOM" WHERE job::text = %s', (job,))
            deleted_count = cur.rowcount
            logger.info(f"Deleted {deleted_count} existing BOM records for job {job}")

            # Use savepoint for atomic operation
            cur.execute("SAVEPOINT before_bom_insert")

            # Insert new BOM items (with validation)
            inserted_count = 0
            for item in bom_items:
                # Log items without MPN but still insert them
                if not item.get('mpn'):
                    logger.warning(f"BOM line {item.get('line')}: Missing MPN (will insert with empty MPN)")

                qty = item.get('qty', 0)
                if qty < 0:
                    logger.warning(f"Line {item.get('line')}: Negative quantity {qty}, using 0")
                    qty = 0

                cost = item.get('cost', 0.0)
                if cost < 0:
                    logger.warning(f"Line {item.get('line')}: Negative cost {cost}, using 0")
                    cost = 0.0
                try:
                    cur.execute("""
                        INSERT INTO pcb_inventory."tblBOM"
                        (job, line, "DESC", man, mpn, aci_pn, qty, pou, loc, cost,
                         job_rev, last_rev, cust, cust_pn, cust_rev, date_loaded)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', 'MM/DD/YYYY HH24:MI:SS'))
                    """, (
                        job,
                        item.get('line'),
                        item.get('desc'),
                        item.get('man'),
                        item.get('mpn'),
                        item.get('aci_pn'),
                        qty,  # Use validated quantity
                        item.get('pou'),
                        item.get('loc'),
                        cost,  # Use validated cost
                        metadata.get('job_rev') or item.get('job_rev', ''),
                        metadata.get('last_rev') or item.get('last_rev', ''),
                        metadata.get('customer') or item.get('cust', ''),
                        metadata.get('cust_pn') or item.get('cust_pn', ''),
                        metadata.get('cust_rev') or item.get('cust_rev', '')
                    ))
                    inserted_count += 1
                except Exception as e:
                    logger.error(f"Failed to insert BOM line {item.get('line')}: {e}")
                    cur.execute("ROLLBACK TO SAVEPOINT before_bom_insert")
                    raise Exception(f"Failed to insert line {item.get('line')}: {e}")

            # Upsert tblJob record
            username = session.get('username', 'Unknown')
            build_qty_val = metadata.get('build_qty')
            try:
                build_qty_int = int(build_qty_val) if build_qty_val else 1
            except (ValueError, TypeError):
                build_qty_int = 1

            cur.execute("""
                INSERT INTO pcb_inventory."tblJob"
                (job_number, customer, cust_pn, build_qty, job_rev, cust_rev, last_rev,
                 wo_number, notes, status, created_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'New', %s)
                ON CONFLICT (job_number) DO UPDATE SET
                    customer = EXCLUDED.customer,
                    cust_pn = EXCLUDED.cust_pn,
                    build_qty = EXCLUDED.build_qty,
                    job_rev = EXCLUDED.job_rev,
                    cust_rev = EXCLUDED.cust_rev,
                    last_rev = EXCLUDED.last_rev,
                    wo_number = EXCLUDED.wo_number,
                    notes = EXCLUDED.notes,
                    updated_at = CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York'
                RETURNING id
            """, (
                job,
                metadata.get('customer', ''),
                metadata.get('cust_pn', ''),
                build_qty_int,
                metadata.get('job_rev', ''),
                metadata.get('cust_rev', ''),
                metadata.get('last_rev', ''),
                metadata.get('wo_number', ''),
                metadata.get('notes', ''),
                username
            ))
            job_id = cur.fetchone()[0]
            logger.info(f"Upserted tblJob record for {job} (id={job_id})")

            conn.commit()
            logger.info(f"Loaded {inserted_count} BOM items for Job {job}")
            log_user_activity('BOM_UPLOAD', f"Loaded BOM for job {job}", f"{inserted_count} items loaded")

            return jsonify({
                'success': True,
                'message': 'BOM loaded successfully',
                'job': job,
                'job_id': job_id,
                'inserted_count': inserted_count,
                'deleted_count': deleted_count
            })

        finally:
            if cur:
                try:
                    cur.close()
                except Exception as e:
                    logger.error(f"Error closing cursor: {e}")

    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error loading BOM: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/history')
@require_auth
def inventory_history_page():
    """Inventory history page showing all changes"""
    return render_template('history.html')

# Admin Login Notifications
@app.route('/admin/notifications')
@require_auth
def admin_notifications():
    """Admin page to view activity notifications.
    Admins see all activity. Theresa sees only James's activity."""
    is_theresa = session.get('username', '').lower() in MANAGE_AUTHORIZED_USERS
    if not is_admin_user() and not is_theresa:
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        if is_theresa and not is_admin_user():
            # Theresa only sees James's activity
            cursor.execute("""
                SELECT id, user_id, username, full_name, action_type, description, details, created_at, seen, seen_at
                FROM pcb_inventory."tblActivityLog"
                WHERE username = 'james@americancircuits.com'
                ORDER BY created_at DESC
                LIMIT 200
            """)
        else:
            # Admins see all activity
            cursor.execute("""
                SELECT id, user_id, username, full_name, action_type, description, details, created_at, seen, seen_at
                FROM pcb_inventory."tblActivityLog"
                ORDER BY created_at DESC
                LIMIT 200
            """)
        notifications = cursor.fetchall()

        if is_theresa and not is_admin_user():
            cursor.execute("""
                SELECT COUNT(*) as count FROM pcb_inventory."tblActivityLog"
                WHERE seen = FALSE AND username = 'james@americancircuits.com'
            """)
        else:
            cursor.execute("""
                SELECT COUNT(*) as count FROM pcb_inventory."tblActivityLog" WHERE seen = FALSE
            """)
        unseen_count = cursor.fetchone()['count']

        return render_template('admin_notifications.html',
                             notifications=notifications,
                             unseen_count=unseen_count)

    except Exception as e:
        logger.error(f"Error loading admin notifications: {e}")
        flash('Error loading notifications.', 'danger')
        return redirect(url_for('index'))
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/admin/notifications/mark-seen', methods=['POST'])
@require_auth
def mark_notifications_seen():
    """Mark all notifications as seen."""
    if not is_admin_user():
        return jsonify({'success': False, 'error': 'Admin access required'}), 403

    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            UPDATE pcb_inventory."tblActivityLog"
            SET seen = TRUE, seen_at = CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York'
            WHERE seen = FALSE
        """)
        conn.commit()

        return jsonify({'success': True, 'message': 'All notifications marked as seen'})

    except Exception as e:
        logger.error(f"Error marking notifications as seen: {e}")
        if conn:
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/admin/notifications/clear', methods=['POST'])
@require_auth
def clear_notifications():
    """Clear all old notifications (keep last 7 days)."""
    if not is_admin_user():
        return jsonify({'success': False, 'error': 'Admin access required'}), 403

    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            DELETE FROM pcb_inventory."tblActivityLog"
            WHERE created_at < NOW() - INTERVAL '7 days'
        """)
        deleted_count = cursor.rowcount
        conn.commit()

        return jsonify({'success': True, 'message': f'Cleared {deleted_count} old notifications'})

    except Exception as e:
        logger.error(f"Error clearing notifications: {e}")
        if conn:
            conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)

@app.route('/api/admin/notification-count')
@require_auth
def get_notification_count():
    """Get count of unseen notifications. Theresa sees only James's count."""
    is_theresa = session.get('username', '').lower() in MANAGE_AUTHORIZED_USERS
    if not is_admin_user() and not is_theresa:
        return jsonify({'count': 0})

    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        if is_theresa and not is_admin_user():
            cursor.execute("""
                SELECT COUNT(*) as count FROM pcb_inventory."tblActivityLog"
                WHERE seen = FALSE AND username = 'james@americancircuits.com'
            """)
        else:
            cursor.execute("""
                SELECT COUNT(*) as count FROM pcb_inventory."tblActivityLog" WHERE seen = FALSE
            """)
        result = cursor.fetchone()

        return jsonify({'count': result['count']})

    except Exception as e:
        logger.error(f"Error getting notification count: {e}")
        return jsonify({'count': 0})
    finally:
        if conn:
            db_manager.return_connection(conn)

# ==================== Jobs Management ====================

@app.route('/jobs')
@require_auth
def jobs_list():
    """Jobs list page with search."""
    search_query = request.args.get('q', '').strip()
    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        if search_query:
            cursor.execute("""
                SELECT id, job_number, description, customer, cust_pn, build_qty,
                       job_rev, status, created_by, created_at
                FROM pcb_inventory."tblJob"
                WHERE job_number ILIKE %s OR customer ILIKE %s OR description ILIKE %s
                ORDER BY created_at DESC
            """, (f'%{search_query}%', f'%{search_query}%', f'%{search_query}%'))
        else:
            cursor.execute("""
                SELECT id, job_number, description, customer, cust_pn, build_qty,
                       job_rev, status, created_by, created_at
                FROM pcb_inventory."tblJob"
                ORDER BY created_at DESC
            """)
        jobs = cursor.fetchall()

        return render_template('jobs.html', jobs=jobs, search_query=search_query)

    except Exception as e:
        logger.error(f"Error loading jobs list: {e}")
        flash('Error loading jobs. Please try again.', 'danger')
        return render_template('jobs.html', jobs=[], search_query=search_query)
    finally:
        if conn:
            db_manager.return_connection(conn)


@app.route('/jobs/<job_number>')
@require_auth
def job_detail(job_number):
    """Job detail page with live inventory lookup."""
    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Get job record
        cursor.execute("""
            SELECT * FROM pcb_inventory."tblJob"
            WHERE job_number = %s
        """, (job_number,))
        job = cursor.fetchone()

        if not job:
            flash(f'Job {job_number} not found.', 'danger')
            return redirect(url_for('jobs_list'))

        build_qty = int(job['build_qty'] or 1)
        order_qty = int(job['order_qty'] or 1)

        # Get latest revision for this job
        cursor.execute("""
            SELECT job_rev FROM pcb_inventory."tblBOM"
            WHERE job = %s AND job_rev IS NOT NULL AND job_rev != ''
            ORDER BY created_at DESC LIMIT 1
        """, (job_number,))
        rev_row = cursor.fetchone()
        job_rev = rev_row['job_rev'] if rev_row else None

        # Get BOM lines with live inventory lookup
        # First deduplicate BOM per aci_pn, then match inventory using warehouse MPN
        cursor.execute("""
            WITH bom_lines AS (
                SELECT DISTINCT ON (b.aci_pn)
                    b.line,
                    b.aci_pn,
                    b."DESC",
                    b.mpn as bom_mpn,
                    b.man,
                    b.qty,
                    b.cost,
                    b.pou,
                    b.job_rev,
                    b.last_rev,
                    b.cust,
                    b.cust_pn,
                    b.cust_rev
                FROM pcb_inventory."tblBOM" b
                WHERE b.job = %s
                    AND (b.job_rev = (SELECT job_rev FROM pcb_inventory."tblBOM" WHERE job = %s AND job_rev IS NOT NULL AND job_rev != '' ORDER BY created_at DESC LIMIT 1)
                         OR NOT EXISTS (SELECT 1 FROM pcb_inventory."tblBOM" WHERE job = %s AND job_rev IS NOT NULL AND job_rev != ''))
                ORDER BY b.aci_pn, b.line
            ),
            inventory_match AS (
                SELECT DISTINCT ON (COALESCE(w.pcn, bl.aci_pn || '_nopcn'), bl.aci_pn)
                    bl.line,
                    bl.aci_pn,
                    bl."DESC",
                    COALESCE(w.mpn, bl.bom_mpn) as mpn,
                    bl.man,
                    bl.qty,
                    bl.cost,
                    bl.pou,
                    bl.job_rev,
                    bl.last_rev,
                    bl.cust,
                    bl.cust_pn,
                    bl.cust_rev,
                    w.pcn,
                    COALESCE(w.item, bl.aci_pn) as item,
                    COALESCE(w.onhandqty, 0) as onhandqty,
                    w.loc_to,
                    CASE WHEN bl.aci_pn = w.item THEN 1 WHEN w.item IS NOT NULL THEN 2 ELSE 3 END as match_priority
                FROM bom_lines bl
                LEFT JOIN pcb_inventory."tblWhse_Inventory" w
                    ON (bl.aci_pn = w.item OR bl.bom_mpn = w.mpn)
                    AND COALESCE(w.loc_to, '') != 'MFG Floor'
                ORDER BY COALESCE(w.pcn, bl.aci_pn || '_nopcn'), bl.aci_pn, match_priority
            )
            SELECT
                line as line_no,
                aci_pn,
                "DESC" as description,
                mpn,
                man as manufacturer,
                CAST(COALESCE(NULLIF(qty, ''), '0') AS INTEGER) as qty,
                COALESCE(SUM(onhandqty), 0) as on_hand,
                pcn,
                item,
                COALESCE(loc_to, '') as location,
                CAST(COALESCE(NULLIF(cost, ''), '0') AS DECIMAL(10,4)) as unit_cost,
                pou,
                job_rev as bom_job_rev,
                last_rev as bom_last_rev,
                cust as bom_cust,
                cust_pn as bom_cust_pn,
                cust_rev as bom_cust_rev
            FROM inventory_match
            GROUP BY line, aci_pn, "DESC", mpn, man, qty, cost, pou, job_rev, last_rev, cust, cust_pn, cust_rev, pcn, item, loc_to
            ORDER BY
                CASE WHEN line ~ '^[0-9]+$' THEN CAST(line AS INTEGER) ELSE 999999 END,
                line
        """, (job_number, job_number, job_number))
        raw_lines = cursor.fetchall()

        # Calculate REQ and shortage for each line
        # REQ = QTY per board × Order QTY (the actual production quantity)
        job_lines = []
        shortage_count = 0
        for line in raw_lines:
            qty = int(line['qty'] or 0)
            req = qty * order_qty
            on_hand = int(line['on_hand'] or 0)
            shortage = on_hand - req
            location = line['location'] if on_hand > 0 else ''

            job_lines.append({
                'line_no': line['line_no'],
                'aci_pn': line['aci_pn'],
                'description': line['description'],
                'mpn': line['mpn'],
                'manufacturer': line['manufacturer'],
                'qty': qty,
                'req': req,
                'on_hand': on_hand,
                'pcn': line['pcn'],
                'item': line['item'],
                'location': location,
                'unit_cost': float(line['unit_cost'] or 0),
                'shortage': shortage,
                'pou': line.get('pou') or '',
                'job_rev': line.get('bom_job_rev') or '',
                'last_rev': line.get('bom_last_rev') or '',
                'cust': line.get('bom_cust') or '',
                'cust_pn': line.get('bom_cust_pn') or '',
                'cust_rev': line.get('bom_cust_rev') or ''
            })
            if shortage < 0:
                shortage_count += 1

        # Compute status from inventory data
        total_bom_lines = len(set(l['line_no'] for l in raw_lines))
        lines_with_location = len(set(l['line_no'] for l in job_lines if l['location']))

        if lines_with_location == 0:
            computed_status = 'New'
        elif lines_with_location < total_bom_lines:
            computed_status = 'In Prep'
        else:
            computed_status = 'In Mfg'

        # Update status if changed
        if computed_status != job['status']:
            cursor.execute("""
                UPDATE pcb_inventory."tblJob"
                SET status = %s, updated_at = CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York'
                WHERE job_number = %s
            """, (computed_status, job_number))
            conn.commit()
            job['status'] = computed_status

        # Fetch related shortage reports for this job
        cursor.execute("""
            SELECT id, report_name, order_qty, total_lines, shortage_lines,
                   created_by, created_at
            FROM pcb_inventory."tblShortageReport"
            WHERE job = %s
            ORDER BY created_at DESC
        """, (job_number,))
        related_reports = cursor.fetchall()

        return render_template('job_detail.html',
                             job=job,
                             lines=job_lines,
                             shortage_count=shortage_count,
                             total_lines=total_bom_lines,
                             related_reports=related_reports,
                             job_rev=job_rev,
                             column_definitions=SHORTAGE_EXPORT_COLUMNS)

    except Exception as e:
        logger.error(f"Error loading job detail: {e}")
        flash('Error loading job. Please try again.', 'danger')
        return redirect(url_for('jobs_list'))
    finally:
        if conn:
            db_manager.return_connection(conn)


@app.route('/jobs/<job_number>/update-build-qty', methods=['POST'])
@require_auth
def job_update_build_qty(job_number):
    """Update build quantity for a job."""
    conn = None
    try:
        data = request.get_json()
        build_qty = int(data.get('build_qty', 1))
        if build_qty < 1:
            build_qty = 1

        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE pcb_inventory."tblJob"
            SET build_qty = %s, updated_at = CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York'
            WHERE job_number = %s
        """, (build_qty, job_number))
        conn.commit()

        return jsonify({'success': True, 'build_qty': build_qty})
    except Exception as e:
        logger.error(f"Error updating build qty: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)


@app.route('/jobs/<job_number>/update-order-qty', methods=['POST'])
@require_auth
def job_update_order_qty(job_number):
    """Update order quantity for a job."""
    conn = None
    try:
        data = request.get_json()
        order_qty = int(data.get('order_qty', 1))
        if order_qty < 1:
            order_qty = 1

        conn = db_manager.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE pcb_inventory."tblJob"
            SET order_qty = %s, updated_at = CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York'
            WHERE job_number = %s
        """, (order_qty, job_number))
        conn.commit()

        return jsonify({'success': True, 'order_qty': order_qty})
    except Exception as e:
        logger.error(f"Error updating order qty: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)


@app.route('/jobs/<job_number>/generate-shortage', methods=['POST'])
@require_auth
def job_generate_shortage(job_number):
    """Generate a shortage report from the job page."""
    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Get job record
        cursor.execute("""
            SELECT * FROM pcb_inventory."tblJob" WHERE job_number = %s
        """, (job_number,))
        job = cursor.fetchone()

        if not job:
            flash(f'Job {job_number} not found.', 'danger')
            return redirect(url_for('jobs_list'))

        order_qty = int(job['order_qty'] or 1)

        # Default report name
        report_name = f"Shortage Report - {job_number} - {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Get latest revision for this job
        cursor.execute("""
            SELECT job_rev FROM pcb_inventory."tblBOM"
            WHERE job = %s AND job_rev IS NOT NULL AND job_rev != ''
            ORDER BY created_at DESC LIMIT 1
        """, (job_number,))
        rev_row = cursor.fetchone()
        job_rev = rev_row['job_rev'] if rev_row else None

        # Get total BOM line count (latest rev only)
        if job_rev:
            cursor.execute("SELECT COUNT(*) as count FROM pcb_inventory.\"tblBOM\" WHERE job = %s AND job_rev = %s", (job_number, job_rev))
        else:
            cursor.execute("SELECT COUNT(*) as count FROM pcb_inventory.\"tblBOM\" WHERE job = %s", (job_number,))
        total_bom_lines = cursor.fetchone()['count']

        if total_bom_lines == 0:
            flash(f'No BOM data found for job {job_number}. Please load BOM first.', 'warning')
            return redirect(url_for('job_detail', job_number=job_number))

        # Deduplicate BOM per aci_pn then match inventory using warehouse MPN
        cursor.execute("""
            WITH bom_lines AS (
                SELECT DISTINCT ON (b.aci_pn)
                    b.line,
                    b.aci_pn,
                    b.mpn as bom_mpn,
                    b.man,
                    b."DESC",
                    b.qty,
                    b.cost
                FROM pcb_inventory."tblBOM" b
                WHERE b.job = %s
                    AND (b.job_rev = (SELECT job_rev FROM pcb_inventory."tblBOM" WHERE job = %s AND job_rev IS NOT NULL AND job_rev != '' ORDER BY created_at DESC LIMIT 1)
                         OR NOT EXISTS (SELECT 1 FROM pcb_inventory."tblBOM" WHERE job = %s AND job_rev IS NOT NULL AND job_rev != ''))
                ORDER BY b.aci_pn, b.line
            ),
            inventory_match AS (
                SELECT DISTINCT ON (COALESCE(w.pcn, bl.aci_pn || '_nopcn'), bl.aci_pn)
                    bl.line,
                    bl.aci_pn,
                    COALESCE(w.mpn, bl.bom_mpn) as mpn,
                    bl.man,
                    bl."DESC",
                    bl.qty,
                    bl.cost,
                    w.pcn,
                    COALESCE(w.item, bl.aci_pn) as item,
                    COALESCE(w.onhandqty, 0) as onhandqty,
                    w.loc_to,
                    CASE WHEN bl.aci_pn = w.item THEN 1 WHEN w.item IS NOT NULL THEN 2 ELSE 3 END as match_priority
                FROM bom_lines bl
                LEFT JOIN pcb_inventory."tblWhse_Inventory" w
                    ON (bl.aci_pn = w.item OR bl.bom_mpn = w.mpn)
                    AND COALESCE(w.loc_to, '') != 'MFG Floor'
                ORDER BY COALESCE(w.pcn, bl.aci_pn || '_nopcn'), bl.aci_pn, match_priority
            )
            SELECT
                line as line_no,
                aci_pn,
                pcn,
                mpn,
                CAST(COALESCE(NULLIF(qty, ''), '0') AS INTEGER) as qty,
                COALESCE(SUM(onhandqty), 0) as qty_on_hand,
                item,
                COALESCE(loc_to, '') as location,
                CAST(COALESCE(NULLIF(cost, ''), '0') AS DECIMAL(10,4)) as unit_cost,
                man as manufacturer,
                "DESC" as description
            FROM inventory_match
            GROUP BY line, aci_pn, mpn, man, "DESC", qty, cost, pcn, item, loc_to
            ORDER BY
                CASE WHEN line ~ '^[0-9]+$' THEN CAST(line AS INTEGER) ELSE 999999 END,
                line
        """, (job_number, job_number, job_number))
        matched_items = cursor.fetchall()

        if not matched_items:
            flash(f'No BOM data found for job {job_number}.', 'warning')
            return redirect(url_for('job_detail', job_number=job_number))

        # Calculate REQ and count shortages
        report_items = []
        shortage_count = 0
        for item in matched_items:
            qty = int(item['qty'] or 0)
            req = qty * order_qty
            on_hand = int(item['qty_on_hand'] or 0)
            item['req'] = req
            item['order_qty'] = order_qty
            if on_hand < req:
                shortage_count += 1
            report_items.append(item)

        total_cost = sum(float(item['qty'] or 0) * float(item['unit_cost'] or 0) for item in report_items)
        shortage_cost = sum(float(item['req'] or 0) * float(item['unit_cost'] or 0) for item in report_items if item['qty_on_hand'] < item['req'])

        username = session.get('username', 'Unknown')
        cursor.execute("""
            INSERT INTO pcb_inventory."tblShortageReport"
            (job, report_name, total_lines, shortage_lines, total_cost, shortage_cost, created_by, notes, order_qty, job_rev)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (job_number, report_name, total_bom_lines, shortage_count, total_cost, shortage_cost, username, '', order_qty, job_rev))
        report_id = cursor.fetchone()['id']

        for item in report_items:
            cursor.execute("""
                INSERT INTO pcb_inventory."tblShortageReportItems"
                (report_id, line_no, aci_pn, pcn, mpn, qty_required, qty_on_hand, order_qty,
                 item, location, unit_cost, line_cost, manufacturer, description, req)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                report_id, item['line_no'], item['aci_pn'], item['pcn'], item['mpn'],
                item['qty'], item['qty_on_hand'], order_qty,
                item['item'], item['location'], float(item['unit_cost'] or 0),
                float(item['qty'] or 0) * float(item['unit_cost'] or 0),
                item['manufacturer'], item['description'], item['req']
            ))

        conn.commit()
        log_user_activity('SHORTAGE_REPORT', f"Generated shortage report for job {job_number}", f"{len(report_items)} items, {shortage_count} shortages")
        flash(f'Shortage report generated! {len(report_items)} items matched, {shortage_count} shortages.', 'success')
        return redirect(url_for('job_detail', job_number=job_number))

    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error generating shortage report from job: {e}")
        flash('Error generating report. Please try again.', 'danger')
        return redirect(url_for('job_detail', job_number=job_number))
    finally:
        if conn:
            db_manager.return_connection(conn)


@app.route('/jobs/<job_number>/export', methods=['GET', 'POST'])
@require_auth
def job_export(job_number):
    """Export job report to Excel with optional column customization."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    conn = None
    try:
        # Parse column config from POST body, or use defaults
        if request.method == 'POST' and request.is_json:
            config = request.get_json()
            selected_columns = config.get('columns', [])
            highlighted_columns = set(config.get('highlighted', []))
            export_filter = config.get('filter', 'all')
        else:
            selected_columns = [c['key'] for c in SHORTAGE_EXPORT_COLUMNS if c['default']]
            highlighted_columns = set()
            export_filter = 'all'

        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Get job record
        cursor.execute("""
            SELECT * FROM pcb_inventory."tblJob" WHERE job_number = %s
        """, (job_number,))
        job = cursor.fetchone()

        if not job:
            return jsonify({'success': False, 'error': f'Job {job_number} not found.'}), 404

        build_qty = int(job['build_qty'] or 1)
        order_qty = int(job['order_qty'] or 1)

        # Get latest revision
        cursor.execute("""
            SELECT job_rev FROM pcb_inventory."tblBOM"
            WHERE job = %s AND job_rev IS NOT NULL AND job_rev != ''
            ORDER BY created_at DESC LIMIT 1
        """, (job_number,))
        rev_row = cursor.fetchone()
        job_rev = rev_row['job_rev'] if rev_row else None

        # Deduplicate BOM per aci_pn then match inventory using warehouse MPN
        cursor.execute("""
            WITH bom_lines AS (
                SELECT DISTINCT ON (b.aci_pn)
                    b.line,
                    b.aci_pn,
                    b.mpn as bom_mpn,
                    b.man,
                    b."DESC",
                    b.qty,
                    b.cost
                FROM pcb_inventory."tblBOM" b
                WHERE b.job = %s
                    AND (b.job_rev = (SELECT job_rev FROM pcb_inventory."tblBOM" WHERE job = %s AND job_rev IS NOT NULL AND job_rev != '' ORDER BY created_at DESC LIMIT 1)
                         OR NOT EXISTS (SELECT 1 FROM pcb_inventory."tblBOM" WHERE job = %s AND job_rev IS NOT NULL AND job_rev != ''))
                ORDER BY b.aci_pn, b.line
            ),
            inventory_match AS (
                SELECT DISTINCT ON (COALESCE(w.pcn, bl.aci_pn || '_nopcn'), bl.aci_pn)
                    bl.line,
                    bl.aci_pn,
                    COALESCE(w.mpn, bl.bom_mpn) as mpn,
                    bl.man,
                    bl."DESC",
                    bl.qty,
                    bl.cost,
                    w.pcn,
                    COALESCE(w.item, bl.aci_pn) as item,
                    COALESCE(w.onhandqty, 0) as onhandqty,
                    w.loc_to,
                    CASE WHEN bl.aci_pn = w.item THEN 1 WHEN w.item IS NOT NULL THEN 2 ELSE 3 END as match_priority
                FROM bom_lines bl
                LEFT JOIN pcb_inventory."tblWhse_Inventory" w
                    ON (bl.aci_pn = w.item OR bl.bom_mpn = w.mpn)
                    AND COALESCE(w.loc_to, '') != 'MFG Floor'
                ORDER BY COALESCE(w.pcn, bl.aci_pn || '_nopcn'), bl.aci_pn, match_priority
            )
            SELECT
                line as line_no,
                aci_pn,
                mpn,
                man as manufacturer,
                "DESC" as description,
                CAST(COALESCE(NULLIF(qty, ''), '0') AS INTEGER) as qty,
                COALESCE(SUM(onhandqty), 0) as on_hand,
                pcn,
                item,
                COALESCE(loc_to, '') as location,
                CAST(COALESCE(NULLIF(cost, ''), '0') AS DECIMAL(10,4)) as unit_cost
            FROM inventory_match
            GROUP BY line, aci_pn, mpn, man, "DESC", qty, cost, pcn, item, loc_to
            ORDER BY
                CASE WHEN line ~ '^[0-9]+$' THEN CAST(line AS INTEGER) ELSE 999999 END,
                line
        """, (job_number, job_number, job_number))
        raw_items = cursor.fetchall()

        # Enrich items with calculated fields
        items = []
        for item in raw_items:
            qty = int(item['qty'] or 0)
            on_hand = int(item['on_hand'] or 0)
            req = qty * order_qty
            item['qty_on_hand'] = on_hand
            item['order_qty'] = order_qty
            item['req'] = req
            item['line_cost'] = float(qty) * float(item['unit_cost'] or 0)
            items.append(item)

        # Apply filter: shortages only
        if export_filter == 'shortages_only':
            items = [i for i in items if (i.get('qty_on_hand') or 0) < (i.get('req') or 0)]

        # Hide zero on-hand rows (default: true, matches UI toggle)
        hide_zero = True
        if request.method == 'POST' and request.is_json:
            hide_zero = config.get('hide_zero', True)
        if hide_zero:
            items = [i for i in items if (i.get('qty_on_hand') or 0) != 0]

        # Build active column list from selection
        col_registry = {c['key']: c for c in SHORTAGE_EXPORT_COLUMNS}
        active_cols = [col_registry[k] for k in selected_columns if k in col_registry]
        if not active_cols:
            active_cols = [col_registry[k] for k in [c['key'] for c in SHORTAGE_EXPORT_COLUMNS if c['default']]]
        num_cols = len(active_cols)
        last_col = get_column_letter(num_cols)

        # Build Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Report"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        shortage_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        highlight_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'] = f"Job Report - {job_number}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'A2:{last_col}2')
        ws['A2'] = f"Customer: {job.get('customer', 'N/A')} | Rev: {job_rev or 'N/A'} | Order Qty: {order_qty}"
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'A3:{last_col}3')
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Items: {len(items)}"
        ws['A3'].alignment = Alignment(horizontal='center')

        # Headers
        for col_idx, col_def in enumerate(active_cols, 1):
            cell = ws.cell(row=5, column=col_idx, value=col_def['label'])
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, item in enumerate(items, 6):
            is_shortage = (item.get('qty_on_hand') or 0) < (item.get('req') or 0)
            for col_idx, col_def in enumerate(active_cols, 1):
                value = get_export_cell_value(item, col_def['key'], order_qty)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if is_shortage:
                    cell.fill = shortage_fill
                elif col_def['key'] in highlighted_columns:
                    cell.fill = highlight_fill

        # Column widths
        for col_idx, col_def in enumerate(active_cols, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def['width']

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Job_{job_number}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        logger.error(f"Error exporting job: {e}")
        if request.method == 'POST':
            return jsonify({'success': False, 'error': str(e)}), 500
        flash('Error exporting job. Please try again.', 'danger')
        return redirect(url_for('job_detail', job_number=job_number))
    finally:
        if conn:
            db_manager.return_connection(conn)


@app.route('/jobs/<job_number>/delete', methods=['POST'])
@require_auth
def job_delete(job_number):
    """Delete a job record."""
    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor()

        cursor.execute('DELETE FROM pcb_inventory."tblJob" WHERE job_number = %s', (job_number,))
        if cursor.rowcount == 0:
            flash(f'Job {job_number} not found.', 'danger')
        else:
            conn.commit()
            flash(f'Job {job_number} deleted successfully.', 'success')

        return redirect(url_for('jobs_list'))

    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error deleting job: {e}")
        flash('Error deleting job. Please try again.', 'danger')
        return redirect(url_for('jobs_list'))
    finally:
        if conn:
            db_manager.return_connection(conn)


@app.route('/jobs/<job_number>/create-revision', methods=['POST'])
@require_auth
def job_create_revision(job_number):
    """Create a new revision by archiving old BOM lines and rebuilding the job under a new rev."""
    conn = None
    try:
        data = request.get_json()
        new_rev = data.get('new_rev', '').strip()
        if not new_rev:
            return jsonify({'success': False, 'error': 'Revision identifier is required.'}), 400

        username = session.get('username', 'unknown')
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Get the current revision
        cursor.execute("""
            SELECT job_rev FROM pcb_inventory."tblBOM"
            WHERE job = %s AND job_rev IS NOT NULL AND job_rev != ''
            ORDER BY created_at DESC LIMIT 1
        """, (job_number,))
        rev_row = cursor.fetchone()
        current_rev = rev_row['job_rev'] if rev_row else None

        # Check if new rev already exists in archive (prevent reuse of archived rev names)
        cursor.execute("""
            SELECT COUNT(*) as count FROM pcb_inventory."tblBOM_Archive"
            WHERE job = %s AND job_rev = %s
        """, (job_number, new_rev))
        if cursor.fetchone()['count'] > 0:
            return jsonify({'success': False, 'error': f'Revision {new_rev} was previously used for this job.'}), 400

        # Get current BOM lines to verify they exist
        if current_rev:
            cursor.execute("""
                SELECT id, line, "DESC", man, mpn, aci_pn, qty, pou, loc, cost,
                       job_rev, last_rev, cust, cust_pn, cust_rev, date_loaded, created_at
                FROM pcb_inventory."tblBOM"
                WHERE job = %s AND job_rev = %s
            """, (job_number, current_rev))
        else:
            cursor.execute("""
                SELECT id, line, "DESC", man, mpn, aci_pn, qty, pou, loc, cost,
                       job_rev, last_rev, cust, cust_pn, cust_rev, date_loaded, created_at
                FROM pcb_inventory."tblBOM"
                WHERE job = %s
            """, (job_number,))

        source_lines = cursor.fetchall()
        if not source_lines:
            return jsonify({'success': False, 'error': 'No BOM lines found for this job.'}), 400

        # Step 1: Archive the old BOM lines
        for line in source_lines:
            cursor.execute("""
                INSERT INTO pcb_inventory."tblBOM_Archive"
                (original_id, job, line, "DESC", man, mpn, aci_pn, qty, pou, loc, cost,
                 job_rev, last_rev, cust, cust_pn, cust_rev, date_loaded, created_at, archived_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                line['id'], job_number, line['line'], line['DESC'], line['man'], line['mpn'],
                line['aci_pn'], line['qty'], line['pou'], line['loc'], line['cost'],
                line['job_rev'], line['last_rev'], line['cust'], line['cust_pn'], line['cust_rev'],
                line['date_loaded'], line['created_at'], username
            ))

        # Step 2: Update existing BOM lines in place with new revision (no duplication)
        if current_rev:
            cursor.execute("""
                UPDATE pcb_inventory."tblBOM"
                SET job_rev = %s, last_rev = %s,
                    date_loaded = TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', 'MM/DD/YYYY HH24:MI:SS')
                WHERE job = %s AND job_rev = %s
            """, (new_rev, current_rev, job_number, current_rev))
        else:
            cursor.execute("""
                UPDATE pcb_inventory."tblBOM"
                SET job_rev = %s, last_rev = %s,
                    date_loaded = TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York', 'MM/DD/YYYY HH24:MI:SS')
                WHERE job = %s
            """, (new_rev, current_rev or '', job_number))

        # Step 3: Update tblJob with new revision info
        cursor.execute("""
            UPDATE pcb_inventory."tblJob"
            SET job_rev = %s, last_rev = %s, updated_at = CURRENT_TIMESTAMP AT TIME ZONE 'America/New_York'
            WHERE job_number = %s
        """, (new_rev, current_rev or '', job_number))

        conn.commit()
        return jsonify({'success': True, 'message': f'Revision {new_rev} created. Old revision {current_rev or "N/A"} archived. {len(source_lines)} lines updated.'})

    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error creating revision: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)


@app.route('/api/bom/job-check/<job_number>')
@require_auth
def api_job_check(job_number):
    """Check if a job already exists in tblJob."""
    conn = None
    try:
        conn = db_manager.get_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        cursor.execute("""
            SELECT id, job_number, status, build_qty, customer, job_rev
            FROM pcb_inventory."tblJob" WHERE job_number = %s
        """, (job_number,))
        job = cursor.fetchone()

        if job:
            return jsonify({'exists': True, 'job': dict(job)})
        return jsonify({'exists': False})

    except Exception as e:
        logger.error(f"Error checking job: {e}")
        return jsonify({'exists': False, 'error': str(e)})
    finally:
        if conn:
            db_manager.return_connection(conn)


# ==================== Admin User Management CRUD ====================

@app.route('/admin/users')
@require_auth
def admin_users():
    """Admin page to manage users - CRUD operations."""
    if not is_admin_user():
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    conn = None
    try:
        conn = db_manager.get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute("""
                SELECT id, userid, username, userlogin, usersecurity
                FROM pcb_inventory."tblUser"
                ORDER BY id
            """)
            users = cursor.fetchall()
        return render_template('user_management.html', users=users)
    except Exception as e:
        logger.error(f"Error loading users: {e}")
        flash('Error loading users.', 'danger')
        return redirect(url_for('index'))
    finally:
        if conn:
            db_manager.return_connection(conn)


@app.route('/admin/users/create', methods=['POST'])
@require_auth
def admin_create_user():
    """Create a new user."""
    if not is_admin_user():
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    userlogin = request.form.get('userlogin', '').strip().lower()
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    usersecurity = request.form.get('usersecurity', 'user').strip()

    # Validation
    if not userlogin:
        flash('Login username is required.', 'danger')
        return redirect(url_for('admin_users'))
    if not username:
        username = userlogin
    if not password or len(password) < 6:
        flash('Password must be at least 6 characters.', 'danger')
        return redirect(url_for('admin_users'))
    if usersecurity not in ('Admin', 'user'):
        usersecurity = 'user'

    conn = None
    try:
        conn = db_manager.get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            # Check for duplicate userlogin
            cursor.execute("""
                SELECT id FROM pcb_inventory."tblUser" WHERE userlogin = %s
            """, (userlogin,))
            if cursor.fetchone():
                flash(f'Username "{userlogin}" already exists.', 'danger')
                return redirect(url_for('admin_users'))

            # Hash password
            password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

            cursor.execute("""
                INSERT INTO pcb_inventory."tblUser" (username, userlogin, password, usersecurity)
                VALUES (%s, %s, %s, %s)
            """, (username, userlogin, password_hash, usersecurity))
            conn.commit()

        flash(f'User "{userlogin}" created successfully.', 'success')
        logger.info(f"Admin {session.get('username')} created user: {userlogin}")
    except Exception as e:
        logger.error(f"Error creating user: {e}")
        flash('Error creating user.', 'danger')
        if conn:
            conn.rollback()
    finally:
        if conn:
            db_manager.return_connection(conn)

    return redirect(url_for('admin_users'))


@app.route('/admin/users/edit/<int:user_id>', methods=['POST'])
@require_auth
def admin_edit_user(user_id):
    """Edit an existing user."""
    if not is_admin_user():
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    username = request.form.get('username', '').strip()
    usersecurity = request.form.get('usersecurity', 'user').strip()
    new_password = request.form.get('password', '').strip()

    if not username:
        flash('Full name is required.', 'danger')
        return redirect(url_for('admin_users'))
    if usersecurity not in ('Admin', 'user'):
        usersecurity = 'user'

    # Prevent admin from demoting themselves
    if user_id == session.get('user_id') and usersecurity != 'Admin':
        flash('You cannot change your own role away from Admin.', 'danger')
        return redirect(url_for('admin_users'))

    conn = None
    try:
        conn = db_manager.get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            if new_password:
                if len(new_password) < 6:
                    flash('Password must be at least 6 characters.', 'danger')
                    return redirect(url_for('admin_users'))
                password_hash = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                cursor.execute("""
                    UPDATE pcb_inventory."tblUser"
                    SET username = %s, usersecurity = %s, password = %s
                    WHERE id = %s
                """, (username, usersecurity, password_hash, user_id))
            else:
                cursor.execute("""
                    UPDATE pcb_inventory."tblUser"
                    SET username = %s, usersecurity = %s
                    WHERE id = %s
                """, (username, usersecurity, user_id))
            conn.commit()

        flash('User updated successfully.', 'success')
        logger.info(f"Admin {session.get('username')} edited user ID: {user_id}")
    except Exception as e:
        logger.error(f"Error editing user: {e}")
        flash('Error updating user.', 'danger')
        if conn:
            conn.rollback()
    finally:
        if conn:
            db_manager.return_connection(conn)

    return redirect(url_for('admin_users'))


@app.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@require_auth
def admin_delete_user(user_id):
    """Delete a user."""
    if not is_admin_user():
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    # Prevent admin from deleting themselves
    if user_id == session.get('user_id'):
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('admin_users'))

    conn = None
    try:
        conn = db_manager.get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            # Get username for logging
            cursor.execute("""
                SELECT userlogin FROM pcb_inventory."tblUser" WHERE id = %s
            """, (user_id,))
            user = cursor.fetchone()

            if not user:
                flash('User not found.', 'danger')
                return redirect(url_for('admin_users'))

            cursor.execute("""
                DELETE FROM pcb_inventory."tblUser" WHERE id = %s
            """, (user_id,))
            conn.commit()

        flash(f'User "{user["userlogin"]}" deleted successfully.', 'success')
        logger.info(f"Admin {session.get('username')} deleted user: {user['userlogin']}")
    except Exception as e:
        logger.error(f"Error deleting user: {e}")
        flash('Error deleting user.', 'danger')
        if conn:
            conn.rollback()
    finally:
        if conn:
            db_manager.return_connection(conn)

    return redirect(url_for('admin_users'))


@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    """Handle CSRF errors - return JSON for API endpoints, HTML for pages"""
    logger.error(f"CSRF validation failed: {e.description}")

    # Check if it's an API request
    if request.path.startswith('/api/'):
        return jsonify({
            'success': False,
            'error': 'CSRF token validation failed. Please refresh the page and try again.'
        }), 400

    # For non-API requests, return simple HTML error message
    return f"<h1>400 Bad Request</h1><p>{e.description}</p><p>Please refresh the page and try again.</p>", 400

@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500

# ==================== Location Management (All Users) ====================

@app.route('/admin/locations')
@require_auth
def admin_locations():
    """Page to manage warehouse locations (tblLoc). Admin + Theresa only."""
    if not can_manage_parts():
        flash('Access denied. You do not have permission to manage locations.', 'danger')
        return redirect(url_for('index'))

    conn = None
    try:
        conn = db_manager.get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute("""
                SELECT id, area, shelf, loc, location
                FROM pcb_inventory."tblLoc"
                ORDER BY location
            """)
            locations = cursor.fetchall()

            # Get unique areas for filter dropdown
            cursor.execute("""
                SELECT DISTINCT area FROM pcb_inventory."tblLoc"
                WHERE area IS NOT NULL ORDER BY area
            """)
            areas = [r['area'] for r in cursor.fetchall()]

        return render_template('location_management.html', locations=locations, areas=areas)
    except Exception as e:
        logger.error(f"Error loading locations: {e}")
        flash('Error loading locations.', 'danger')
        return redirect(url_for('index'))
    finally:
        if conn:
            db_manager.return_connection(conn)


@app.route('/admin/locations/create', methods=['POST'])
@require_auth
def admin_create_location():
    """Create a new location. Admin + Theresa only."""
    if not can_manage_parts():
        flash('Access denied.', 'danger')
        return redirect(url_for('index'))

    location = request.form.get('location', '').strip()
    area = request.form.get('area', '').strip()
    shelf = request.form.get('shelf', '').strip()
    loc = request.form.get('loc', '').strip()

    # Validate: must be exactly 7 digits
    if not re.match(r'^\d{7}$', location):
        flash('Location must be exactly 7 digits.', 'danger')
        return redirect(url_for('admin_locations'))

    conn = None
    try:
        conn = db_manager.get_connection()
        with conn.cursor() as cursor:
            # Check for duplicate
            cursor.execute('SELECT COUNT(*) FROM pcb_inventory."tblLoc" WHERE location = %s', (location,))
            if cursor.fetchone()[0] > 0:
                flash(f'Location {location} already exists.', 'danger')
                return redirect(url_for('admin_locations'))

            cursor.execute("""
                INSERT INTO pcb_inventory."tblLoc" (area, shelf, loc, location)
                VALUES (%s, %s, %s, %s)
            """, (
                int(area) if area else None,
                shelf or None,
                loc or None,
                location
            ))
            conn.commit()
            log_user_activity('CREATE_LOCATION', f"Created location {location}", f"Area: {area}, Shelf: {shelf}, Loc: {loc}")
            flash(f'Location {location} created successfully.', 'success')
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error creating location: {e}")
        flash(f'Error creating location: {e}', 'danger')
    finally:
        if conn:
            db_manager.return_connection(conn)

    return redirect(url_for('admin_locations'))


@app.route('/admin/locations/delete/<int:location_id>', methods=['POST'])
@require_auth
def admin_delete_location(location_id):
    """Delete a location. Admin + Theresa only."""
    if not can_manage_parts():
        flash('Access denied.', 'danger')
        return redirect(url_for('index'))

    conn = None
    try:
        conn = db_manager.get_connection()
        with conn.cursor() as cursor:
            cursor.execute('SELECT location FROM pcb_inventory."tblLoc" WHERE id = %s', (location_id,))
            row = cursor.fetchone()
            if not row:
                flash('Location not found.', 'danger')
                return redirect(url_for('admin_locations'))

            loc_name = row[0]
            cursor.execute('DELETE FROM pcb_inventory."tblLoc" WHERE id = %s', (location_id,))
            conn.commit()
            log_user_activity('DELETE_LOCATION', f"Deleted location {loc_name}", f"ID: {location_id}")
            flash(f'Location {loc_name} deleted.', 'success')
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error deleting location: {e}")
        flash(f'Error deleting location: {e}', 'danger')
    finally:
        if conn:
            db_manager.return_connection(conn)

    return redirect(url_for('admin_locations'))


# ──────────────────────────────────────────────
# ACI Number Creator
# ──────────────────────────────────────────────

@app.route('/aci-numbers')
@require_auth
def aci_numbers():
    """ACI Number Creator page - create consecutive ACI part numbers for non-BOM parts."""
    if not can_manage_parts():
        flash('Access denied. You do not have permission to access ACI Numbers.', 'danger')
        return redirect(url_for('index'))
    response = make_response(render_template('aci_numbers.html'))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return response


@app.route('/api/aci-numbers/next', methods=['GET'])
@require_auth
def api_aci_next_number():
    if not can_manage_parts():
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    """Get the next available ACI number by scanning tblPN_List and tblACI_PartNumbers."""
    conn = None
    try:
        conn = db_manager.get_connection()
        with conn.cursor() as cursor:
            # Find the highest ACI-XXXXX number in the 5-digit sequence (10000-99999)
            cursor.execute("""
                SELECT MAX(num) FROM (
                    SELECT CAST(SUBSTRING(item FROM 5) AS INTEGER) as num
                    FROM pcb_inventory."tblPN_List"
                    WHERE item ~ '^ACI-[0-9]{5}$'
                    UNION ALL
                    SELECT CAST(SUBSTRING(aci_pn FROM 5) AS INTEGER) as num
                    FROM pcb_inventory."tblACI_PartNumbers"
                    WHERE aci_pn ~ '^ACI-[0-9]{5}$'
                ) sub
            """)
            row = cursor.fetchone()
            max_num = row[0] if row and row[0] else 10000
            next_num = max_num + 1
            return jsonify({'success': True, 'next_number': next_num, 'next_aci_pn': f'ACI-{next_num}'})
    except Exception as e:
        logger.error(f"Error getting next ACI number: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)


@app.route('/api/aci-numbers/create', methods=['POST'])
@require_auth
def api_aci_create():
    """Create one or more ACI part numbers. Expects JSON array of parts."""
    if not can_manage_parts():
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    conn = None
    try:
        data = request.get_json()
        if not data or 'parts' not in data:
            return jsonify({'success': False, 'error': 'No parts provided'}), 400

        parts = data['parts']
        if not parts or len(parts) == 0:
            return jsonify({'success': False, 'error': 'No parts provided'}), 400

        if len(parts) > 100:
            return jsonify({'success': False, 'error': 'Maximum 100 parts per batch'}), 400

        conn = db_manager.get_connection()
        created = []
        errors = []

        with conn.cursor() as cursor:
            # Lock to prevent race conditions on concurrent creates
            cursor.execute("LOCK TABLE pcb_inventory.\"tblACI_PartNumbers\" IN EXCLUSIVE MODE")

            # Get the current max ACI number in the 5-digit sequence
            cursor.execute("""
                SELECT MAX(num) FROM (
                    SELECT CAST(SUBSTRING(item FROM 5) AS INTEGER) as num
                    FROM pcb_inventory."tblPN_List"
                    WHERE item ~ '^ACI-[0-9]{5}$'
                    UNION ALL
                    SELECT CAST(SUBSTRING(aci_pn FROM 5) AS INTEGER) as num
                    FROM pcb_inventory."tblACI_PartNumbers"
                    WHERE aci_pn ~ '^ACI-[0-9]{5}$'
                ) sub
            """)
            row = cursor.fetchone()
            current_max = row[0] if row and row[0] else 10000

            username = session.get('username', 'unknown')

            for part in parts:
                manufacturer = (part.get('manufacturer') or '').strip()
                mpn = (part.get('mpn') or '').strip()
                description = (part.get('description') or '').strip()
                comment = (part.get('comment') or '').strip()
                loaded = (part.get('loaded') or 'N').strip().upper()
                if loaded not in ('Y', 'N'):
                    loaded = 'N'

                if not manufacturer and not mpn and not description:
                    errors.append('Skipped empty row')
                    continue

                current_max += 1
                aci_pn = f'ACI-{current_max}'

                # Insert into tblACI_PartNumbers
                cursor.execute("""
                    INSERT INTO pcb_inventory."tblACI_PartNumbers"
                    (aci_pn, manufacturer, mpn, description, comment, loaded, created_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (aci_pn, manufacturer, mpn, description, comment, loaded, username))

                # Also insert into tblPN_List so it shows up in inventory lookups
                cursor.execute("""
                    INSERT INTO pcb_inventory."tblPN_List" (item, "DESC")
                    VALUES (%s, %s)
                    ON CONFLICT DO NOTHING
                """, (aci_pn, description))

                created.append({
                    'aci_pn': aci_pn,
                    'manufacturer': manufacturer,
                    'mpn': mpn,
                    'description': description,
                    'loaded': loaded
                })

        conn.commit()

        if created:
            log_user_activity(
                'ACI_NUMBER_CREATE',
                f"Created {len(created)} ACI number(s): {created[0]['aci_pn']}" +
                (f" - {created[-1]['aci_pn']}" if len(created) > 1 else ''),
                json.dumps(created)
            )

        return jsonify({
            'success': True,
            'created': created,
            'count': len(created),
            'errors': errors
        })

    except psycopg2.errors.UniqueViolation:
        if conn:
            conn.rollback()
        return jsonify({'success': False, 'error': 'Duplicate ACI number detected. Please try again.'}), 409
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error creating ACI numbers: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)


@app.route('/api/aci-numbers/history', methods=['GET'])
@require_auth
def api_aci_history():
    """Get recently created ACI numbers."""
    if not can_manage_parts():
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    conn = None
    try:
        conn = db_manager.get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            cursor.execute("""
                SELECT aci_pn, manufacturer, mpn, description, comment, created_by,
                       TO_CHAR(created_at, 'MM/DD/YYYY HH12:MI AM') as created_at_fmt
                FROM pcb_inventory."tblACI_PartNumbers"
                ORDER BY id DESC
                LIMIT 100
            """)
            rows = cursor.fetchall()
            return jsonify({'success': True, 'history': rows})
    except Exception as e:
        logger.error(f"Error fetching ACI history: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if conn:
            db_manager.return_connection(conn)


if __name__ == '__main__':
    # Test database connection on startup
    try:
        test_inventory = db_manager.get_current_inventory()
        logger.info(f"Database connection successful. Found {len(test_inventory)} inventory items.")
    except Exception as e:
        logger.error(f"Database connection failed: {e}")
        print("Database connection failed. Check if PostgreSQL container is running.")
    
    # Run the application
    app.run(debug=False, host='0.0.0.0', port=5000)